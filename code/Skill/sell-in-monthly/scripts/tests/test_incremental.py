from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPT_DIR))

from extraction import OUTPUT_COLUMNS  # noqa: E402
from incremental import (  # noqa: E402
    build_incremental_plan,
    commit_incremental_plan,
)


SOURCE_ROW = [
    "MT",
    "MT",
    date(2026, 7, 10),
    "KH-001",
    "KHACH HANG",
    "130100001",
    "SAN PHAM",
    10,
    100,
    1000,
    "Don hang ban",
    "",
]


def save_source(path: Path, rows: list[list]) -> None:
    workbook = Workbook()
    ws = workbook.active
    for _ in range(4):
        ws.append([])
    ws.append(OUTPUT_COLUMNS[:12])
    for row in rows:
        ws.append(row)
    workbook.save(path)
    workbook.close()


def save_output(path: Path, rows: list[list]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Sell in"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row + [7, 2026])
    workbook.save(path)
    workbook.close()


class IncrementalPlanTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.source_dir = self.root / "Data" / "Data ERP"
        self.output_dir = (
            self.root / "Data" / "out put" / "Sell in hang  thang"
        )
        self.state_file = (
            self.root
            / "Data"
            / "Logs"
            / "Tach data logs"
            / "incremental_state.json"
        )
        self.source_dir.mkdir(parents=True)
        self.output_dir.mkdir(parents=True)
        self.source_file = (
            self.source_dir / "BCDonHangBanTrongKyNPP_Vikoda_T7_2026.xlsx"
        )
        self.output_file = self.output_dir / "Sell in T07_2026.xlsx"
        save_source(self.source_file, [SOURCE_ROW])
        save_output(self.output_file, [SOURCE_ROW])

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def baseline_and_commit(self) -> dict:
        plan = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(plan["rebuild_periods"], [])
        self.assertEqual(plan["skipped_periods"], ["2026-07"])
        self.assertIn(
            "baseline_source_and_output_match",
            plan["periods"][0]["reasons"],
        )
        commit_incremental_plan(plan, self.state_file)
        return plan

    def test_baseline_then_unchanged_skips_without_rescan(self) -> None:
        self.baseline_and_commit()
        second = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(second["rebuild_periods"], [])
        self.assertEqual(second["periods"][0]["reasons"], ["unchanged"])
        source = next(iter(second["source_files"].values()))
        self.assertTrue(source["reused_fingerprint"])
        self.assertTrue(second["periods"][0]["output"]["reused_fingerprint"])

    def test_resaved_source_with_same_dates_and_counts_skips(self) -> None:
        self.baseline_and_commit()
        workbook = load_workbook(self.source_file)
        workbook.active.title = "Du lieu"
        workbook.save(self.source_file)
        workbook.close()

        plan = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(plan["rebuild_periods"], [])
        self.assertIn(
            "source_resaved_no_invoice_date_change",
            plan["periods"][0]["reasons"],
        )

    def test_new_invoice_date_rebuilds_period(self) -> None:
        self.baseline_and_commit()
        new_row = list(SOURCE_ROW)
        new_row[2] = date(2026, 7, 11)
        save_source(self.source_file, [SOURCE_ROW, new_row])

        plan = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(plan["rebuild_periods"], ["2026-07"])
        self.assertTrue(
            any(
                reason.startswith("new_invoice_dates:2026-07-11")
                for reason in plan["periods"][0]["reasons"]
            )
        )

    def test_changed_count_on_existing_date_rebuilds_period(self) -> None:
        self.baseline_and_commit()
        save_source(self.source_file, [SOURCE_ROW, SOURCE_ROW])

        plan = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(plan["rebuild_periods"], ["2026-07"])
        self.assertTrue(
            any(
                reason.startswith("invoice_date_counts_changed:2026-07-10")
                for reason in plan["periods"][0]["reasons"]
            )
        )

    def test_missing_output_and_force_period_rebuild(self) -> None:
        self.baseline_and_commit()
        self.output_file.unlink()
        missing = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(missing["rebuild_periods"], ["2026-07"])
        self.assertIn("output_missing", missing["periods"][0]["reasons"])

        save_output(self.output_file, [SOURCE_ROW])
        forced = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
            force_periods=["2026-07"],
        )
        self.assertEqual(forced["rebuild_periods"], ["2026-07"])
        self.assertIn("force_period", forced["periods"][0]["reasons"])

    def test_changed_output_and_added_source_file_rebuild(self) -> None:
        self.baseline_and_commit()
        workbook = load_workbook(self.output_file)
        workbook.active["E2"] = "TEN BI SUA"
        workbook.save(self.output_file)
        workbook.close()

        changed_output = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(changed_output["rebuild_periods"], ["2026-07"])
        self.assertIn(
            "output_changed_since_last_verified_run",
            changed_output["periods"][0]["reasons"],
        )

        save_output(self.output_file, [SOURCE_ROW])
        added_source = (
            self.source_dir
            / "BCDonHangBanTrongKyNPP_VKD_T7_2026.xlsx"
        )
        save_source(added_source, [])
        changed_membership = build_incremental_plan(
            self.source_dir,
            self.output_dir,
            self.state_file,
        )
        self.assertEqual(changed_membership["rebuild_periods"], ["2026-07"])
        self.assertTrue(
            any(
                reason.startswith("source_files_added:")
                for reason in changed_membership["periods"][0]["reasons"]
            )
        )


if __name__ == "__main__":
    unittest.main()
