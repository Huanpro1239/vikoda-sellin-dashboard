from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPT_DIR))

from master_data import (  # noqa: E402
    CANDIDATE_COLUMNS,
    CUSTOMER_MASTER_COLUMNS,
    analyze_master_data,
    apply_approved_customers_portable,
    build_approval_plan,
    verify_master_data_artifacts,
    write_review_workbook_portable,
)


SELL_IN_COLUMNS = [
    "Vung",
    "KhuVuc",
    "NgayHoaDon",
    "MaKhachHangMoi",
    "TenKhachHang",
    "MaSanPhamMoi",
    "TenSanPham",
    "SoLuong",
    "DonGia",
    "ThanhTien",
    "LoaiDonHang",
    "GhiChu",
    "Thang",
    "Nam",
]


def save_rows(path: Path, sheet_name: str, rows: list[list]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    workbook.save(path)
    workbook.close()


class MasterDataTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.customer_master = (
            self.root
            / "Data"
            / "Danh muc KH"
            / "Thong tin khach hang.xlsx"
        )
        self.product_master = (
            self.root
            / "Data"
            / "Danh muc SP"
            / "Danh Muc San Pham.xlsx"
        )
        self.output_dir = (
            self.root
            / "Data"
            / "out put"
            / "Sell in hang  thang"
        )
        self.source_dir = self.root / "Data" / "Data ERP"
        self.candidate_dir = (
            self.root
            / "Data"
            / "Work"
            / "sell_in"
            / "new_customers"
        )

        save_rows(
            self.customer_master,
            "DM KH",
            [
                CUSTOMER_MASTER_COLUMNS + [None],
                [
                    "KH-001",
                    "KHACH CU",
                    "KHACH CU",
                    "DIA CHI CU",
                    "GT",
                    "Other",
                    None,
                    "Miền Trung",
                    "Nam Trung Bộ",
                    "Khánh Hòa",
                    "NHA TRANG",
                    None,
                    None,
                    None,
                ],
            ],
        )
        save_rows(
            self.product_master,
            "DanhMucSanPham",
            [
                [
                    "MaSanPhamMoi_Vikoda",
                    "MaSanPhamMoi_VKD",
                    "TenSanPham",
                ],
                ["130100001", "230100001", "SAN PHAM CU"],
            ],
        )
        save_rows(
            self.output_dir / "Sell in T08_2026.xlsx",
            "Sell in",
            [
                SELL_IN_COLUMNS,
                [
                    "MT",
                    "MT",
                    date(2026, 8, 12),
                    "KH-NEW",
                    "KHACH HANG MOI",
                    "130100999",
                    "SAN PHAM MOI",
                    10,
                    100,
                    1000,
                    "Đơn hàng bán",
                    "",
                    8,
                    2026,
                ],
                [
                    "MT",
                    "MT",
                    date(2026, 8, 13),
                    "KH-NEW",
                    "KHACH HANG MOI",
                    "130100999",
                    "SAN PHAM MOI",
                    5,
                    100,
                    500,
                    "Đơn hàng bán",
                    "",
                    8,
                    2026,
                ],
            ],
        )
        erp_headers = [
            "MaKhachHangMoi",
            "TenKhachHang",
            "NgayHoaDon",
            "DiaChiGiaoHangImport",
            "DiaChiGiaoHang",
            "KenhBanHang",
            "TinhThanh",
            "QuanHuyen",
        ]
        erp_path = (
            self.source_dir
            / "BCDonHangBanTrongKyNPP_VKD_T8_2026.xlsm"
        )
        erp_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        ws = workbook.active
        for _ in range(8):
            ws.append([])
        ws.append(erp_headers)
        ws.append(
            [
                "KH-NEW",
                "KHACH HANG MOI",
                date(2026, 8, 13),
                "12 ĐƯỜNG MỚI",
                "",
                "GT",
                "Khánh Hòa",
                "NHA TRANG",
            ]
        )
        workbook.save(erp_path)
        workbook.close()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_analyze_new_customer_and_missing_product(self) -> None:
        analysis = analyze_master_data(
            self.output_dir,
            self.source_dir,
            self.customer_master,
            self.product_master,
            self.candidate_dir,
        )
        self.assertEqual(analysis["latest_period"]["label"], "T08_2026")
        self.assertEqual(analysis["candidate_count"], 1)
        candidate = analysis["candidate_rows"][0]
        self.assertEqual(candidate["MaKhachHangMoi"], "KH-NEW")
        self.assertEqual(candidate["DiaChi"], "12 ĐƯỜNG MỚI")
        self.assertEqual(candidate["MIEN"], "Miền Trung")
        self.assertEqual(candidate["VUNG"], "Nam Trung Bộ")
        self.assertEqual(candidate["Loaikhachhang"], "")
        self.assertEqual(candidate["Hethong MT"], "")
        self.assertEqual(candidate["CODE"], "")
        self.assertEqual(candidate["TENKH"], "")
        self.assertEqual(candidate["TrangThaiDuyet"], "CHO DUYET")
        self.assertEqual(analysis["missing_product_count"], 1)
        self.assertEqual(
            analysis["missing_products"][0]["MaSanPhamMoi"],
            "130100999",
        )

    def test_approved_customer_is_appended_once_with_backup(self) -> None:
        candidate_path = (
            self.candidate_dir / "Khach hang moi T08_2026.xlsx"
        )
        approved = {
            column: ""
            for column in CANDIDATE_COLUMNS
        }
        approved.update(
            {
                "MaKhachHangMoi": "KH-APPROVED",
                "TenKhachHang": "KHACH DA DUYET",
                "TenKhachHangdaydu": "KHACH DA DUYET",
                "TrangThaiDuyet": "DUYET",
            }
        )
        save_rows(
            candidate_path,
            "Khach hang moi",
            [
                CANDIDATE_COLUMNS,
                [approved[column] for column in CANDIDATE_COLUMNS],
            ],
        )

        plan = build_approval_plan(
            self.customer_master,
            self.candidate_dir,
        )
        self.assertEqual(plan["approved_count"], 1)
        self.assertEqual(plan["errors"], [])
        backup_dir = self.root / "Data" / "Logs" / "Danh muc KH backups"
        report = apply_approved_customers_portable(plan, backup_dir)
        self.assertTrue(report["changed"])
        self.assertEqual(report["appended_codes"], ["KH-APPROVED"])
        self.assertTrue(Path(report["backup_file"]).exists())

        second_plan = build_approval_plan(
            self.customer_master,
            self.candidate_dir,
        )
        self.assertEqual(second_plan["approved_count"], 0)
        self.assertEqual(len(second_plan["already_exists"]), 1)

    def test_portable_review_workbook_passes_verification(self) -> None:
        analysis = analyze_master_data(
            self.output_dir,
            self.source_dir,
            self.customer_master,
            self.product_master,
            self.candidate_dir,
        )
        output_path = write_review_workbook_portable(analysis)
        self.assertTrue(output_path.exists())
        report = verify_master_data_artifacts(
            analysis,
            {"appended_codes": []},
        )
        self.assertTrue(report["ok"], report["problems"])
        workbook = load_workbook(output_path, read_only=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["Khach hang moi", "Ma SP chua co", "Huong dan"],
            )
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
