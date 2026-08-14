"""Gộp toàn bộ workbook Sell In hàng tháng thành một file CSV phẳng cho Looker.

Looker Studio KHÔNG đọc được .xlsx nằm trên Google Drive: nguồn dùng được chỉ có
Google Sheets, CSV (File Upload) hoặc BigQuery. Script này đọc mọi file
`Sell in T<MM>_<YYYY>.xlsx` trong thư mục output, nối lại thành một bảng duy nhất
rồi ghi ra CSV để:

  * import vào Google Sheets (Tệp > Nhập > Thay thế trang tính hiện tại), hoặc
  * nạp trực tiếp bằng connector "File Upload" của Looker Studio.

Khác biệt so với workbook tháng:
  * Thêm cột `KyBaoCao` dạng YYYY-MM để Looker có sẵn một chiều thời gian tháng.
  * `NgayHoaDon` chuẩn hoá thành YYYY-MM-DD (Looker tự nhận là kiểu Date).
  * Số ghi dạng thuần, không tách nghìn, dấu thập phân là dấu chấm.
  * CSV ghi UTF-8 KHÔNG BOM: Looker Studio và Google Sheets đều đọc đúng tiếng
    Việt; nếu để BOM thì tên cột đầu tiên trong Looker bị dính ký tự lạ.

Ví dụ:
    python build_looker_dataset.py \
        --output-dir "Data/out put/Sell in hang  thang" \
        --csv-file "Data/Work/sell_in/looker/Sell in tong hop.csv" \
        --report-file "Data/Work/sell_in/verification/looker_report.json"
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from extraction import OUTPUT_COLUMNS

SHEET_NAME = "Sell in"
WORKBOOK_PATTERN = re.compile(
    r"^Sell in T(?P<month>\d{2})_(?P<year>\d{4})\.xlsx$",
    re.IGNORECASE,
)
# Cột thêm cho Looker, đứng đầu để dễ chọn làm chiều thời gian.
LOOKER_COLUMNS = ["KyBaoCao"] + OUTPUT_COLUMNS
NUMERIC_COLUMNS = ("SoLuong", "DonGia", "ThanhTien")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Thư mục chứa các file Sell in T<MM>_<YYYY>.xlsx.",
    )
    parser.add_argument(
        "--csv-file",
        required=True,
        help="Đường dẫn CSV gộp cần ghi ra.",
    )
    parser.add_argument(
        "--report-file",
        default="",
        help="Tuỳ chọn: ghi báo cáo đối soát dạng JSON.",
    )
    return parser.parse_args()


def discover_workbooks(output_dir: Path) -> list[tuple[int, int, Path]]:
    """Trả về [(năm, tháng, đường dẫn)] đã sắp xếp theo kỳ tăng dần."""
    found: list[tuple[int, int, Path]] = []
    for path in output_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        match = WORKBOOK_PATTERN.match(path.name)
        if match is None:
            continue
        found.append(
            (int(match.group("year")), int(match.group("month")), path)
        )
    found.sort(key=lambda item: (item[0], item[1]))
    return found


def format_cell(column_name: str, value: object) -> str:
    if value is None:
        return ""
    if column_name == "NgayHoaDon":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip()
    if column_name in NUMERIC_COLUMNS:
        if isinstance(value, bool):
            return str(int(value))
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def read_workbook_rows(
    path: Path,
    year: int,
    month: int,
) -> tuple[list[list[str]], dict]:
    """Đọc sheet 'Sell in', bỏ hàng tiêu đề, trả về hàng CSV và tóm tắt kỳ."""
    workbook = openpyxl.load_workbook(
        path, read_only=True, data_only=True
    )
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"{path.name}: khong tim thay sheet '{SHEET_NAME}'."
            )
        worksheet = workbook[SHEET_NAME]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration:
            raise ValueError(f"{path.name}: sheet trong, khong co tieu de.")

        header_values = [
            str(cell).strip() if cell is not None else ""
            for cell in header[: len(OUTPUT_COLUMNS)]
        ]
        if header_values != OUTPUT_COLUMNS:
            raise ValueError(
                f"{path.name}: tieu de khong dung chuan.\n"
                f"  Mong doi: {OUTPUT_COLUMNS}\n"
                f"  Thuc te : {header_values}"
            )

        period_key = f"{year:04d}-{month:02d}"
        rows: list[list[str]] = []
        amount_total = 0.0
        quantity_total = 0.0
        amount_index = OUTPUT_COLUMNS.index("ThanhTien")
        quantity_index = OUTPUT_COLUMNS.index("SoLuong")

        for raw_row in iterator:
            padded = list(raw_row[: len(OUTPUT_COLUMNS)])
            padded += [None] * (len(OUTPUT_COLUMNS) - len(padded))
            if all(cell is None or str(cell).strip() == "" for cell in padded):
                continue

            if isinstance(padded[amount_index], (int, float)) and not isinstance(
                padded[amount_index], bool
            ):
                amount_total += float(padded[amount_index])
            if isinstance(padded[quantity_index], (int, float)) and not isinstance(
                padded[quantity_index], bool
            ):
                quantity_total += float(padded[quantity_index])

            rows.append(
                [period_key]
                + [
                    format_cell(name, cell)
                    for name, cell in zip(OUTPUT_COLUMNS, padded)
                ]
            )

        summary = {
            "period": period_key,
            "file": path.name,
            "row_count": len(rows),
            "quantity_total": quantity_total,
            "amount_total": amount_total,
        }
        return rows, summary
    finally:
        workbook.close()


def build_looker_csv(
    output_dir: Path,
    csv_file: Path,
    report_file: Path | None = None,
    verbose: bool = True,
) -> dict:
    """Gộp mọi workbook trong `output_dir` thành một CSV, trả về báo cáo đối soát.

    Tách riêng khỏi `main()` để luồng chuyển giao gọi trực tiếp bằng import:
    `portable_sell_in.py` được đóng gói thành EXE nên không thể chạy script con.
    """
    output_dir = Path(output_dir)
    csv_file = Path(csv_file)

    if not output_dir.is_dir():
        raise FileNotFoundError(f"Khong tim thay thu muc output: {output_dir}")

    workbooks = discover_workbooks(output_dir)
    if not workbooks:
        raise FileNotFoundError(
            f"Khong co file 'Sell in T<MM>_<YYYY>.xlsx' trong {output_dir}"
        )

    csv_file.parent.mkdir(parents=True, exist_ok=True)

    periods: list[dict] = []
    total_rows = 0
    total_amount = 0.0
    total_quantity = 0.0

    # utf-8 khong BOM + newline='' de csv module tu quan ly ky tu ket dong.
    with csv_file.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\r\n")
        writer.writerow(LOOKER_COLUMNS)
        for year, month, path in workbooks:
            rows, summary = read_workbook_rows(path, year, month)
            writer.writerows(rows)
            periods.append(summary)
            total_rows += summary["row_count"]
            total_amount += summary["amount_total"]
            total_quantity += summary["quantity_total"]
            if verbose:
                print(
                    f"  - {summary['period']}: {summary['row_count']:,} dong"
                    f" | ThanhTien {summary['amount_total']:,.0f}"
                )

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "output_dir": str(output_dir),
        "csv_file": str(csv_file),
        "columns": LOOKER_COLUMNS,
        "encoding": "utf-8 (no BOM)",
        "workbook_count": len(workbooks),
        "row_count": total_rows,
        "quantity_total": total_quantity,
        "amount_total": total_amount,
        "periods": periods,
    }

    report["size_bytes"] = csv_file.stat().st_size

    if report_file is not None:
        report_file = Path(report_file)
        report_file.parent.mkdir(parents=True, exist_ok=True)
        report_file.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    if verbose:
        size_mb = report["size_bytes"] / (1024 * 1024)
        print(
            f"Gop {len(workbooks)} thang -> {total_rows:,} dong"
            f" ({size_mb:.1f} MB): {csv_file}"
        )
    return report


def main() -> int:
    args = parse_args()
    try:
        build_looker_csv(
            output_dir=Path(args.output_dir),
            csv_file=Path(args.csv_file),
            report_file=Path(args.report_file) if args.report_file else None,
        )
    except FileNotFoundError as error:
        print(error)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
