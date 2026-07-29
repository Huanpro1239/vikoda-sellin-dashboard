from __future__ import annotations

import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from extraction import (
    FILE_PATTERN,
    OUTPUT_COLUMNS,
    clean_id,
    iso_invoice_date,
    locate_header,
    normalize_company,
    parse_invoice_date,
)


STATE_SCHEMA_VERSION = 1
OUTPUT_FILE_PATTERN = "Sell in T{month:02d}_{year}.xlsx"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def parse_source_file(path: Path) -> dict[str, Any] | None:
    if path.name.startswith("~$") or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    match = FILE_PATTERN.search(path.name)
    if not match:
        return None
    month = int(match.group("month"))
    if not 1 <= month <= 12:
        return None
    year = int(match.group("year"))
    return {
        "file": path.name,
        "path": str(path.resolve()),
        "company": normalize_company(match.group("company")),
        "month": month,
        "year": year,
        "period": f"{year}-{month:02d}",
    }


def validate_period(period: str) -> str:
    parts = period.split("-")
    if len(parts) != 2 or not all(part.isdigit() for part in parts):
        raise ValueError(f"Tháng bắt buộc theo dạng YYYY-MM: {period}")
    year, month = (int(part) for part in parts)
    if year < 1900 or not 1 <= month <= 12:
        raise ValueError(f"Tháng không hợp lệ: {period}")
    return f"{year:04d}-{month:02d}"


def scan_source_file(path: Path, metadata: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        ws = workbook.worksheets[0]
        header_row, header_map = locate_header(ws)
        date_counts: Counter[str] = Counter()
        eligible_rows = 0
        invalid_invoice_dates = 0

        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            customer_code = clean_id(values[header_map["MaKhachHangMoi"]])
            if (
                metadata["company"] == "Vikoda"
                and customer_code.upper() == "VKD3"
            ):
                continue
            product_code = clean_id(values[header_map["MaSanPhamMoi"]])
            if not product_code or product_code[0] not in {"1", "2"}:
                continue

            # Khóa `date_counts` phải là chuỗi ISO để so khớp với output và ghi
            # được vào incremental_state.json.
            invoice_date = iso_invoice_date(
                parse_invoice_date(
                    values[header_map["NgayHoaDon"]],
                    workbook.epoch,
                )
            )
            eligible_rows += 1
            if invoice_date is None:
                invalid_invoice_dates += 1
            else:
                date_counts[invoice_date] += 1

        return {
            **metadata,
            "sheet": ws.title,
            "header_row": header_row,
            "eligible_rows": eligible_rows,
            "invalid_invoice_dates": invalid_invoice_dates,
            "date_counts": dict(sorted(date_counts.items())),
        }
    finally:
        workbook.close()


def scan_output_file(path: Path, year: int, month: int) -> dict[str, Any]:
    result: dict[str, Any] = {
        "file": path.name,
        "path": str(path.resolve()),
        "exists": path.exists(),
        "valid": False,
        "rows": 0,
        "date_counts": {},
        "problems": [],
    }
    if not path.exists():
        result["problems"].append("output_missing")
        return result

    result["size"] = path.stat().st_size
    result["sha256"] = sha256_file(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        result["problems"].append(f"cannot_open:{exc}")
        return result

    try:
        if len(workbook.worksheets) != 1:
            result["problems"].append("sheet_count")
        ws = workbook.worksheets[0]
        if ws.title != "Sell in":
            result["problems"].append("sheet_name")
        headers = [
            cell.value
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        if headers != OUTPUT_COLUMNS:
            result["problems"].append("headers")

        date_counts: Counter[str] = Counter()
        invalid_dates = 0
        invalid_period = 0
        rows = 0
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            rows += 1
            invoice_date = values[2] if len(values) > 2 else None
            if isinstance(invoice_date, datetime):
                invoice_key = invoice_date.date().isoformat()
            elif isinstance(invoice_date, date):
                invoice_key = invoice_date.isoformat()
            else:
                invoice_key = None
            if invoice_key is None:
                invalid_dates += 1
            else:
                date_counts[invoice_key] += 1
            if (
                len(values) < len(OUTPUT_COLUMNS)
                or values[12] != month
                or values[13] != year
            ):
                invalid_period += 1

        result.update(
            {
                "sheet": ws.title,
                "rows": rows,
                "date_counts": dict(sorted(date_counts.items())),
                "invalid_dates": invalid_dates,
                "invalid_period": invalid_period,
            }
        )
        if invalid_dates:
            result["problems"].append("invalid_dates")
        if invalid_period:
            result["problems"].append("invalid_period")
        result["valid"] = not result["problems"]
        return result
    finally:
        workbook.close()


def load_state(state_file: Path) -> dict[str, Any]:
    if not state_file.exists():
        return {
            "schema_version": STATE_SCHEMA_VERSION,
            "source_files": {},
            "outputs": {},
        }
    try:
        state = json.loads(state_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Không đọc được trạng thái tăng dần: {state_file}") from exc
    if state.get("schema_version") != STATE_SCHEMA_VERSION:
        raise ValueError(
            "Phiên bản trạng thái tăng dần không được hỗ trợ. "
            f"File: {state_file}"
        )
    return state


def _reuse_source_fingerprint(
    previous: dict[str, Any] | None,
    metadata: dict[str, Any],
    source_hash: str,
) -> dict[str, Any] | None:
    if (
        not previous
        or previous.get("sha256") != source_hash
        or "date_counts" not in previous
        or previous.get("period") != metadata["period"]
    ):
        return None
    reused = dict(previous)
    reused.update(metadata)
    reused["reused_fingerprint"] = True
    return reused


def _reuse_output_fingerprint(
    previous: dict[str, Any] | None,
    path: Path,
    output_hash: str,
) -> dict[str, Any] | None:
    if (
        not previous
        or previous.get("sha256") != output_hash
        or not previous.get("verified")
        or "date_counts" not in previous
    ):
        return None
    reused = dict(previous)
    reused.update(
        {
            "file": path.name,
            "path": str(path.resolve()),
            "exists": True,
            "valid": True,
            "reused_fingerprint": True,
        }
    )
    return reused


def _aggregate_source_period(
    sources: list[dict[str, Any]],
) -> tuple[dict[str, int], int, int]:
    date_counts: Counter[str] = Counter()
    rows = 0
    invalid_dates = 0
    for source in sources:
        date_counts.update(source.get("date_counts", {}))
        rows += int(source.get("eligible_rows", 0))
        invalid_dates += int(source.get("invalid_invoice_dates", 0))
    return dict(sorted(date_counts.items())), rows, invalid_dates


def _date_difference_reasons(
    source_counts: dict[str, int],
    output_counts: dict[str, int],
) -> list[str]:
    missing_dates = sorted(set(source_counts) - set(output_counts))
    extra_dates = sorted(set(output_counts) - set(source_counts))
    changed_dates = sorted(
        invoice_date
        for invoice_date in set(source_counts) & set(output_counts)
        if source_counts[invoice_date] != output_counts[invoice_date]
    )
    reasons: list[str] = []
    if missing_dates:
        reasons.append(
            f"new_invoice_dates:{','.join(missing_dates[:5])}"
            + ("..." if len(missing_dates) > 5 else "")
        )
    if changed_dates:
        reasons.append(
            f"invoice_date_counts_changed:{','.join(changed_dates[:5])}"
            + ("..." if len(changed_dates) > 5 else "")
        )
    if extra_dates:
        reasons.append(
            f"output_has_extra_dates:{','.join(extra_dates[:5])}"
            + ("..." if len(extra_dates) > 5 else "")
        )
    return reasons


def build_incremental_plan(
    source_dir: Path,
    output_dir: Path,
    state_file: Path,
    force_periods: list[str] | None = None,
    force_all: bool = False,
) -> dict[str, Any]:
    source_dir = source_dir.resolve()
    output_dir = output_dir.resolve()
    state_file = state_file.resolve()
    forced = {
        validate_period(period)
        for period in (force_periods or [])
    }
    state = load_state(state_file)
    previous_sources = state.get("source_files", {})
    previous_outputs = state.get("outputs", {})

    current_sources: dict[str, dict[str, Any]] = {}
    ignored_files: list[str] = []
    for path in sorted(source_dir.iterdir(), key=lambda item: item.name.lower()):
        metadata = parse_source_file(path)
        if metadata is None:
            if (
                not path.name.startswith("~$")
                and path.suffix.lower() in {".xlsx", ".xlsm"}
            ):
                ignored_files.append(path.name)
            continue
        source_hash = sha256_file(path)
        source = _reuse_source_fingerprint(
            previous_sources.get(path.name),
            metadata,
            source_hash,
        )
        if source is None:
            source = scan_source_file(path, metadata)
            source["reused_fingerprint"] = False
        source.update(
            {
                "sha256": source_hash,
                "size": path.stat().st_size,
                "mtime_ns": path.stat().st_mtime_ns,
            }
        )
        current_sources[path.name] = source

    if not current_sources:
        raise RuntimeError("Không tìm thấy file ERP hợp lệ để lập kế hoạch.")

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for source in current_sources.values():
        grouped[source["period"]].append(source)

    previous_members: dict[str, set[str]] = defaultdict(set)
    for name, source in previous_sources.items():
        if source.get("period"):
            previous_members[source["period"]].add(name)

    problems: list[str] = []
    missing_periods = sorted(set(previous_members) - set(grouped))
    if missing_periods:
        problems.append(
            "Toàn bộ file nguồn của các tháng đã theo dõi bị mất: "
            + ", ".join(missing_periods)
        )
    unknown_forced = sorted(forced - set(grouped))
    if unknown_forced:
        problems.append(
            "Không có file nguồn cho ForcePeriod: " + ", ".join(unknown_forced)
        )

    periods: list[dict[str, Any]] = []
    for period, sources in sorted(grouped.items()):
        year, month = (int(part) for part in period.split("-"))
        source_counts, source_rows, invalid_source_dates = (
            _aggregate_source_period(sources)
        )
        if invalid_source_dates:
            problems.append(
                f"{period}: có {invalid_source_dates} dòng nguồn hợp lệ "
                "nhưng NgayHoaDon không đọc được."
            )

        output_path = output_dir / OUTPUT_FILE_PATTERN.format(
            month=month,
            year=year,
        )
        previous_output = previous_outputs.get(period)
        if output_path.exists():
            output_hash = sha256_file(output_path)
            output = _reuse_output_fingerprint(
                previous_output,
                output_path,
                output_hash,
            )
            if output is None:
                output = scan_output_file(output_path, year, month)
        else:
            output = scan_output_file(output_path, year, month)

        current_members = {source["file"] for source in sources}
        prior_members = previous_members.get(period, set())
        changed_sources = sorted(
            source["file"]
            for source in sources
            if not source.get("reused_fingerprint")
        )

        reasons: list[str] = []
        action = "SKIP"
        if force_all:
            reasons.append("force_all")
        elif period in forced:
            reasons.append("force_period")
        if not output.get("exists"):
            reasons.append("output_missing")
        elif not output.get("valid"):
            reasons.append(
                "output_invalid:" + ",".join(output.get("problems", []))
            )
        if prior_members and current_members != prior_members:
            added = sorted(current_members - prior_members)
            removed = sorted(prior_members - current_members)
            if added:
                reasons.append("source_files_added:" + ",".join(added))
            if removed:
                reasons.append("source_files_removed:" + ",".join(removed))
        if (
            previous_output
            and output.get("exists")
            and previous_output.get("sha256")
            and output.get("sha256") != previous_output.get("sha256")
        ):
            reasons.append("output_changed_since_last_verified_run")
        if output.get("valid") and source_counts != output.get("date_counts", {}):
            reasons.extend(
                _date_difference_reasons(
                    source_counts,
                    output.get("date_counts", {}),
                )
            )
            if not reasons:
                reasons.append("source_output_counts_changed")

        if reasons:
            action = "REBUILD"
        elif not previous_sources:
            reasons.append("baseline_source_and_output_match")
        elif changed_sources:
            reasons.append("source_resaved_no_invoice_date_change")
        else:
            reasons.append("unchanged")

        periods.append(
            {
                "period": period,
                "year": year,
                "month": month,
                "action": action,
                "reasons": reasons,
                "source_files": sorted(current_members),
                "changed_source_files": changed_sources,
                "source_rows": source_rows,
                "source_date_counts": source_counts,
                "output_file": str(output_path.resolve()),
                "output": output,
            }
        )

    rebuild_periods = [
        item["period"] for item in periods if item["action"] == "REBUILD"
    ]
    skipped_periods = [
        item["period"] for item in periods if item["action"] == "SKIP"
    ]
    return {
        "schema_version": STATE_SCHEMA_VERSION,
        "generated_at": utc_now(),
        "source_dir": str(source_dir),
        "output_dir": str(output_dir),
        "state_file": str(state_file),
        "force_all": force_all,
        "force_periods": sorted(forced),
        "periods": periods,
        "rebuild_periods": rebuild_periods,
        "skipped_periods": skipped_periods,
        "source_files": current_sources,
        "ignored_files": ignored_files,
        "problems": problems,
    }


def write_plan(plan: dict[str, Any], plan_file: Path) -> None:
    plan_file.parent.mkdir(parents=True, exist_ok=True)
    plan_file.write_text(
        json.dumps(plan, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def commit_incremental_plan(
    plan: dict[str, Any],
    state_file: Path | None = None,
) -> dict[str, Any]:
    if plan.get("problems"):
        raise RuntimeError(
            "Không thể ghi trạng thái vì kế hoạch còn lỗi: "
            + "; ".join(plan["problems"])
        )
    output_state: dict[str, dict[str, Any]] = {}
    for period_item in plan["periods"]:
        period = period_item["period"]
        if period_item["action"] == "REBUILD":
            output_path = Path(period_item["output_file"])
            output = scan_output_file(
                output_path,
                period_item["year"],
                period_item["month"],
            )
        else:
            output = dict(period_item["output"])
            output.pop("reused_fingerprint", None)
        if not output.get("valid"):
            raise RuntimeError(
                f"{period}: chưa thể ghi trạng thái vì output không hợp lệ: "
                + ", ".join(output.get("problems", []))
            )
        if output.get("date_counts", {}) != period_item["source_date_counts"]:
            raise RuntimeError(
                f"{period}: số dòng theo NgayHoaDon của output không khớp nguồn."
            )
        output["verified"] = True
        output["verified_at"] = utc_now()
        output_state[period] = output

    source_state: dict[str, dict[str, Any]] = {}
    for name, source in plan["source_files"].items():
        clean_source = dict(source)
        clean_source.pop("reused_fingerprint", None)
        source_state[name] = clean_source

    target = (
        state_file.resolve()
        if state_file is not None
        else Path(plan["state_file"]).resolve()
    )
    state = {
        "schema_version": STATE_SCHEMA_VERSION,
        "updated_at": utc_now(),
        "source_dir": plan["source_dir"],
        "output_dir": plan["output_dir"],
        "source_files": source_state,
        "outputs": output_state,
    }
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(target.suffix + ".tmp")
    temporary.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(target)
    return state
