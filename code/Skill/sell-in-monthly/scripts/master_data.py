from __future__ import annotations

import json
import math
import os
import re
import shutil
from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


CUSTOMER_MASTER_COLUMNS = [
    "MaKhachHangMoi",
    "TenKhachHang",
    "TenKhachHangdaydu",
    "DiaChi",
    "KenhBanHang",
    "Loaikhachhang",
    "Hethong MT",
    "MIEN",
    "VUNG",
    "TINHTHANH",
    "QUANHUYEN",
    "CODE",
    "TENKH",
]

CANDIDATE_COLUMNS = CUSTOMER_MASTER_COLUMNS + [
    "KyDuLieu",
    "FileNguonERP",
    "NgayHoaDonGanNhat",
    "SoDongSellIn",
    "TruongXungDot",
    "TrangThaiDuyet",
    "GhiChuDuyet",
]

PRODUCT_REVIEW_COLUMNS = [
    "MaSanPhamMoi",
    "TenSanPham",
    "SoDongSellIn",
    "KyDuLieu",
    "GhiChu",
]

APPROVAL_STATUSES = {
    "CHO DUYET",
    "DUYET",
    "TU CHOI",
    "CAN BO SUNG",
}

SELL_IN_PATTERN = re.compile(
    r"Sell in T(?P<month>\d{2})_(?P<year>\d{4})\.xlsx$",
    re.IGNORECASE,
)

ERP_PATTERN = re.compile(
    r"_(?P<company>Vikoda|Vkoda|VKD)_T(?P<month>\d{1,2})_(?P<year>\d{4})\.(?:xlsx|xlsm)$",
    re.IGNORECASE,
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).strip()


def normalize_key(value: Any) -> str:
    return clean_id(value).strip().upper()


def normalize_status(value: Any) -> str:
    status = re.sub(r"\s+", " ", clean_text(value)).upper()
    return status


def unique_nonblank(values: Iterable[Any]) -> list[str]:
    result: dict[str, str] = {}
    for value in values:
        text = clean_text(value)
        key = text.upper()
        if text and key not in result:
            result[key] = text
    return list(result.values())


def parse_excel_date(value: Any, epoch: datetime) -> date | None:
    parsed: date | datetime | None = None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value, epoch)
        except (TypeError, ValueError, OverflowError):
            parsed = None
    elif value is not None:
        text = str(value).strip()
        for pattern in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, pattern)
                break
            except ValueError:
                continue
    if parsed is None:
        return None
    return parsed.date() if isinstance(parsed, datetime) else parsed


def header_map(values: Iterable[Any]) -> dict[str, int]:
    return {
        clean_text(value): index
        for index, value in enumerate(values)
        if clean_text(value)
    }


def require_headers(
    mapping: dict[str, int],
    required: Iterable[str],
    source_name: str,
) -> None:
    missing = [column for column in required if column not in mapping]
    if missing:
        raise ValueError(
            f"{source_name} thiếu cột bắt buộc: {', '.join(missing)}"
        )


def locate_erp_header(ws) -> tuple[int, dict[str, int]]:
    required = {"MaKhachHangMoi", "TenKhachHang", "NgayHoaDon"}
    for row_number in range(1, min(ws.max_row, 40) + 1):
        mapping = header_map(cell.value for cell in ws[row_number])
        if required.issubset(mapping):
            return row_number, mapping
    raise ValueError(
        f"Không tìm thấy tiêu đề ERP hợp lệ trong 40 hàng đầu của sheet {ws.title}."
    )


def find_latest_sell_in(output_dir: Path) -> tuple[Path, int, int]:
    periods: list[tuple[int, int, Path]] = []
    for file_path in output_dir.glob("Sell in T??_????.xlsx"):
        match = SELL_IN_PATTERN.fullmatch(file_path.name)
        if not match:
            continue
        periods.append(
            (int(match.group("year")), int(match.group("month")), file_path)
        )
    if not periods:
        raise FileNotFoundError(
            f"Không tìm thấy file Sell In tháng tại {output_dir}"
        )
    year, month, file_path = max(periods)
    return file_path, month, year


def read_customer_master(customer_master: Path) -> dict[str, Any]:
    workbook = load_workbook(
        customer_master,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        ws = workbook.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(
                f"Danh mục khách hàng đang trống: {customer_master}"
            ) from exc
        mapping = header_map(headers)
        require_headers(mapping, CUSTOMER_MASTER_COLUMNS, customer_master.name)

        records: list[dict[str, str]] = []
        code_counter: Counter[str] = Counter()
        last_data_row = 1
        for row_number, values in enumerate(rows, start=2):
            code = clean_id(values[mapping["MaKhachHangMoi"]])
            if not code:
                continue
            record = {
                column: clean_text(values[mapping[column]])
                for column in CUSTOMER_MASTER_COLUMNS
            }
            record["MaKhachHangMoi"] = code
            records.append(record)
            code_counter[normalize_key(code)] += 1
            last_data_row = row_number

        return {
            "sheet_name": ws.title,
            "headers": headers,
            "records": records,
            "codes": set(code_counter),
            "duplicate_codes": sorted(
                code for code, count in code_counter.items() if count > 1
            ),
            "last_data_row": last_data_row,
            "max_column": ws.max_column,
        }
    finally:
        workbook.close()


def read_existing_candidate(candidate_path: Path) -> dict[str, dict[str, Any]]:
    if not candidate_path.exists():
        return {}
    workbook = load_workbook(
        candidate_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        ws = (
            workbook["Khach hang moi"]
            if "Khach hang moi" in workbook.sheetnames
            else workbook.worksheets[0]
        )
        rows = ws.iter_rows(values_only=True)
        try:
            mapping = header_map(next(rows))
        except StopIteration:
            return {}
        if "MaKhachHangMoi" not in mapping:
            return {}
        result: dict[str, dict[str, Any]] = {}
        for values in rows:
            code = clean_id(values[mapping["MaKhachHangMoi"]])
            if not code:
                continue
            row = {
                column: (
                    values[mapping[column]]
                    if column in mapping and mapping[column] < len(values)
                    else None
                )
                for column in CANDIDATE_COLUMNS
            }
            result[normalize_key(code)] = row
        return result
    finally:
        workbook.close()


def build_approval_plan(
    customer_master: Path,
    candidate_dir: Path,
) -> dict[str, Any]:
    master = read_customer_master(customer_master)
    existing_codes = set(master["codes"])
    approved_by_code: dict[str, dict[str, Any]] = {}
    source_by_code: dict[str, str] = {}
    already_exists: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []
    reviewed_files: list[str] = []

    for candidate_path in sorted(
        candidate_dir.glob("Khach hang moi T??_????.xlsx"),
        key=lambda path: path.name.lower(),
    ):
        if candidate_path.name.startswith("~$"):
            continue
        reviewed_files.append(str(candidate_path))
        workbook = load_workbook(
            candidate_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            ws = (
                workbook["Khach hang moi"]
                if "Khach hang moi" in workbook.sheetnames
                else workbook.worksheets[0]
            )
            rows = ws.iter_rows(values_only=True)
            try:
                mapping = header_map(next(rows))
            except StopIteration:
                continue
            require_headers(mapping, CANDIDATE_COLUMNS, candidate_path.name)

            for row_number, values in enumerate(rows, start=2):
                status = normalize_status(
                    values[mapping["TrangThaiDuyet"]]
                )
                if status != "DUYET":
                    continue
                code = clean_id(values[mapping["MaKhachHangMoi"]])
                name = clean_text(values[mapping["TenKhachHang"]])
                code_key = normalize_key(code)
                if not code:
                    errors.append(
                        {
                            "file": str(candidate_path),
                            "row": str(row_number),
                            "error": "Dòng DUYET thiếu MaKhachHangMoi.",
                        }
                    )
                    continue
                if not name:
                    errors.append(
                        {
                            "file": str(candidate_path),
                            "row": str(row_number),
                            "code": code,
                            "error": "Dòng DUYET thiếu TenKhachHang.",
                        }
                    )
                    continue
                if code_key in existing_codes:
                    already_exists.append(
                        {
                            "file": str(candidate_path),
                            "code": code,
                        }
                    )
                    continue

                row = {
                    column: clean_text(values[mapping[column]])
                    for column in CUSTOMER_MASTER_COLUMNS
                }
                row["MaKhachHangMoi"] = code
                if code_key in approved_by_code:
                    if approved_by_code[code_key] != row:
                        errors.append(
                            {
                                "file": str(candidate_path),
                                "row": str(row_number),
                                "code": code,
                                "error": (
                                    "Mã khách hàng được DUYET nhiều lần với "
                                    "nội dung khác nhau."
                                ),
                            }
                        )
                    continue
                approved_by_code[code_key] = row
                source_by_code[code_key] = str(candidate_path)
        finally:
            workbook.close()

    approved_rows = [
        {
            **approved_by_code[code],
            "_source_candidate_file": source_by_code[code],
        }
        for code in sorted(approved_by_code)
    ]
    return {
        "customer_master": str(customer_master.resolve()),
        "customer_master_sheet": master["sheet_name"],
        "customer_master_headers": CUSTOMER_MASTER_COLUMNS,
        "last_data_row": master["last_data_row"],
        "max_column": master["max_column"],
        "baseline_duplicate_codes": master["duplicate_codes"],
        "reviewed_candidate_files": reviewed_files,
        "approved_rows": approved_rows,
        "approved_count": len(approved_rows),
        "already_exists": already_exists,
        "errors": errors,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def scan_sell_in(sell_in_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        sell_in_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        ws = workbook.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        mapping = header_map(next(rows))
        required = [
            "NgayHoaDon",
            "MaKhachHangMoi",
            "TenKhachHang",
            "MaSanPhamMoi",
            "TenSanPham",
        ]
        require_headers(mapping, required, sell_in_path.name)

        customers: dict[str, dict[str, Any]] = {}
        products: dict[str, dict[str, Any]] = {}
        for values in rows:
            if not any(value not in (None, "") for value in values):
                continue
            customer_code = clean_id(values[mapping["MaKhachHangMoi"]])
            customer_key = normalize_key(customer_code)
            if customer_key and customer_key != "VKD3":
                customer = customers.setdefault(
                    customer_key,
                    {
                        "code": customer_code,
                        "names": [],
                        "dates": [],
                        "row_count": 0,
                    },
                )
                customer["row_count"] += 1
                name = clean_text(values[mapping["TenKhachHang"]])
                if name:
                    customer["names"].append(name)
                invoice_date = parse_excel_date(
                    values[mapping["NgayHoaDon"]],
                    workbook.epoch,
                )
                if invoice_date:
                    customer["dates"].append(invoice_date)

            product_code = clean_id(values[mapping["MaSanPhamMoi"]])
            product_key = normalize_key(product_code)
            if product_key:
                product = products.setdefault(
                    product_key,
                    {
                        "code": product_code,
                        "names": [],
                        "row_count": 0,
                    },
                )
                product["row_count"] += 1
                product_name = clean_text(
                    values[mapping["TenSanPham"]]
                )
                if product_name:
                    product["names"].append(product_name)

        for customer in customers.values():
            customer["names"] = unique_nonblank(customer["names"])
            customer["latest_date"] = (
                max(customer["dates"]).isoformat()
                if customer["dates"]
                else None
            )
            del customer["dates"]
        for product in products.values():
            product["names"] = unique_nonblank(product["names"])
        return {"customers": customers, "products": products}
    finally:
        workbook.close()


def read_product_codes(product_master: Path) -> set[str]:
    workbook = load_workbook(
        product_master,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        ws = workbook.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        mapping = header_map(next(rows))
        required = ["MaSanPhamMoi_Vikoda", "MaSanPhamMoi_VKD"]
        require_headers(mapping, required, product_master.name)
        codes: set[str] = set()
        for values in rows:
            for column in required:
                code = normalize_key(values[mapping[column]])
                if code:
                    codes.add(code)
        return codes
    finally:
        workbook.close()


def province_mappings(
    customer_records: list[dict[str, str]],
) -> dict[str, dict[str, list[str]]]:
    source: dict[str, dict[str, list[str]]] = defaultdict(
        lambda: {"MIEN": [], "VUNG": []}
    )
    for record in customer_records:
        province_key = clean_text(record["TINHTHANH"]).upper()
        if not province_key:
            continue
        for field in ("MIEN", "VUNG"):
            value = clean_text(record[field])
            if value:
                source[province_key][field].append(value)
    return {
        province: {
            field: unique_nonblank(values)
            for field, values in fields.items()
        }
        for province, fields in source.items()
    }


def choose_suggestion(
    values: Iterable[Any],
    latest_value: Any = None,
) -> tuple[str, bool]:
    unique = unique_nonblank(values)
    if not unique:
        return "", False
    if len(unique) == 1:
        return unique[0], False
    latest = clean_text(latest_value)
    return (latest if latest else unique[-1]), True


def matching_erp_files(
    source_dir: Path,
    month: int,
    year: int,
) -> list[Path]:
    result: list[Path] = []
    for file_path in sorted(source_dir.iterdir(), key=lambda p: p.name.lower()):
        if (
            file_path.name.startswith("~$")
            or file_path.suffix.lower() not in {".xlsx", ".xlsm"}
        ):
            continue
        match = ERP_PATTERN.search(file_path.name)
        if not match:
            continue
        if (
            int(match.group("month")) == month
            and int(match.group("year")) == year
        ):
            result.append(file_path)
    return result


def read_erp_customer_details(
    source_files: list[Path],
    candidate_codes: set[str],
) -> dict[str, Any]:
    details: dict[str, dict[str, Any]] = {}
    fields = [
        "TenKhachHang",
        "DiaChiGiaoHangImport",
        "DiaChiGiaoHang",
        "KenhBanHang",
        "TinhThanh",
        "QuanHuyen",
    ]
    for file_path in source_files:
        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            ws = workbook.worksheets[0]
            header_row, mapping = locate_erp_header(ws)
            for values in ws.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            ):
                code = clean_id(values[mapping["MaKhachHangMoi"]])
                code_key = normalize_key(code)
                if code_key not in candidate_codes:
                    continue
                detail = details.setdefault(
                    code_key,
                    {
                        "values": defaultdict(list),
                        "latest_date": None,
                        "latest_values": {},
                    },
                )
                invoice_date = parse_excel_date(
                    values[mapping["NgayHoaDon"]],
                    workbook.epoch,
                )
                row_values: dict[str, str] = {}
                for field in fields:
                    value = (
                        clean_text(values[mapping[field]])
                        if field in mapping
                        else ""
                    )
                    row_values[field] = value
                    if value:
                        detail["values"][field].append(value)
                if (
                    invoice_date is not None
                    and (
                        detail["latest_date"] is None
                        or invoice_date >= detail["latest_date"]
                    )
                ):
                    detail["latest_date"] = invoice_date
                    detail["latest_values"] = row_values
        finally:
            workbook.close()
    return details


def analyze_master_data(
    output_dir: Path,
    source_dir: Path,
    customer_master: Path,
    product_master: Path,
    candidate_dir: Path,
) -> dict[str, Any]:
    sell_in_path, month, year = find_latest_sell_in(output_dir)
    period_label = f"T{month:02d}_{year}"
    candidate_path = candidate_dir / f"Khach hang moi {period_label}.xlsx"

    master = read_customer_master(customer_master)
    sell_in = scan_sell_in(sell_in_path)
    missing_customer_keys = sorted(
        set(sell_in["customers"]) - set(master["codes"])
    )
    erp_files = matching_erp_files(source_dir, month, year)
    erp_details = read_erp_customer_details(
        erp_files,
        set(missing_customer_keys),
    )
    province_map = province_mappings(master["records"])
    existing_candidates = read_existing_candidate(candidate_path)

    candidates: list[dict[str, Any]] = []
    for code_key in missing_customer_keys:
        sell_customer = sell_in["customers"][code_key]
        detail = erp_details.get(
            code_key,
            {
                "values": defaultdict(list),
                "latest_date": None,
                "latest_values": {},
            },
        )
        values = detail["values"]
        latest_values = detail["latest_values"]
        conflicts: list[str] = []

        customer_name, conflict = choose_suggestion(
            values.get("TenKhachHang", []) or sell_customer["names"],
            latest_values.get("TenKhachHang"),
        )
        if conflict:
            conflicts.append("TenKhachHang")

        imported_address, imported_conflict = choose_suggestion(
            values.get("DiaChiGiaoHangImport", []),
            latest_values.get("DiaChiGiaoHangImport"),
        )
        fallback_address, fallback_conflict = choose_suggestion(
            values.get("DiaChiGiaoHang", []),
            latest_values.get("DiaChiGiaoHang"),
        )
        address = imported_address or fallback_address
        if (
            (imported_address and imported_conflict)
            or (not imported_address and fallback_conflict)
        ):
            conflicts.append("DiaChi")

        channel, conflict = choose_suggestion(
            values.get("KenhBanHang", []),
            latest_values.get("KenhBanHang"),
        )
        if conflict:
            conflicts.append("KenhBanHang")

        province, conflict = choose_suggestion(
            values.get("TinhThanh", []),
            latest_values.get("TinhThanh"),
        )
        if conflict:
            conflicts.append("TINHTHANH")

        district, conflict = choose_suggestion(
            values.get("QuanHuyen", []),
            latest_values.get("QuanHuyen"),
        )
        if conflict:
            conflicts.append("QUANHUYEN")

        region = ""
        domain = ""
        if province:
            mapped = province_map.get(province.upper(), {})
            domains = mapped.get("MIEN", [])
            regions = mapped.get("VUNG", [])
            if len(domains) == 1:
                domain = domains[0]
            elif len(domains) > 1:
                conflicts.append("MIEN")
            if len(regions) == 1:
                region = regions[0]
            elif len(regions) > 1:
                conflicts.append("VUNG")

        latest_date = (
            detail["latest_date"].isoformat()
            if detail["latest_date"] is not None
            else sell_customer["latest_date"]
        )
        row: dict[str, Any] = {
            "MaKhachHangMoi": sell_customer["code"],
            "TenKhachHang": customer_name,
            "TenKhachHangdaydu": customer_name,
            "DiaChi": address,
            "KenhBanHang": channel,
            "Loaikhachhang": "",
            "Hethong MT": "",
            "MIEN": domain,
            "VUNG": region,
            "TINHTHANH": province,
            "QUANHUYEN": district,
            "CODE": "",
            "TENKH": "",
            "KyDuLieu": period_label,
            "FileNguonERP": "; ".join(path.name for path in erp_files),
            "NgayHoaDonGanNhat": latest_date,
            "SoDongSellIn": sell_customer["row_count"],
            "TruongXungDot": "; ".join(sorted(set(conflicts))),
            "TrangThaiDuyet": (
                "CAN BO SUNG" if conflicts else "CHO DUYET"
            ),
            "GhiChuDuyet": (
                "Địa chỉ lấy từ dữ liệu giao hàng ERP; kiểm tra lại "
                "trước khi duyệt."
                if address
                else ""
            ),
        }

        existing = existing_candidates.get(code_key)
        if existing:
            for column in CUSTOMER_MASTER_COLUMNS:
                existing_value = existing.get(column)
                if existing_value not in (None, ""):
                    row[column] = existing_value
            existing_status = normalize_status(
                existing.get("TrangThaiDuyet")
            )
            if existing_status in APPROVAL_STATUSES:
                row["TrangThaiDuyet"] = existing_status
            existing_note = clean_text(existing.get("GhiChuDuyet"))
            if existing_note:
                row["GhiChuDuyet"] = existing_note
        candidates.append(row)

    product_codes = read_product_codes(product_master)
    missing_product_keys = sorted(
        set(sell_in["products"]) - product_codes
    )
    missing_products = []
    for code_key in missing_product_keys:
        product = sell_in["products"][code_key]
        missing_products.append(
            {
                "MaSanPhamMoi": product["code"],
                "TenSanPham": (
                    product["names"][0] if product["names"] else ""
                ),
                "SoDongSellIn": product["row_count"],
                "KyDuLieu": period_label,
                "GhiChu": (
                    "Chưa có trong hai cột mã của Danh Muc San Pham.xlsx; "
                    "không tự động bổ sung."
                ),
            }
        )

    return {
        "latest_period": {
            "month": month,
            "year": year,
            "label": period_label,
        },
        "latest_sell_in_file": str(sell_in_path.resolve()),
        "erp_source_files": [str(path.resolve()) for path in erp_files],
        "customer_master": str(customer_master.resolve()),
        "product_master": str(product_master.resolve()),
        "candidate_file": str(candidate_path.resolve()),
        "candidate_count": len(candidates),
        "candidate_rows": candidates,
        "missing_product_count": len(missing_products),
        "missing_products": missing_products,
        "customer_master_preexisting_duplicate_codes": master[
            "duplicate_codes"
        ],
        "warnings": (
            []
            if erp_files
            else [
                "Không tìm thấy file ERP cùng kỳ với file Sell In mới nhất."
            ]
        ),
    }


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def apply_approved_customers_portable(
    plan: dict[str, Any],
    backup_dir: Path,
) -> dict[str, Any]:
    if plan["errors"]:
        raise ValueError(
            "Không thể cập nhật danh mục khách hàng: "
            + "; ".join(item["error"] for item in plan["errors"])
        )
    master_path = Path(plan["customer_master"])
    approved_rows = plan["approved_rows"]
    result: dict[str, Any] = {
        "customer_master": str(master_path),
        "approved_count": len(approved_rows),
        "appended_codes": [],
        "already_exists": plan["already_exists"],
        "backup_file": None,
        "changed": False,
    }
    if not approved_rows:
        return result

    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / (
        f"{master_path.stem}_{timestamp}{master_path.suffix}"
    )
    shutil.copy2(master_path, backup_path)

    workbook = load_workbook(master_path)
    temp_path = master_path.with_name(
        f".{master_path.stem}.{os.getpid()}.tmp{master_path.suffix}"
    )
    try:
        ws = workbook[plan["customer_master_sheet"]]
        last_row = int(plan["last_data_row"])
        source_style_row = max(2, last_row)
        for offset, approved in enumerate(approved_rows, start=1):
            target_row = last_row + offset
            copy_row_style(ws, source_style_row, target_row)
            for column_number, column in enumerate(
                CUSTOMER_MASTER_COLUMNS,
                start=1,
            ):
                ws.cell(target_row, column_number).value = approved[column]
            if ws.max_column >= 14:
                ws.cell(target_row, 14).value = None
            ws.cell(target_row, 1).number_format = "@"
            ws.cell(target_row, 12).number_format = "@"
            result["appended_codes"].append(
                approved["MaKhachHangMoi"]
            )
        workbook.save(temp_path)
    finally:
        workbook.close()

    try:
        os.replace(temp_path, master_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Hãy đóng file danh mục khách hàng trước khi chạy: {master_path}"
        ) from exc
    finally:
        if temp_path.exists():
            temp_path.unlink()

    verification = read_customer_master(master_path)
    missing = [
        code
        for code in result["appended_codes"]
        if normalize_key(code) not in verification["codes"]
    ]
    if missing:
        raise RuntimeError(
            "Không xác minh được mã khách hàng vừa thêm: "
            + ", ".join(missing)
        )
    result["backup_file"] = str(backup_path)
    result["changed"] = True
    return result


def style_review_header(ws, end_column: int, color: str) -> None:
    header_fill = PatternFill("solid", fgColor=color)
    header_font = Font(
        name="Aptos",
        size=10,
        bold=True,
        color="FFFFFF",
    )
    header_border = Border(
        bottom=Side(style="medium", color=color),
    )
    for cell in ws[1][:end_column]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = header_border
    ws.row_dimensions[1].height = 34


def write_review_workbook_portable(
    analysis: dict[str, Any],
    preview_path: Path | None = None,
) -> Path:
    output_path = Path(analysis["candidate_file"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(
        f".{output_path.stem}.{os.getpid()}.tmp.xlsx"
    )

    workbook = Workbook()
    customer_ws = workbook.active
    customer_ws.title = "Khach hang moi"
    customer_ws.sheet_view.showGridLines = False
    customer_ws.freeze_panes = "A2"
    customer_ws.append(CANDIDATE_COLUMNS)
    for row in analysis["candidate_rows"]:
        values = [row.get(column) for column in CANDIDATE_COLUMNS]
        if values[15]:
            values[15] = datetime.strptime(
                values[15],
                "%Y-%m-%d",
            ).date()
        customer_ws.append(values)
    style_review_header(customer_ws, len(CANDIDATE_COLUMNS), "00A651")

    customer_widths = [
        18, 28, 32, 42, 15, 18, 17, 14, 18, 17,
        18, 18, 22, 12, 54, 16, 14, 24, 18, 48,
    ]
    last_customer_row = len(analysis["candidate_rows"]) + 1
    for column_number, width in enumerate(customer_widths, start=1):
        customer_ws.column_dimensions[
            get_column_letter(column_number)
        ].width = width
    for row_number in range(2, last_customer_row + 1):
        customer_ws.cell(row_number, 1).number_format = "@"
        customer_ws.cell(row_number, 12).number_format = "@"
        customer_ws.cell(row_number, 16).number_format = "dd/mm/yyyy"
        customer_ws.cell(row_number, 17).number_format = "#,##0"
        for column_number in range(1, len(CANDIDATE_COLUMNS) + 1):
            customer_ws.cell(
                row_number,
                column_number,
            ).alignment = Alignment(vertical="center")
    if analysis["candidate_rows"]:
        table = Table(
            displayName=(
                "KhachHangMoi_"
                f"{analysis['latest_period']['year']}_"
                f"{analysis['latest_period']['month']:02d}"
            ),
            ref=f"A1:T{last_customer_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False,
        )
        customer_ws.add_table(table)
        validation = DataValidation(
            type="list",
            formula1='"CHO DUYET,DUYET,TU CHOI,CAN BO SUNG"',
            allow_blank=False,
        )
        customer_ws.add_data_validation(validation)
        validation.add(f"S2:S{last_customer_row}")
        customer_ws.conditional_formatting.add(
            f"S2:S{last_customer_row}",
            FormulaRule(
                formula=['$S2="DUYET"'],
                fill=PatternFill("solid", fgColor="DCFCE7"),
                font=Font(color="166534", bold=True),
            ),
        )
        customer_ws.conditional_formatting.add(
            f"S2:S{last_customer_row}",
            FormulaRule(
                formula=['$S2="TU CHOI"'],
                fill=PatternFill("solid", fgColor="FEE2E2"),
                font=Font(color="991B1B", bold=True),
            ),
        )
        customer_ws.conditional_formatting.add(
            f"S2:S{last_customer_row}",
            FormulaRule(
                formula=['$S2="CAN BO SUNG"'],
                fill=PatternFill("solid", fgColor="FEF3C7"),
                font=Font(color="92400E", bold=True),
            ),
        )

    product_ws = workbook.create_sheet("Ma SP chua co")
    product_ws.sheet_view.showGridLines = False
    product_ws.freeze_panes = "A2"
    product_ws.append(PRODUCT_REVIEW_COLUMNS)
    for row in analysis["missing_products"]:
        product_ws.append(
            [row.get(column) for column in PRODUCT_REVIEW_COLUMNS]
        )
    style_review_header(
        product_ws,
        len(PRODUCT_REVIEW_COLUMNS),
        "1F4E78",
    )
    for column_number, width in enumerate(
        [20, 58, 16, 12, 58],
        start=1,
    ):
        product_ws.column_dimensions[
            get_column_letter(column_number)
        ].width = width
    if analysis["missing_products"]:
        last_product_row = len(analysis["missing_products"]) + 1
        table = Table(
            displayName=(
                "SanPhamThieu_"
                f"{analysis['latest_period']['year']}_"
                f"{analysis['latest_period']['month']:02d}"
            ),
            ref=f"A1:E{last_product_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False,
        )
        product_ws.add_table(table)

    guide_ws = workbook.create_sheet("Huong dan")
    guide_ws.sheet_view.showGridLines = False
    guide_ws["A1"] = "RÀ SOÁT DỮ LIỆU DANH MỤC"
    guide_ws["A1"].font = Font(
        name="Aptos Display",
        size=16,
        bold=True,
        color="FFFFFF",
    )
    guide_ws["A1"].fill = PatternFill("solid", fgColor="0F766E")
    guide_ws.merge_cells("A1:D1")
    guide_ws["A3"] = "Kỳ dữ liệu"
    guide_ws["B3"] = analysis["latest_period"]["label"]
    guide_ws["A4"] = "File Sell In"
    guide_ws["B4"] = Path(analysis["latest_sell_in_file"]).name
    guide_ws["A5"] = "Khách hàng chờ rà soát"
    guide_ws["B5"] = analysis["candidate_count"]
    guide_ws["A6"] = "Mã sản phẩm chưa có"
    guide_ws["B6"] = analysis["missing_product_count"]
    guide_ws["A8"] = "Cách duyệt khách hàng"
    guide_ws["B8"] = (
        "Kiểm tra các trường trên sheet Khach hang moi, chọn DUYET "
        "tại cột TrangThaiDuyet, lưu và chạy lại quy trình. Chỉ dòng "
        "DUYET mới được nối thêm vào danh mục khách hàng."
    )
    guide_ws["A9"] = "Trường nhập thủ công"
    guide_ws["B9"] = (
        "Loaikhachhang, Hethong MT, CODE và TENKH không được tự suy "
        "diễn. Bổ sung khi nghiệp vụ yêu cầu."
    )
    guide_ws["A10"] = "Danh mục sản phẩm"
    guide_ws["B10"] = (
        "Sheet Ma SP chua co chỉ để cảnh báo. Quy trình không tự động "
        "thêm sản phẩm vào Danh Muc San Pham.xlsx."
    )
    guide_ws.column_dimensions["A"].width = 25
    guide_ws.column_dimensions["B"].width = 95
    guide_ws.column_dimensions["C"].width = 3
    guide_ws.column_dimensions["D"].width = 3
    for row_number in range(3, 11):
        guide_ws.cell(row_number, 1).font = Font(bold=True)
        guide_ws.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
    for row_number in (8, 9, 10):
        guide_ws.row_dimensions[row_number].height = 42

    try:
        workbook.save(temp_path)
    finally:
        workbook.close()
    try:
        os.replace(temp_path, output_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Hãy đóng file rà soát trước khi chạy: {output_path}"
        ) from exc
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return output_path


def verify_master_data_artifacts(
    analysis: dict[str, Any],
    apply_report: dict[str, Any],
) -> dict[str, Any]:
    candidate_path = Path(analysis["candidate_file"])
    problems: list[str] = []
    workbook = load_workbook(
        candidate_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        expected_sheets = [
            "Khach hang moi",
            "Ma SP chua co",
            "Huong dan",
        ]
        if workbook.sheetnames != expected_sheets:
            problems.append(
                "Sai danh sách sheet trong file rà soát dữ liệu danh mục."
            )

        customer_ws = workbook["Khach hang moi"]
        customer_rows = customer_ws.iter_rows(values_only=True)
        customer_headers = list(next(customer_rows))
        if customer_headers != CANDIDATE_COLUMNS:
            problems.append("Sai cột của sheet Khach hang moi.")
        candidate_codes: list[str] = []
        invalid_statuses: list[str] = []
        candidate_count = 0
        for values in customer_rows:
            if not any(value not in (None, "") for value in values):
                continue
            candidate_count += 1
            code = clean_id(values[0])
            candidate_codes.append(normalize_key(code))
            status = normalize_status(values[18])
            if status not in APPROVAL_STATUSES:
                invalid_statuses.append(status)
        if candidate_count != analysis["candidate_count"]:
            problems.append("Sai số dòng khách hàng chờ rà soát.")
        if len(candidate_codes) != len(set(candidate_codes)):
            problems.append("Trùng MaKhachHangMoi trong file rà soát.")
        if invalid_statuses:
            problems.append(
                "Có TrangThaiDuyet không hợp lệ: "
                + ", ".join(sorted(set(invalid_statuses)))
            )

        product_ws = workbook["Ma SP chua co"]
        product_rows = product_ws.iter_rows(values_only=True)
        product_headers = list(next(product_rows))
        if product_headers != PRODUCT_REVIEW_COLUMNS:
            problems.append("Sai cột của sheet Ma SP chua co.")
        product_count = sum(
            1
            for values in product_rows
            if any(value not in (None, "") for value in values)
        )
        if product_count != analysis["missing_product_count"]:
            problems.append("Sai số mã sản phẩm chưa có.")
    finally:
        workbook.close()

    master = read_customer_master(Path(analysis["customer_master"]))
    missing_appended_codes = [
        code
        for code in apply_report.get("appended_codes", [])
        if normalize_key(code) not in master["codes"]
    ]
    if missing_appended_codes:
        problems.append(
            "Danh mục khách hàng thiếu mã đã duyệt: "
            + ", ".join(missing_appended_codes)
        )

    return {
        "candidate_file": str(candidate_path),
        "candidate_count": analysis["candidate_count"],
        "missing_product_count": analysis["missing_product_count"],
        "approved_appended_count": len(
            apply_report.get("appended_codes", [])
        ),
        "missing_appended_codes": missing_appended_codes,
        "warnings": analysis.get("warnings", []),
        "problems": problems,
        "ok": not problems,
    }
