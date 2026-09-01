from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COLUMNS = [
    "Vung",
    "KhuVuc",
    "NgayHoaDon",
    "MaKhachHangMoi",
    "TenKhachHang",
    "MaSanPhamMoi",
    "TenSanPham",
    "SoLuong",
    "DonGia",
    "ThanhTien",
    "LoaiDonHang",
    "GhiChu",
    "Thang",
    "Nam",
]

FILE_PATTERN = re.compile(r"^Sell in T(\d{1,2})_(20\d{2})\.xlsx$", re.IGNORECASE)

ALLOWED_INVOICE_TYPE_KEYS = {
    "DON HANG BAN",
    "DON TRA HANG",
    "HOA DON BAN",
    "HOA DON TRA HANG",
    "NHAP HOA DON TRA HANG",
}

ALLOWED_INVOICE_TYPES = [
    "Đơn hàng bán (tương đương Hóa đơn bán)",
    "Đơn trả hàng (tương đương Hóa đơn trả hàng)",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def normalized_key(value: Any) -> str:
    text = clean_text(value)
    text = text.replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(
        character for character in text
        if not unicodedata.combining(character)
    )
    return text.upper()


def is_revenue_invoice_type(value: Any) -> bool:
    return normalized_key(value) in ALLOWED_INVOICE_TYPE_KEYS


def numeric_value(value: Any, *, field: str, location: str) -> int | float:
    if value is None or clean_text(value) == "":
        return 0
    if isinstance(value, bool):
        raise ValueError(f"{location}: {field} không được là Boolean")
    if isinstance(value, (int, float)):
        return value
    text = clean_text(value)
    if text.startswith("="):
        raise ValueError(f"{location}: {field} không được là công thức")
    try:
        number = Decimal(text.replace(",", ""))
    except InvalidOperation as error:
        raise ValueError(f"{location}: {field} không phải số: {text}") from error
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def date_value(value: Any, *, location: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"{location}: NgayHoaDon không phải ngày Excel: {value!r}")


import calendar


def parse_as_of(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def derive_as_of(source_dir: Path, requested_as_of: date | None = None) -> date:
    if requested_as_of is not None:
        return requested_as_of

    discovered_periods: list[tuple[int, int]] = []
    for path in source_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        match = FILE_PATTERN.match(path.name)
        if match:
            month = int(match.group(1))
            year = int(match.group(2))
            discovered_periods.append((year, month))

    if discovered_periods:
        max_year = max(year for year, _ in discovered_periods)
        max_month = max(month for year, month in discovered_periods if year == max_year)
        today = date.today()
        if today.year == max_year and today.month == max_month:
            return today
        last_day = calendar.monthrange(max_year, max_month)[1]
        return date(max_year, max_month, last_day)

    return date.today()


def expected_periods(as_of: date) -> list[tuple[int, int]]:
    periods = []
    for year in (as_of.year - 1, as_of.year):
        for month in range(1, as_of.month + 1):
            periods.append((year, month))
    return periods


def main() -> int:
    parser = argparse.ArgumentParser(description="Ghép dữ liệu Sell In cùng kỳ")
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--staging-dir", required=True)
    parser.add_argument("--as-of-date")
    args = parser.parse_args()

    source_dir = Path(args.source_dir).resolve()
    staging_dir = Path(args.staging_dir).resolve()
    staging_dir.mkdir(parents=True, exist_ok=True)
    data_file = staging_dir / "sell_in_data.json"
    audit_file = staging_dir / "sell_in_audit.json"
    as_of = derive_as_of(source_dir, parse_as_of(args.as_of_date))

    problems: list[str] = []
    warnings: list[str] = []
    expected = expected_periods(as_of)
    expected_set = set(expected)
    files_by_period: dict[tuple[int, int], list[Path]] = {}

    for path in source_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        match = FILE_PATTERN.match(path.name)
        if not match:
            continue
        month = int(match.group(1))
        year = int(match.group(2))
        if (year, month) in expected_set:
            files_by_period.setdefault((year, month), []).append(path)

    selected_files: list[tuple[int, int, Path]] = []
    for year, month in expected:
        paths = files_by_period.get((year, month), [])
        if not paths:
            problems.append(f"Thiếu file Sell in T{month:02d}_{year}.xlsx")
            continue
        if len(paths) > 1:
            problems.append(
                f"Có nhiều file cho kỳ {year}{month:02d}: "
                + ", ".join(path.name for path in paths)
            )
            continue
        selected_files.append((year, month, paths[0]))

    rows: list[list[Any]] = []
    period_reports: list[dict[str, Any]] = []
    for year, month, path in selected_files:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if "Sell in" not in workbook.sheetnames:
                problems.append(f"{path.name}: không có sheet Sell in")
                continue
            worksheet = workbook["Sell in"]
            headers = [
                worksheet.cell(1, column).value
                for column in range(1, len(COLUMNS) + 1)
            ]
            if headers != COLUMNS:
                problems.append(f"{path.name}: tiêu đề không đúng: {headers}")
                continue

            report = {
                "period": f"{year}{month:02d}",
                "file": path.name,
                "source_rows": 0,
                "rows": 0,
                "quantity": Decimal(0),
                "amount": 0,
                "excluded_rows": 0,
                "excluded_quantity": Decimal(0),
                "excluded_amount": 0,
                "invoice_types": {},
            }
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    min_col=1,
                    max_col=len(COLUMNS),
                    values_only=True,
                ),
                start=2,
            ):
                if not any(value is not None for value in values):
                    continue
                location = f"{path.name}!R{row_number}"
                if any(isinstance(value, str) and value.startswith("=") for value in values):
                    problems.append(f"{location}: không được có công thức")
                    continue
                try:
                    normalized = [
                        clean_text(values[0]),
                        clean_text(values[1]),
                        date_value(values[2], location=location),
                        clean_text(values[3]),
                        clean_text(values[4]),
                        clean_text(values[5]),
                        clean_text(values[6]),
                        numeric_value(values[7], field="SoLuong", location=location),
                        numeric_value(values[8], field="DonGia", location=location),
                        numeric_value(values[9], field="ThanhTien", location=location),
                        clean_text(values[10]),
                        clean_text(values[11]),
                        int(values[12]),
                        int(values[13]),
                    ]
                except (TypeError, ValueError) as error:
                    problems.append(str(error))
                    continue

                if normalized[12] != month or normalized[13] != year:
                    problems.append(
                        f"{location}: Thang/Nam {normalized[12]}/{normalized[13]} "
                        f"không khớp tên file {month}/{year}"
                    )
                    continue

                report["source_rows"] += 1
                invoice_type = normalized[10] or "(trống)"
                invoice_type_report = report["invoice_types"].setdefault(
                    invoice_type,
                    {
                        "invoice_type": invoice_type,
                        "included": is_revenue_invoice_type(invoice_type),
                        "rows": 0,
                        "quantity": Decimal(0),
                        "amount": 0,
                    },
                )
                invoice_type_report["rows"] += 1
                invoice_type_report["quantity"] += Decimal(str(normalized[7]))
                invoice_type_report["amount"] += int(round(normalized[9]))

                if not invoice_type_report["included"]:
                    report["excluded_rows"] += 1
                    report["excluded_quantity"] += Decimal(str(normalized[7]))
                    report["excluded_amount"] += int(round(normalized[9]))
                    continue

                if not normalized[3] or not normalized[5]:
                    warnings.append(f"{location}: thiếu mã khách hàng hoặc mã sản phẩm")

                rows.append(normalized)
                report["rows"] += 1
                report["quantity"] += Decimal(str(normalized[7]))
                report["amount"] += int(round(normalized[9]))

            report["quantity"] = str(report["quantity"].normalize())
            report["excluded_quantity"] = str(
                report["excluded_quantity"].normalize()
            )
            report["invoice_types"] = [
                {
                    **item,
                    "quantity": str(item["quantity"].normalize()),
                }
                for _, item in sorted(report["invoice_types"].items())
            ]
            period_reports.append(report)
        finally:
            workbook.close()

    rows.sort(
        key=lambda row: (
            row[13],
            row[12],
            row[2],
            row[3],
            row[5],
        )
    )

    payload = {
        "schema_version": 1,
        "generated_at": datetime.now().astimezone().isoformat(),
        "as_of_date": as_of.isoformat(),
        "current_year": as_of.year,
        "through_month": as_of.month,
        "columns": COLUMNS,
        "rows": rows,
    }
    audit = {
        "generated_at": payload["generated_at"],
        "source_dir": str(source_dir),
        "data_file": str(data_file),
        "as_of_date": as_of.isoformat(),
        "expected_periods": [f"{year}{month:02d}" for year, month in expected],
        "invoice_type_filter": {
            "source_column": "LoaiDonHang",
            "business_meaning": "Loại hóa đơn",
            "allowed": ALLOWED_INVOICE_TYPES,
        },
        "periods": period_reports,
        "total_rows": len(rows),
        "total_source_rows": sum(
            report["source_rows"] for report in period_reports
        ),
        "total_excluded_rows": sum(
            report["excluded_rows"] for report in period_reports
        ),
        "total_quantity": str(
            sum((Decimal(report["quantity"]) for report in period_reports), Decimal(0))
        ),
        "total_amount": sum(report["amount"] for report in period_reports),
        "total_excluded_quantity": str(
            sum(
                (
                    Decimal(report["excluded_quantity"])
                    for report in period_reports
                ),
                Decimal(0),
            )
        ),
        "total_excluded_amount": sum(
            report["excluded_amount"] for report in period_reports
        ),
        "warnings": warnings,
        "problems": problems,
    }
    data_file.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    audit_file.write_text(
        json.dumps(audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "data_file": str(data_file),
                "audit_file": str(audit_file),
                "records": len(rows),
                "source_records": audit["total_source_rows"],
                "excluded_records": audit["total_excluded_rows"],
                "excluded_amount": audit["total_excluded_amount"],
                "periods": len(period_reports),
                "warnings": len(warnings),
                "problems": len(problems),
            },
            ensure_ascii=False,
        )
    )
    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
