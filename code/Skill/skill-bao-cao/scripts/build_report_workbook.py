from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


TARGET_COLUMNS = [
    "Ky",
    "Nam",
    "Thang",
    "MaKhachHangMoi",
    "TenKhachHang",
    "TargetVikoda",
    "TargetTong",
    "NgayCapNhatNguon",
    "NguonFile",
]

DATA_COLUMNS = [
    "Vung",
    "KhuVuc",
    "NgayHoaDon",
    "MaKhachHangMoi",
    "TenKhachHang",
    "MaSanPhamMoi",
    "TenSanPham",
    "SoLuong",
    "DonGia",
    "ThanhTien",
    "LoaiDonHang",
    "GhiChu",
    "Thang",
    "Nam",
]

DMKH_COLUMNS = [
    "MaKhachHangMoi",
    "TenKhachHang",
    "TenKhachHangdaydu",
    "DiaChi",
    "KenhBanHang",
    "Loaikhachhang",
    "Hethong MT",
    "MIEN",
    "VUNG",
    "TINHTHANH",
    "QUANHUYEN",
    "CODE",
    "TENKH",
    "ThongTinBoSungNguon",
]


def load_payload(path: Path, expected_columns: list[str], row_key: str):
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("columns") != expected_columns:
        raise ValueError(
            f"Cot staging khong dung tai {path}: {payload.get('columns')}"
        )
    rows = payload.get(row_key)
    if not isinstance(rows, list) or not rows:
        raise ValueError(f"Khong co du lieu staging tai {path}")
    return payload


def parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    parsed = datetime.fromisoformat(value)
    return parsed.replace(tzinfo=None)


def style_header(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    header_border = Border(bottom=Side(style="medium", color="17365D"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    worksheet.row_dimensions[1].height = 30


def configure_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    style_header(worksheet)


def add_table(worksheet, name: str, max_column: str) -> None:
    table = Table(
        displayName=name,
        ref=f"A1:{max_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def set_widths(worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def create_target_sheet(workbook: Workbook, payload) -> int:
    worksheet = workbook.active
    worksheet.title = "Target"
    worksheet.append(TARGET_COLUMNS)
    for record in payload["records"]:
        worksheet.append(
            [
                str(record["Ky"]),
                int(record["Nam"]),
                int(record["Thang"]),
                str(record["MaKhachHangMoi"]),
                str(record["TenKhachHang"]),
                int(record["TargetVikoda"]),
                int(record["TargetTong"]),
                parse_datetime(record.get("NgayCapNhatNguon")),
                str(record["NguonFile"]),
            ]
        )
    configure_sheet(worksheet)
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 1).number_format = "@"
        worksheet.cell(row, 2).number_format = "0"
        worksheet.cell(row, 3).number_format = "0"
        worksheet.cell(row, 4).number_format = "@"
        worksheet.cell(row, 6).number_format = "#,##0"
        worksheet.cell(row, 7).number_format = "#,##0"
        worksheet.cell(row, 8).number_format = "dd/mm/yyyy hh:mm"
    set_widths(
        worksheet,
        {
            "A": 11,
            "B": 9,
            "C": 9,
            "D": 19,
            "E": 36,
            "F": 19,
            "G": 19,
            "H": 21,
            "I": 46,
        },
    )
    add_table(worksheet, "tblTarget", "I")
    return len(payload["records"])


def create_data_sheet(workbook: Workbook, payload) -> int:
    worksheet = workbook.create_sheet("Data")
    worksheet.append(DATA_COLUMNS)
    for raw_row in payload["rows"]:
        worksheet.append(
            [
                str(raw_row[0] or ""),
                str(raw_row[1] or ""),
                date.fromisoformat(raw_row[2]),
                str(raw_row[3] or ""),
                str(raw_row[4] or ""),
                str(raw_row[5] or ""),
                str(raw_row[6] or ""),
                raw_row[7],
                raw_row[8],
                raw_row[9],
                str(raw_row[10] or ""),
                str(raw_row[11] or ""),
                int(raw_row[12]),
                int(raw_row[13]),
            ]
        )
    configure_sheet(worksheet)
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 3).number_format = "dd/mm/yyyy"
        worksheet.cell(row, 4).number_format = "@"
        worksheet.cell(row, 6).number_format = "@"
        quantity_cell = worksheet.cell(row, 8)
        quantity = Decimal(str(quantity_cell.value))
        quantity_cell.number_format = (
            "#,##0" if quantity == quantity.to_integral_value() else "#,##0.##"
        )
        worksheet.cell(row, 9).number_format = "#,##0"
        worksheet.cell(row, 10).number_format = "#,##0"
        worksheet.cell(row, 13).number_format = "0"
        worksheet.cell(row, 14).number_format = "0"
    set_widths(
        worksheet,
        {
            "A": 12,
            "B": 14,
            "C": 13,
            "D": 18,
            "E": 34,
            "F": 18,
            "G": 56,
            "H": 13,
            "I": 15,
            "J": 17,
            "K": 18,
            "L": 46,
            "M": 9,
            "N": 9,
        },
    )
    add_table(worksheet, "tblDataSellIn", "N")
    return len(payload["rows"])


def create_dmkh_sheet(workbook: Workbook, payload) -> int:
    worksheet = workbook.create_sheet("DMKH")
    worksheet.append(DMKH_COLUMNS)
    for raw_row in payload["rows"]:
        worksheet.append([str(value or "") for value in raw_row])
    configure_sheet(worksheet)
    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(DMKH_COLUMNS),
    ):
        for cell in row:
            cell.number_format = "@"
    set_widths(
        worksheet,
        {
            "A": 18,
            "B": 28,
            "C": 44,
            "D": 58,
            "E": 14,
            "F": 20,
            "G": 18,
            "H": 18,
            "I": 20,
            "J": 24,
            "K": 22,
            "L": 18,
            "M": 26,
            "N": 24,
        },
    )
    add_table(worksheet, "tblDMKH", "N")
    return len(payload["rows"])


def main() -> None:
    parser = argparse.ArgumentParser(description="Tao Bao_Cao_Sell_in.xlsx")
    parser.add_argument("--target-data-file", required=True)
    parser.add_argument("--sell-in-data-file", required=True)
    parser.add_argument("--dmkh-data-file", required=True)
    parser.add_argument("--output-file", required=True)
    args = parser.parse_args()

    target_data_file = Path(args.target_data_file).resolve()
    sell_in_data_file = Path(args.sell_in_data_file).resolve()
    dmkh_data_file = Path(args.dmkh_data_file).resolve()
    output_file = Path(args.output_file).resolve()
    output_file.parent.mkdir(parents=True, exist_ok=True)

    target_payload = load_payload(target_data_file, TARGET_COLUMNS, "records")
    sell_in_payload = load_payload(sell_in_data_file, DATA_COLUMNS, "rows")
    dmkh_payload = load_payload(dmkh_data_file, DMKH_COLUMNS, "rows")

    workbook = Workbook()
    try:
        target_records = create_target_sheet(workbook, target_payload)
        sell_in_records = create_data_sheet(workbook, sell_in_payload)
        dmkh_records = create_dmkh_sheet(workbook, dmkh_payload)
        workbook.create_sheet("PIVOT")
        pivot_data_sheet = workbook.create_sheet("PVT_DATA")
        pivot_data_sheet.sheet_state = "hidden"
        workbook.active = 0

        temporary_file = output_file.with_name(
            f"{output_file.stem}.building{output_file.suffix}"
        )
        if temporary_file.exists():
            temporary_file.unlink()
        workbook.save(temporary_file)
        temporary_file.replace(output_file)
    finally:
        workbook.close()

    print(
        json.dumps(
            {
                "output_file": str(output_file),
                "target_records": target_records,
                "sell_in_records": sell_in_records,
                "dmkh_records": dmkh_records,
            }
        )
    )


if __name__ == "__main__":
    main()
