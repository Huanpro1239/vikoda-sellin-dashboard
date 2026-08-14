"""Build a refreshable Power BI Project for the Sell In report.

The Excel report remains the operational source of truth.  This builder consumes
the same staging JSON files as the Excel pipeline and writes a portable Power BI
Project plus a star-schema CSV layer.  The project is deliberately generated as
text (PBIP/TMSL/PBIR) so it can be reviewed, diffed and regenerated every month.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import shutil
import sys
import unicodedata
import uuid
from collections import OrderedDict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from report_model import REPORTING_STRUCTURE, normalize_key, normalize_reporting_pair


PACKAGE_VERSION = "2.0.0"
REPORT_NAME = "Vikoda_SellIn_PowerBI"
REPORT_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/report/definition"
SEMANTIC_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/semanticModel"
# Revenue and target are already divided by 1,000,000 in the DAX measures.
# Keep the executive view compact: no decimal places for million-VND amounts.
MONEY_FORMAT = '#,##0 "tr";(#,##0 "tr");-'
QUANTITY_FORMAT = '#,##0;(#,##0);-'
# Tăng trưởng hiển thị dấu rõ ràng. Kiểu kế toán `(73,4%)` dễ bị đọc nhầm
# thành số dương khi liếc nhanh trên thẻ KPI.
GROWTH_FORMAT = "+0.0%;-0.0%;0.0%"

# Shared executive rail geometry.  Keeping the filter rail on one grid avoids
# slicers being clipped by the KPI cards on the CEO page and keeps every page
# visually consistent.  The slicer height deliberately leaves enough room for
# both the label and the dropdown value rendered by Power BI.
SIDEBAR_X = 12
SIDEBAR_WIDTH = 196
SIDEBAR_SLICER_Y = 308
SIDEBAR_SLICER_HEIGHT = 60
SIDEBAR_SLICER_STEP = 68
SIDEBAR_DATE_SLICER_HEIGHT = 72
SIDEBAR_CARD_Y = 390
SIDEBAR_CARD_WIDTH = SIDEBAR_WIDTH
SIDEBAR_CARD_HEIGHT = 62
SIDEBAR_CARD_STEP = 68

# Rail nav: sáu trang nên nút phải gọn lại để không đè lên khối slicer bên dưới.
NAV_BUTTON_Y = 128
NAV_BUTTON_HEIGHT = 26
NAV_BUTTON_STEP = 29

# Lưới vùng nội dung: x 236..1264 (rộng 1028), y 92..704 (cao 612).
CONTENT_X = 236
CONTENT_Y = 92
CONTENT_WIDTH = 1028
ROW1_HEIGHT = 286
ROW2_Y = 386
ROW2_HEIGHT = 318
# Ba cột đều nhau trong vùng nội dung.
COL3_X = (236, 581, 926)
COL3_WIDTH = 338
# Hai cột đều nhau.
COL2_X = (236, 754)
COL2_WIDTH = 510


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


# File cấu hình cục bộ của Power BI Desktop cần giữ qua các lần dựng lại.
# `cache.abf` cố tình KHÔNG nằm trong danh sách: đó là bản chụp số liệu cũ.
PRESERVED_LOCAL_FILES = ("localSettings.json", "editorSettings.json")


def remove_stale_pbix(output_dir: Path) -> str | None:
    """Xóa file `.pbix` sót lại từ quy trình cũ.

    PBIX là bản chụp tĩnh: không script nào ghi được vào nó, nên sau mỗi lần
    cập nhật dữ liệu nó lập tức lệch với CSV mà không có dấu hiệu nào. Dự án
    chuyển hẳn sang PBIP để chỉ còn một nguồn duy nhất.
    """
    stale = output_dir / f"{REPORT_NAME}.pbix"
    if not stale.is_file():
        return None
    try:
        stale.unlink()
    except OSError as error:  # đang mở trong Power BI Desktop
        print(
            f"Canh bao: khong xoa duoc {stale.name} ({error}). "
            "Dong Power BI Desktop roi chay lai neu muon don sach.",
            file=sys.stderr,
        )
        return None
    return stale.name


def preserve_local_settings(*directories: Path) -> dict[Path, bytes]:
    """Đọc trước các file `.pbi/*.json` trước khi xóa thư mục để dựng lại.

    Giữ `localSettings.json` để Power BI Desktop không hỏi lại quyền đọc file
    CSV sau mỗi lần chạy; người dùng chỉ việc mở và bấm Refresh.
    """
    saved: dict[Path, bytes] = {}
    for directory in directories:
        pbi_dir = directory / ".pbi"
        if not pbi_dir.is_dir():
            continue
        for name in PRESERVED_LOCAL_FILES:
            candidate = pbi_dir / name
            if candidate.is_file():
                saved[candidate] = candidate.read_bytes()
    return saved


def restore_local_settings(saved: dict[Path, bytes]) -> None:
    """Ghi lại các file cấu hình cục bộ sau khi dựng xong."""
    for path, payload in saved.items():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(payload)


def platform_payload(item_type: str, display_name: str) -> dict[str, Any]:
    logical_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"vikoda:{REPORT_NAME}:{item_type}"))
    return {
        "$schema": "https://developer.microsoft.com/json-schemas/fabric/gitIntegration/platformProperties/2.0.0/schema.json",
        "metadata": {"type": item_type, "displayName": display_name},
        "config": {"version": "2.0", "logicalId": logical_id},
    }


def safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"Gia tri khong hop le: {value!r}")
    return number


def parse_iso_date(value: Any) -> date:
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        raise ValueError("Thieu ngay hoa don trong staging")
    return datetime.fromisoformat(text[:10]).date()


def period_date(year: int, month: int) -> date:
    return date(year, month, 1)


def previous_month(year: int, month: int) -> tuple[int, int]:
    return (year - 1, 12) if month == 1 else (year, month - 1)


def month_label(year: int, month: int) -> str:
    return f"T{month:02d}/{year}"


def unique_name(values: Iterable[str], fallback: str) -> str:
    cleaned = sorted({str(value or "").strip() for value in values if str(value or "").strip()})
    if not cleaned:
        return fallback
    return max(cleaned, key=lambda value: (len(value), value))


def customer_key(code: Any, name: Any, area: Any, region: Any) -> str:
    """Keep special target rows with code 0 from collapsing into one customer."""
    code_text = str(code or "").strip()
    if code_text and normalize_key(code_text) not in {"0", "NONE", "NAN"}:
        return code_text
    name_text = normalize_key(name) or "KHONG RO"
    return f"0|{name_text}|{normalize_key(area)}|{normalize_key(region)}"


def product_key(code: Any, name: Any) -> str:
    code_text = str(code or "").strip()
    if code_text:
        return code_text
    return f"NAME|{normalize_key(name) or '-'}"


def tidy_label(value: Any) -> str:
    """Gộp khoảng trắng thừa trong nhãn hiển thị.

    DMSP có vài tên kiểu `Vikoda  Fusion Cam Ranh  RGB 450` với hai dấu cách
    liền nhau. Trên trục biểu đồ, mỗi khoảng trắng thừa lại đẩy tên tới ngưỡng
    bị cắt bằng dấu ba chấm, nên chuẩn hóa ngay từ lúc dựng chiều.
    """
    return " ".join(str(value or "").split())


def axis_label(short_name: str) -> str:
    """Nhãn rút gọn dành riêng cho trục biểu đồ.

    Bỏ tiền tố `Vikoda ` vì đây là báo cáo sell-in của chính Vikoda — tiền tố
    lặp trên 43/67 SKU chỉ ăn chỗ và đẩy tên tới ngưỡng bị cắt. Tên đầy đủ vẫn
    giữ nguyên ở `ProductShortName` cho các bảng Reporting.
    """
    trimmed = re.sub(r"^Vikoda\s+", "", tidy_label(short_name), flags=re.IGNORECASE)
    return trimmed or tidy_label(short_name)


def catalog_code(value: Any) -> str:
    """Normalize Excel numeric/text product codes to the staging key format."""
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def optional_positive_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) and parsed > 0 else None


def load_product_catalog(path: Path | None) -> dict[str, dict[str, Any]]:
    """Read DMSP packaging metadata without modifying the source workbook."""
    if path is None:
        return {}
    if not path.exists():
        raise FileNotFoundError(f"Khong tim thay DMSP: {path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - portable runtime supplies vendor package
        raise RuntimeError("Can openpyxl de doc DMSP") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["DanhMucSanPham"] if "DanhMucSanPham" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = {normalize_key(str(value or "")).replace(" ", "").replace("_", ""): index for index, value in enumerate(rows[0])}
    code_positions = [headers.get("MASANPHAMMOIVIKODA", 0), headers.get("MASANPHAMMOIVKD", 1)]
    name_position = headers.get("TENSANPHAM", 2)
    short_name_position = headers.get("TENTHUGON", 3)
    pack_size_position = headers.get("QUYCACH", 4)
    pack_unit_position = headers.get("DVT", 5)
    catalog: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        if not row:
            continue
        item = {
            "ProductCode": "",
            "ProductNameDMSP": tidy_label(row[name_position]),
            "ProductShortName": tidy_label(row[short_name_position]),
            "PackSize": optional_positive_float(row[pack_size_position]),
            "PackUnit": str(row[pack_unit_position] or "").strip(),
            "CoBrand": str(row[7] or "").strip() if len(row) > 7 else "",
            "ProductType": str(row[8] or "").strip() if len(row) > 8 else "",
            "PackagingType": str(row[9] or "").strip() if len(row) > 9 else "",
        }
        for position in code_positions:
            code = catalog_code(row[position] if position < len(row) else None)
            if code:
                item["ProductCode"] = code
                catalog[code] = dict(item)
    workbook.close()
    return catalog


def build_customer_map(dmkh_rows: list[list]) -> dict[str, dict[str, Any]]:
    customers: dict[str, dict[str, Any]] = {}
    for row in dmkh_rows:
        if not row:
            continue
        code = str(row[0] or "").strip()
        if not code:
            continue
        area, region = normalize_reporting_pair(row[7] if len(row) > 7 else "", row[8] if len(row) > 8 else "")
        item = customers.setdefault(code, {
            "CustomerCode": code,
            "CustomerName": str(row[1] or "").strip(),
            "CustomerNameFull": str(row[2] or row[1] or "").strip(),
            "Channel": str(row[4] or "").strip(),
            "CustomerType": str(row[5] or "").strip(),
            "SystemMT": str(row[6] or "").strip(),
            "Mien": area,
            "Vung": region,
            "Province": str(row[9] or "").strip() if len(row) > 9 else "",
            "District": str(row[10] or "").strip() if len(row) > 10 else "",
        })
        # Keep the most descriptive name when the source has multiple variants.
        item["CustomerName"] = unique_name(
            [item.get("CustomerName"), row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else ""],
            code,
        )
        if not item["CustomerNameFull"]:
            item["CustomerNameFull"] = item["CustomerName"]
    return customers


def build_dimensions(
    sell_in_payload: dict[str, Any],
    target_payload: dict[str, Any],
    dmkh_payload: dict[str, Any],
    product_catalog: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Return DimDate, DimCustomer, DimProduct, DimTerritory and fact rows."""
    current_year = int(sell_in_payload["current_year"])
    through_month = int(sell_in_payload["through_month"])
    as_of_date = parse_iso_date(sell_in_payload.get("as_of_date"))
    customer_map = build_customer_map(dmkh_payload.get("rows", []))
    product_catalog = product_catalog or {}

    sell_rows: list[dict[str, Any]] = []
    product_names: dict[str, set[str]] = {}
    customer_rows: dict[str, dict[str, Any]] = {}
    territory_lookup: dict[tuple[str, str], int] = {}
    for area_index, (area, regions) in enumerate(REPORTING_STRUCTURE, start=1):
        for region_index, region in enumerate(regions, start=1):
            territory_lookup[(area, region)] = area_index * 100 + region_index

    for row in sell_in_payload.get("rows", []):
        if len(row) < 14:
            continue
        invoice_date = parse_iso_date(row[2])
        reporting_year = int(row[13])
        reporting_month = int(row[12])
        reporting_date = period_date(reporting_year, reporting_month)
        code = str(row[3] or "").strip()
        name = str(row[4] or "").strip()
        product_code = str(row[5] or "").strip()
        product_name = str(row[6] or "").strip() or "-"
        mapped = customer_map.get(code)
        if mapped is None:
            area, region = "Other", "Other"
            channel = ""
            customer_type = ""
            system_mt = ""
        else:
            area, region = mapped["Mien"], mapped["Vung"]
            channel = mapped.get("Channel", "")
            customer_type = mapped.get("CustomerType", "")
            system_mt = mapped.get("SystemMT", "")
        ckey = customer_key(code, name, area, region)
        pkey = product_key(product_code, product_name)
        product_names.setdefault(pkey, set()).add(product_name)
        customer_rows.setdefault(ckey, {
            "CustomerKey": ckey,
            "CustomerCode": code or "0",
            "CustomerName": name or "Khong ro ten",
            "CustomerNameFull": name or "Khong ro ten",
            "Channel": channel,
            "CustomerType": customer_type,
            "SystemMT": system_mt,
            "Mien": area,
            "Vung": region,
            "Province": "",
            "District": "",
        })
        amount = safe_float(row[9])
        quantity = safe_float(row[7])
        catalog_item = product_catalog.get(product_code, {})
        pack_size = optional_positive_float(catalog_item.get("PackSize"))
        pack_unit = str(catalog_item.get("PackUnit") or "").strip()
        invoice_type = str(row[10] or "").strip()
        sell_rows.append({
            "Date": reporting_date.isoformat(),
            "InvoiceDate": invoice_date.isoformat(),
            "DateKey": reporting_date.strftime("%Y%m%d"),
            "PeriodKey": f"{reporting_year}{reporting_month:02d}",
            "Year": reporting_year,
            "MonthNumber": reporting_month,
            "CustomerKey": ckey,
            "CustomerCode": code or "0",
            "ProductKey": pkey,
            "ProductCode": product_code,
            "TerritoryKey": territory_lookup.get((area, region), territory_lookup[("Other", "Other")]),
            "Mien": area,
            "Vung": region,
            "Quantity": quantity,
            "ConvertedQuantity": quantity / pack_size if pack_size else None,
            "PackUnit": pack_unit,
            "RevenueVND": amount,
            "InvoiceType": invoice_type,
            "IsReturn": invoice_type == "Đơn trả hàng",
            "IsVikoda": "VIKODA" in normalize_key(product_name),
            "IsKDT": "KDT" in normalize_key(product_name),
            "ProductName": product_name,
            "CustomerName": name or "Khong ro ten",
        })

    # Add customers from target rows, including special code-0 rows not present in DMKH.
    target_rows: list[dict[str, Any]] = []
    for record in target_payload.get("records", []):
        try:
            year = int(record.get("Nam"))
            month = int(record.get("Thang"))
        except (TypeError, ValueError):
            key = str(record.get("Ky") or "")
            year, month = int(key[:4]), int(key[4:6])
        code = str(record.get("MaKhachHangMoi") or "").strip()
        name = str(record.get("TenKhachHang") or "").strip()
        area_hint, region_hint = record.get("MienBaoCao"), record.get("VungBaoCao")
        if normalize_key(code) == "B2C" or normalize_key(name) == "B2C":
            area, region = "B2C", "B2C"
        elif normalize_key(name) == "OTHER":
            area, region = "Other", "Other"
        elif area_hint or region_hint:
            area, region = normalize_reporting_pair(area_hint, region_hint)
        elif code in customer_map:
            area, region = customer_map[code]["Mien"], customer_map[code]["Vung"]
        else:
            area, region = "Other", "Other"
        ckey = customer_key(code, name, area, region)
        customer_rows.setdefault(ckey, {
            "CustomerKey": ckey,
            "CustomerCode": code or "0",
            "CustomerName": name or "Khong ro ten",
            "CustomerNameFull": name or "Khong ro ten",
            "Channel": customer_map.get(code, {}).get("Channel", ""),
            "CustomerType": customer_map.get(code, {}).get("CustomerType", ""),
            "SystemMT": customer_map.get(code, {}).get("SystemMT", ""),
            "Mien": area,
            "Vung": region,
            "Province": customer_map.get(code, {}).get("Province", ""),
            "District": customer_map.get(code, {}).get("District", ""),
        })
        target_rows.append({
            "Date": period_date(year, month).isoformat(),
            "DateKey": period_date(year, month).strftime("%Y%m%d"),
            "PeriodKey": f"{year}{month:02d}",
            "Year": year,
            "MonthNumber": month,
            "CustomerKey": ckey,
            "CustomerCode": code or "0",
            "TerritoryKey": territory_lookup.get((area, region), territory_lookup[("Other", "Other")]),
            "Mien": area,
            "Vung": region,
            "TargetTotalVND": safe_float(record.get("TargetTong")),
            "TargetVikodaVND": safe_float(record.get("TargetVikoda")),
            "SourceFile": str(record.get("NguonFile") or ""),
        })

    reporting_periods = [period_date(int(row[13]), int(row[12])) for row in sell_in_payload.get("rows", [])]
    min_date = min(reporting_periods or [period_date(current_year - 1, 1)])
    max_target_year = max([int(record.get("Nam")) for record in target_payload.get("records", []) if record.get("Nam")] or [current_year])
    max_date = date(max_target_year, 12, 31)
    dim_date: list[dict[str, Any]] = []
    current_period = f"{current_year}{through_month:02d}"
    last_year_period = f"{current_year - 1}{through_month:02d}"
    prior_year, prior_month = previous_month(current_year, through_month)
    prior_period = f"{prior_year}{prior_month:02d}"
    cursor = min_date
    while cursor <= max_date:
        period = cursor.strftime("%Y%m")
        month_start = cursor.replace(day=1)
        dim_date.append({
            "Date": cursor.isoformat(),
            "DateKey": cursor.strftime("%Y%m%d"),
            "PeriodKey": period,
            "Year": cursor.year,
            "Quarter": f"Q{((cursor.month - 1) // 3) + 1}",
            "MonthNumber": cursor.month,
            "MonthName": f"Tháng {cursor.month}",
            "MonthLabel": month_label(cursor.year, cursor.month),
            # Short year-first labels stay readable on compact executive charts
            # and remain chronologically sortable without sortByColumn metadata.
            "MonthAxis": f"{cursor.year % 100:02d}/{cursor.month:02d}",
            "MonthStart": month_start.isoformat(),
            "IsCurrentYTD": cursor.year == current_year and cursor.month <= through_month,
            "IsPriorYTD": cursor.year == current_year - 1 and cursor.month <= through_month,
            "IsCurrentMonth": period == current_period,
            "IsLastYearMonth": period == last_year_period,
            "IsPriorMonth": period == prior_period,
        })
        cursor += timedelta(days=1)

    dim_customer = sorted(customer_rows.values(), key=lambda item: (item["Mien"], item["Vung"], item["CustomerName"], item["CustomerKey"]))
    for row in dim_customer:
        row["CustomerName"] = unique_name([row.get("CustomerName"), row.get("CustomerNameFull")], row["CustomerCode"])
    dim_product: list[dict[str, Any]] = []
    for key, names in sorted(product_names.items()):
        product_code = key if not key.startswith("NAME|") else ""
        source_name = unique_name(names, "-")
        catalog_item = product_catalog.get(product_code, {})
        catalog_name = str(catalog_item.get("ProductNameDMSP") or "").strip()
        display_name = catalog_name or source_name
        dim_product.append({
            "ProductKey": key,
            "ProductCode": product_code,
            "ProductName": display_name,
            "ProductShortName": tidy_label(catalog_item.get("ProductShortName") or display_name),
            "ProductAxisLabel": axis_label(catalog_item.get("ProductShortName") or display_name),
            "ProductGroup": "Vikoda" if "VIKODA" in normalize_key(display_name) else ("KDT" if "KDT" in normalize_key(display_name) else "Khác"),
            "PackSize": optional_positive_float(catalog_item.get("PackSize")),
            "PackUnit": str(catalog_item.get("PackUnit") or "").strip(),
            "CoBrand": str(catalog_item.get("CoBrand") or "").strip(),
            "ProductType": str(catalog_item.get("ProductType") or "").strip(),
            "PackagingType": str(catalog_item.get("PackagingType") or "").strip(),
            "IsVikoda": "VIKODA" in normalize_key(display_name),
            "IsKDT": "KDT" in normalize_key(display_name),
        })
    dim_territory: list[dict[str, Any]] = []
    for area_index, (area, regions) in enumerate(REPORTING_STRUCTURE, start=1):
        for region_index, region in enumerate(regions, start=1):
            dim_territory.append({
                "TerritoryKey": area_index * 100 + region_index,
                "Mien": area,
                "Vung": region,
                "MienSort": area_index,
                "VungSort": region_index,
            })
    return dim_date, dim_customer, dim_product, dim_territory, (sell_rows, target_rows)


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in columns})


def quote_m(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def m_partition(file_path: str, columns: list[tuple[str, str]]) -> str:
    type_map = {
        "string": "type text",
        "int64": "Int64.Type",
        "double": "type number",
        "dateTime": "type date",
        "boolean": "type logical",
    }
    path = file_path
    type_list = ", ".join(f"{{\"{name}\", {type_map[dtype]}}}" for name, dtype in columns)
    return (
        "let\n"
        f"    Source = Csv.Document(File.Contents({quote_m(path)}), [Delimiter=\",\", Encoding=65001, QuoteStyle=QuoteStyle.Csv]),\n"
        "    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars=true]),\n"
        # CSV numeric fields always use a dot decimal separator.  Pin the
        # conversion culture so a vi-VN Power BI Desktop session does not
        # interpret a value such as ``46189020042.0`` as 461,890,200,420.
        f"    Typed = Table.TransformColumnTypes(PromotedHeaders, {{{type_list}}}, \"en-US\")\n"
        "in\n"
        "    Typed"
    )


def tmsl_column(name: str, dtype: str, *, key: bool = False, format_string: str | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "name": name,
        "dataType": dtype,
        "sourceColumn": name,
        "summarizeBy": "none" if dtype in {"string", "dateTime", "boolean"} else "sum",
    }
    if key:
        result["isKey"] = True
    if format_string:
        result["formatString"] = format_string
    return result


def measure(name: str, expression: str, fmt: str = "#,##0;(#,##0);-") -> dict[str, Any]:
    return {"name": name, "expression": expression, "formatString": fmt, "displayFolder": "KPI"}


def build_measures() -> list[dict[str, Any]]:
    has_date_filter = "(ISCROSSFILTERED(DimDate[Date]) || ISFILTERED(DimDate))"
    latest_period = "CALCULATE(MAX(FactSellIn[PeriodKey]), REMOVEFILTERS(DimDate), REMOVEFILTERS(FactSellIn))"
    latest_date = "CALCULATE(MAX(FactSellIn[InvoiceDate]), REMOVEFILTERS(DimDate), REMOVEFILTERS(FactSellIn))"
    actual_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN DIVIDE(IF(HasDateFilter, SUM(FactSellIn[RevenueVND]), CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod)), 1000000)"
    target_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN DIVIDE(IF(HasDateFilter, SUM(FactTarget[TargetTotalVND]), CALCULATE(SUM(FactTarget[TargetTotalVND]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod)), 1000000)"
    target_vikoda_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN DIVIDE(IF(HasDateFilter, SUM(FactTarget[TargetVikodaVND]), CALCULATE(SUM(FactTarget[TargetVikodaVND]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod)), 1000000)"
    quantity_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN IF(HasDateFilter, SUM(FactSellIn[Quantity]), CALCULATE(SUM(FactSellIn[Quantity]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod))"
    customers_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN IF(HasDateFilter, DISTINCTCOUNT(FactSellIn[CustomerKey]), CALCULATE(DISTINCTCOUNT(FactSellIn[CustomerKey]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod))"
    converted_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN IF(HasDateFilter, SUM(FactSellIn[ConvertedQuantity]), CALCULATE(SUM(FactSellIn[ConvertedQuantity]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod))"
    unconverted_expression = f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} RETURN IF(HasDateFilter, CALCULATE(SUM(FactSellIn[Quantity]), FactSellIn[ConvertedQuantity] = BLANK()), CALCULATE(SUM(FactSellIn[Quantity]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod, FactSellIn[ConvertedQuantity] = BLANK()))"

    def prior_unit_expression(unit: str) -> str:
        unit_literal = unit.replace('"', '""')
        return (
            f'VAR HasDateFilter = {has_date_filter} VAR MaxDataDate = {latest_date} '
            f'VAR LastYearPeriod = FORMAT(DATE(YEAR(MaxDataDate) - 1, MONTH(MaxDataDate), 1), "yyyyMM") '
            f'RETURN IF(HasDateFilter, '
            f'CALCULATE(SUM(FactSellIn[ConvertedQuantity]), KEEPFILTERS(FactSellIn[PackUnit] = "{unit_literal}"), SAMEPERIODLASTYEAR(DimDate[Date])), '
            f'CALCULATE(SUM(FactSellIn[ConvertedQuantity]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LastYearPeriod, KEEPFILTERS(FactSellIn[PackUnit] = "{unit_literal}")))'
        )

    cases_ly = prior_unit_expression("Két")
    cartons_ly = prior_unit_expression("Thùng")
    bottles_ly = prior_unit_expression("Bình")

    sku_expression = (
        f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} "
        "RETURN IF(HasDateFilter, DISTINCTCOUNT(FactSellIn[ProductKey]), "
        "CALCULATE(DISTINCTCOUNT(FactSellIn[ProductKey]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod))"
    )
    customers_ly_expression = (
        f"VAR HasDateFilter = {has_date_filter} VAR MaxDataDate = {latest_date} "
        'VAR LastYearPeriod = FORMAT(DATE(YEAR(MaxDataDate) - 1, MONTH(MaxDataDate), 1), "yyyyMM") '
        "RETURN IF(HasDateFilter, "
        "CALCULATE(DISTINCTCOUNT(FactSellIn[CustomerKey]), SAMEPERIODLASTYEAR(DimDate[Date])), "
        "CALCULATE(DISTINCTCOUNT(FactSellIn[CustomerKey]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LastYearPeriod))"
    )

    def customer_movement_expression(direction: str) -> str:
        """Đếm khách hàng mới / ngừng mua so với cùng kỳ năm trước.

        DAX không cho `IF` trả về bảng, nên cả hai nhánh (có lọc ngày và mặc
        định về kỳ mới nhất) được tính sẵn thành scalar rồi mới chọn.
        """
        pair = "FilteredCurrent, FilteredPrior" if direction == "new" else "FilteredPrior, FilteredCurrent"
        default_pair = "DefaultCurrent, DefaultPrior" if direction == "new" else "DefaultPrior, DefaultCurrent"
        return (
            f"VAR HasDateFilter = {has_date_filter} VAR LatestPeriod = {latest_period} VAR MaxDataDate = {latest_date} "
            'VAR LastYearPeriod = FORMAT(DATE(YEAR(MaxDataDate) - 1, MONTH(MaxDataDate), 1), "yyyyMM") '
            "VAR FilteredCurrent = VALUES(FactSellIn[CustomerKey]) "
            "VAR FilteredPrior = CALCULATETABLE(VALUES(FactSellIn[CustomerKey]), SAMEPERIODLASTYEAR(DimDate[Date])) "
            "VAR DefaultCurrent = CALCULATETABLE(VALUES(FactSellIn[CustomerKey]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LatestPeriod) "
            "VAR DefaultPrior = CALCULATETABLE(VALUES(FactSellIn[CustomerKey]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LastYearPeriod) "
            f"VAR FilteredResult = COUNTROWS(EXCEPT({pair})) "
            f"VAR DefaultResult = COUNTROWS(EXCEPT({default_pair})) "
            "RETURN COALESCE(IF(HasDateFilter, FilteredResult, DefaultResult), 0)"
        )
    seasonality_forecast = (
        f"VAR HasDateFilter = {has_date_filter} VAR GlobalMaxDate = {latest_date} "
        "VAR MaxDataDate = IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), GlobalMaxDate) "
        "VAR MonthStart = DATE(YEAR(MaxDataDate), MONTH(MaxDataDate), 1) "
        "VAR ActualMTD = DIVIDE(CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DATESBETWEEN(DimDate[Date], MonthStart, MaxDataDate)), 1000000) "
        "VAR PriorMonthStart = EDATE(MonthStart, -12) "
        "VAR PriorMonthEnd = EOMONTH(PriorMonthStart, 0) "
        "VAR PriorSamePoint = EDATE(MaxDataDate, -12) "
        "VAR PriorTotal = DIVIDE(CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DATESBETWEEN(DimDate[Date], PriorMonthStart, PriorMonthEnd)), 1000000) "
        "VAR PriorMTD = DIVIDE(CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DATESBETWEEN(DimDate[Date], PriorMonthStart, PriorSamePoint)), 1000000) "
        "VAR PriorRemaining = MAX(0, PriorTotal - PriorMTD) "
        "VAR YoYGrowth = [Tăng trưởng YoY] "
        "VAR ExpectedRemaining = PriorRemaining * (1 + IF(ISBLANK(YoYGrowth), 0, YoYGrowth)) "
        "RETURN IF(PriorRemaining > 0, ActualMTD + ExpectedRemaining, [Run-rate dự báo tháng])"
    )

    return [
        measure("Doanh thu Sell In", actual_expression, MONEY_FORMAT),
        measure("Doanh thu bán", "CALCULATE([Doanh thu Sell In], FactSellIn[IsReturn] = FALSE())", MONEY_FORMAT),
        measure("Doanh thu trả hàng", "CALCULATE([Doanh thu Sell In], FactSellIn[IsReturn] = TRUE())", MONEY_FORMAT),
        measure("Target", target_expression, MONEY_FORMAT),
        measure("Target Vikoda", target_vikoda_expression, MONEY_FORMAT),
        measure("Khoảng cách Target", "[Doanh thu Sell In] - [Target]", MONEY_FORMAT),
        measure("Tỷ lệ đạt Target", "DIVIDE([Doanh thu Sell In], [Target])", "0.0%;(0.0%);-"),
        measure("Doanh thu LY", f"VAR HasDateFilter = {has_date_filter} VAR MaxDataDate = {latest_date} VAR LastYearPeriod = FORMAT(DATE(YEAR(MaxDataDate) - 1, MONTH(MaxDataDate), 1), \"yyyyMM\") RETURN DIVIDE(IF(HasDateFilter, CALCULATE(SUM(FactSellIn[RevenueVND]), SAMEPERIODLASTYEAR(DimDate[Date])), CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = LastYearPeriod)), 1000000)", MONEY_FORMAT),
        measure("Tăng trưởng YoY", "DIVIDE([Doanh thu Sell In] - [Doanh thu LY], [Doanh thu LY])", GROWTH_FORMAT),
        measure("Doanh thu tháng trước", f"VAR HasDateFilter = {has_date_filter} VAR MaxDataDate = {latest_date} VAR PreviousPeriod = FORMAT(EOMONTH(MaxDataDate, -1), \"yyyyMM\") RETURN DIVIDE(IF(HasDateFilter, CALCULATE(SUM(FactSellIn[RevenueVND]), DATEADD(DimDate[Date], -1, MONTH)), CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DimDate[PeriodKey] = PreviousPeriod)), 1000000)", MONEY_FORMAT),
        measure("Tăng trưởng MoM", "DIVIDE([Doanh thu Sell In] - [Doanh thu tháng trước], [Doanh thu tháng trước])", GROWTH_FORMAT),
        measure("Doanh thu Vikoda", "CALCULATE([Doanh thu Sell In], KEEPFILTERS(DimProduct[IsVikoda] = TRUE()))", MONEY_FORMAT),
        measure("Đạt Target Vikoda", "DIVIDE([Doanh thu Vikoda], [Target Vikoda])", "0.0%;(0.0%);-"),
        measure("Khoảng cách Vikoda", "[Doanh thu Vikoda] - [Target Vikoda]", MONEY_FORMAT),
        measure("Doanh thu KDT", "CALCULATE([Doanh thu Sell In], KEEPFILTERS(DimProduct[IsKDT] = TRUE()))", MONEY_FORMAT),
        measure("Sản lượng", quantity_expression, QUANTITY_FORMAT),
        measure("Số khách hàng", customers_expression, "#,##0;(#,##0);-"),
        measure("Ngày dữ liệu mới nhất", f"VAR HasDateFilter = {has_date_filter} VAR MaxDataDate = {latest_date} RETURN IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), MaxDataDate)", "dd/mm/yyyy"),
        # --- Chỉ số Nhịp độ & Dự báo Mùa vụ ngành Nước giải khát ---
        measure("% Thời gian tháng đã qua", f"VAR HasDateFilter = {has_date_filter} VAR GlobalMaxDate = {latest_date} VAR MaxDataDate = IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), GlobalMaxDate) VAR DaysElapsed = DAY(MaxDataDate) VAR DaysInMonth = DAY(EOMONTH(MaxDataDate, 0)) RETURN DIVIDE(DaysElapsed, DaysInMonth)", "0.0%"),
        measure("Nhịp độ bán hàng", "DIVIDE([Tỷ lệ đạt Target], [% Thời gian tháng đã qua])", "0.0%;(0.0%);-"),
        measure("Doanh thu ngày bình quân MTD", f"VAR HasDateFilter = {has_date_filter} VAR GlobalMaxDate = {latest_date} VAR MaxDataDate = IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), GlobalMaxDate) VAR DaysElapsed = DAY(MaxDataDate) VAR MonthStart = DATE(YEAR(MaxDataDate), MONTH(MaxDataDate), 1) VAR ActualMTD = DIVIDE(CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DATESBETWEEN(DimDate[Date], MonthStart, MaxDataDate)), 1000000) RETURN DIVIDE(ActualMTD, IF(DaysElapsed > 0, DaysElapsed, 1))", MONEY_FORMAT),
        measure("Hệ số gia tốc cần đạt", "VAR DailyActual = [Doanh thu ngày bình quân MTD] VAR DailyRequired = [Cần doanh thu mỗi ngày] RETURN IF(DailyActual > 0, DIVIDE(DailyRequired, DailyActual), BLANK())", "0.0x"),
        measure("Run-rate dự báo tháng", f"VAR HasDateFilter = {has_date_filter} VAR GlobalMaxDate = {latest_date} VAR MaxDataDate = IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), GlobalMaxDate) VAR DaysElapsed = DAY(MaxDataDate) VAR DaysInMonth = DAY(EOMONTH(MaxDataDate, 0)) VAR MonthStart = DATE(YEAR(MaxDataDate), MONTH(MaxDataDate), 1) VAR ActualMTD = DIVIDE(CALCULATE(SUM(FactSellIn[RevenueVND]), REMOVEFILTERS(DimDate), DATESBETWEEN(DimDate[Date], MonthStart, MaxDataDate)), 1000000) RETURN DIVIDE(ActualMTD, DaysElapsed) * DaysInMonth", MONEY_FORMAT),
        measure("Dự báo EOM Mùa vụ", seasonality_forecast, MONEY_FORMAT),
        measure("Dự báo đạt Target", "DIVIDE([Run-rate dự báo tháng], [Target])", "0.0%;(0.0%);-"),
        measure("Còn thiếu để đạt Target", "MAX(0, [Target] - [Doanh thu Sell In])", MONEY_FORMAT),
        measure("Cần doanh thu mỗi ngày", f"VAR HasDateFilter = {has_date_filter} VAR GlobalMaxDate = {latest_date} VAR MaxDataDate = IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), GlobalMaxDate) VAR DaysLeft = DAY(EOMONTH(MaxDataDate, 0)) - DAY(MaxDataDate) RETURN DIVIDE([Còn thiếu để đạt Target], DaysLeft)", MONEY_FORMAT),
        measure("Target 3 tháng tới", f"VAR HasDateFilter = {has_date_filter} VAR GlobalMaxDate = {latest_date} VAR MaxDataDate = IF(HasDateFilter, MAX(FactSellIn[InvoiceDate]), GlobalMaxDate) RETURN CALCULATE([Target], DATESBETWEEN(DimDate[Date], EOMONTH(MaxDataDate, 0) + 1, EOMONTH(MaxDataDate, 3)))", MONEY_FORMAT),
        measure("Tỷ lệ trả hàng", "COALESCE(DIVIDE(ABS([Doanh thu trả hàng]), ABS([Doanh thu bán])), 0)", "0.0%;(0.0%);0.0%"),
        measure("Tỷ trọng Vikoda", "DIVIDE([Doanh thu Vikoda], [Doanh thu Sell In])", "0.0%;(0.0%);-"),
        measure("Xếp hạng vùng", "RANKX(ALL(DimTerritory[Vung]), [Doanh thu Sell In],, DESC, Dense)", "#,##0"),
        measure("Lũy kế Pareto", "VAR CurrentRank = [Xếp hạng vùng] VAR TotalRevenue = CALCULATE([Doanh thu Sell In], ALL(DimTerritory[Vung])) RETURN DIVIDE(CALCULATE([Doanh thu Sell In], FILTER(ALL(DimTerritory[Vung]), [Xếp hạng vùng] <= CurrentRank)), TotalRevenue)", "0.0%;(0.0%);-"),
        # --- Cơ cấu bao bì Két / Thùng / Bình ---
        measure("Sản lượng quy đổi KTB", converted_expression, QUANTITY_FORMAT),
        measure("SL Két (K)", "CALCULATE([Sản lượng quy đổi KTB], KEEPFILTERS(DimProduct[PackUnit] = \"Két\"))", QUANTITY_FORMAT),
        measure("SL Thùng (T)", "CALCULATE([Sản lượng quy đổi KTB], KEEPFILTERS(DimProduct[PackUnit] = \"Thùng\"))", QUANTITY_FORMAT),
        measure("SL Bình (B)", "CALCULATE([Sản lượng quy đổi KTB], KEEPFILTERS(DimProduct[PackUnit] = \"Bình\"))", QUANTITY_FORMAT),
        measure("Tỷ trọng Két %", "DIVIDE([SL Két (K)], [Sản lượng quy đổi KTB])", "0.0%"),
        measure("Tỷ trọng Thùng %", "DIVIDE([SL Thùng (T)], [Sản lượng quy đổi KTB])", "0.0%"),
        measure("Tỷ trọng Bình %", "DIVIDE([SL Bình (B)], [Sản lượng quy đổi KTB])", "0.0%"),
        measure("SL Két (K) LY", cases_ly, QUANTITY_FORMAT),
        measure("SL Thùng (T) LY", cartons_ly, QUANTITY_FORMAT),
        measure("SL Bình (B) LY", bottles_ly, QUANTITY_FORMAT),
        measure("Tăng trưởng Két (K)", "DIVIDE([SL Két (K)] - [SL Két (K) LY], [SL Két (K) LY])", GROWTH_FORMAT),
        measure("Tăng trưởng Thùng (T)", "DIVIDE([SL Thùng (T)] - [SL Thùng (T) LY], [SL Thùng (T) LY])", GROWTH_FORMAT),
        measure("Tăng trưởng Bình (B)", "DIVIDE([SL Bình (B)] - [SL Bình (B) LY], [SL Bình (B) LY])", GROWTH_FORMAT),
        measure("SL chưa quy đổi", unconverted_expression, QUANTITY_FORMAT),
        # --- Vikoda vs KDT: tách riêng cùng kỳ và tăng trưởng cho từng nhánh hàng ---
        measure("Doanh thu Vikoda LY", "CALCULATE([Doanh thu LY], KEEPFILTERS(DimProduct[IsVikoda] = TRUE()))", MONEY_FORMAT),
        measure("Tăng trưởng Vikoda", "DIVIDE([Doanh thu Vikoda] - [Doanh thu Vikoda LY], [Doanh thu Vikoda LY])", GROWTH_FORMAT),
        measure("Doanh thu KDT LY", "CALCULATE([Doanh thu LY], KEEPFILTERS(DimProduct[IsKDT] = TRUE()))", MONEY_FORMAT),
        measure("Tăng trưởng KDT", "DIVIDE([Doanh thu KDT] - [Doanh thu KDT LY], [Doanh thu KDT LY])", GROWTH_FORMAT),
        measure("Tỷ trọng KDT", "DIVIDE([Doanh thu KDT], [Doanh thu Sell In])", "0.0%;(0.0%);-"),
        # --- Độ phủ danh mục và tệp khách hàng ---
        measure("Số SKU", sku_expression, "#,##0;(#,##0);-"),
        measure("Số khách hàng LY", customers_ly_expression, "#,##0;(#,##0);-"),
        measure("Tăng trưởng số khách hàng", "DIVIDE([Số khách hàng] - [Số khách hàng LY], [Số khách hàng LY])", GROWTH_FORMAT),
        measure("Doanh thu bình quân KH", "DIVIDE([Doanh thu Sell In], [Số khách hàng])", MONEY_FORMAT),
        measure("Doanh thu / Điểm bán Active", "DIVIDE([Doanh thu Sell In], [Số khách hàng])", MONEY_FORMAT),
        measure("Khách hàng mới", customer_movement_expression("new"), "#,##0;(#,##0);-"),
        measure("Khách hàng ngừng mua", customer_movement_expression("churn"), "#,##0;(#,##0);-"),
        # --- Xếp hạng và đóng góp, dùng cho bảng Top/Bottom ---
        measure("Doanh thu hệ thống MT", 'CALCULATE([Doanh thu Sell In], KEEPFILTERS(NOT ISBLANK(DimCustomer[SystemMT])), KEEPFILTERS(DimCustomer[SystemMT] <> ""))', MONEY_FORMAT),
        measure("Xếp hạng SKU", "RANKX(ALL(DimProduct[ProductShortName]), [Doanh thu Sell In],, DESC, Dense)", "#,##0"),
        measure("Đóng góp doanh thu", "DIVIDE([Doanh thu Sell In], CALCULATE([Doanh thu Sell In], ALLSELECTED()))", "0.0%;(0.0%);-"),
    ]


def model_definition(csv_dir: Path, row_counts: dict[str, int]) -> dict[str, Any]:
    files = {
        "DimDate": "DimDate.csv",
        "DimCustomer": "DimCustomer.csv",
        "DimProduct": "DimProduct.csv",
        "DimTerritory": "DimTerritory.csv",
        "FactSellIn": "FactSellIn.csv",
        "FactTarget": "FactTarget.csv",
    }
    schemas: dict[str, list[tuple[str, str]]] = {
        "DimDate": [("Date", "dateTime"), ("DateKey", "string"), ("PeriodKey", "string"), ("Year", "int64"), ("Quarter", "string"), ("MonthNumber", "int64"), ("MonthName", "string"), ("MonthLabel", "string"), ("MonthAxis", "string"), ("MonthStart", "dateTime"), ("IsCurrentYTD", "boolean"), ("IsPriorYTD", "boolean"), ("IsCurrentMonth", "boolean"), ("IsLastYearMonth", "boolean"), ("IsPriorMonth", "boolean")],
        "DimCustomer": [("CustomerKey", "string"), ("CustomerCode", "string"), ("CustomerName", "string"), ("CustomerNameFull", "string"), ("Channel", "string"), ("CustomerType", "string"), ("SystemMT", "string"), ("Mien", "string"), ("Vung", "string"), ("Province", "string"), ("District", "string")],
        "DimProduct": [("ProductKey", "string"), ("ProductCode", "string"), ("ProductName", "string"), ("ProductShortName", "string"), ("ProductAxisLabel", "string"), ("ProductGroup", "string"), ("PackSize", "double"), ("PackUnit", "string"), ("CoBrand", "string"), ("ProductType", "string"), ("PackagingType", "string"), ("IsVikoda", "boolean"), ("IsKDT", "boolean")],
        "DimTerritory": [("TerritoryKey", "int64"), ("Mien", "string"), ("Vung", "string"), ("MienSort", "int64"), ("VungSort", "int64")],
        "FactSellIn": [("Date", "dateTime"), ("InvoiceDate", "dateTime"), ("DateKey", "string"), ("PeriodKey", "string"), ("Year", "int64"), ("MonthNumber", "int64"), ("CustomerKey", "string"), ("CustomerCode", "string"), ("ProductKey", "string"), ("ProductCode", "string"), ("TerritoryKey", "int64"), ("Mien", "string"), ("Vung", "string"), ("Quantity", "double"), ("ConvertedQuantity", "double"), ("PackUnit", "string"), ("RevenueVND", "double"), ("InvoiceType", "string"), ("IsReturn", "boolean"), ("IsVikoda", "boolean"), ("IsKDT", "boolean"), ("ProductName", "string"), ("CustomerName", "string")],
        "FactTarget": [("Date", "dateTime"), ("DateKey", "string"), ("PeriodKey", "string"), ("Year", "int64"), ("MonthNumber", "int64"), ("CustomerKey", "string"), ("CustomerCode", "string"), ("TerritoryKey", "int64"), ("Mien", "string"), ("Vung", "string"), ("TargetTotalVND", "double"), ("TargetVikodaVND", "double"), ("SourceFile", "string")],
    }
    tables: list[dict[str, Any]] = []
    for table_name, columns in schemas.items():
        table: dict[str, Any] = {
            "name": table_name,
            "columns": [
                tmsl_column(
                    name,
                    {"string": "string", "int64": "int64", "double": "double", "dateTime": "dateTime", "boolean": "boolean"}[dtype],
                    key=(name in {"Date", "CustomerKey", "ProductKey", "TerritoryKey"} and table_name.startswith("Dim")),
                    format_string=("yyyy-mm-dd" if name == "Date" else "yy/MM" if name == "MonthStart" else None),
                )
                for name, dtype in columns
            ],
            "partitions": [{
                "name": table_name,
                "mode": "import",
                "source": {"type": "m", "expression": m_partition(str((csv_dir / files[table_name]).resolve()), columns)},
            }],
        }
        if table_name == "DimDate":
            table["lineageTag"] = "Vikoda-DimDate"
        if table_name == "FactSellIn":
            table["measures"] = build_measures()
        tables.append(table)

    relationships = [
        {"name": "DimDate-Date-FactSellIn", "fromTable": "FactSellIn", "fromColumn": "Date", "toTable": "DimDate", "toColumn": "Date", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
        {"name": "DimDate-Date-FactTarget", "fromTable": "FactTarget", "fromColumn": "Date", "toTable": "DimDate", "toColumn": "Date", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
        {"name": "DimCustomer-Key-FactSellIn", "fromTable": "FactSellIn", "fromColumn": "CustomerKey", "toTable": "DimCustomer", "toColumn": "CustomerKey", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
        {"name": "DimCustomer-Key-FactTarget", "fromTable": "FactTarget", "fromColumn": "CustomerKey", "toTable": "DimCustomer", "toColumn": "CustomerKey", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
        {"name": "DimProduct-Key-FactSellIn", "fromTable": "FactSellIn", "fromColumn": "ProductKey", "toTable": "DimProduct", "toColumn": "ProductKey", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
        {"name": "DimTerritory-Key-FactSellIn", "fromTable": "FactSellIn", "fromColumn": "TerritoryKey", "toTable": "DimTerritory", "toColumn": "TerritoryKey", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
        {"name": "DimTerritory-Key-FactTarget", "fromTable": "FactTarget", "fromColumn": "TerritoryKey", "toTable": "DimTerritory", "toColumn": "TerritoryKey", "crossFilteringBehavior": "oneDirection", "fromCardinality": "many", "toCardinality": "one", "isActive": True},
    ]
    return {
        "name": REPORT_NAME,
        "compatibilityLevel": 1600,
        "model": {
            "culture": "vi-VN",
            "defaultPowerBIDataSourceVersion": "powerBI_V3",
            "sourceQueryCulture": "vi-VN",
            "discourageImplicitMeasures": True,
            "tables": tables,
            "relationships": relationships,
            "annotations": [
                {"name": "__PBI_TimeIntelligenceEnabled", "value": "0"},
                {"name": "PBIDesktopVersion", "value": "2.0.0"},
                {"name": "VikodaPackageVersion", "value": PACKAGE_VERSION},
            ],
        },
    }


def field_column(entity: str, prop: str) -> dict[str, Any]:
    return {"Column": {"Expression": {"SourceRef": {"Entity": entity}}, "Property": prop}}


def field_measure(entity: str, prop: str) -> dict[str, Any]:
    return {"Measure": {"Expression": {"SourceRef": {"Entity": entity}}, "Property": prop}}


def projection(field: dict[str, Any], query_ref: str, display_name: str | None = None) -> dict[str, Any]:
    item: dict[str, Any] = {"field": field, "queryRef": query_ref, "nativeQueryRef": display_name or query_ref.split(".")[-1]}
    if display_name:
        item["displayName"] = display_name
    return item


def literal(value: str) -> dict[str, Any]:
    return {"expr": {"Literal": {"Value": value}}}


def solid_color(color: str) -> dict[str, Any]:
    return {"solid": {"color": literal(quote_dax_literal(color))}}


def container_objects(title: str) -> dict[str, Any]:
    return {
        "title": [{"properties": {
            "show": literal("true"),
            "text": literal(quote_dax_literal(title)),
            "fontSize": literal("11D"),
            "bold": literal("true"),
            "fontColor": solid_color("#172033"),
            "fontFamily": literal("'Segoe UI Semibold'"),
            "titleWrap": literal("true"),
            "alignment": literal("'left'"),
        }}],
        "background": [{"properties": {"show": literal("true"), "color": solid_color("#FFFFFF"), "transparency": literal("0D")}}],
        "border": [{"properties": {"show": literal("true"), "color": solid_color("#DCE3ED"), "radius": literal("10D"), "width": literal("1D")}}],
        "padding": [{"properties": {"top": literal("8D"), "bottom": literal("8D"), "left": literal("8D"), "right": literal("8D")}}],
        "dropShadow": [{"properties": {"show": literal("true"), "angle": literal("90L"), "shadowDistance": literal("2L"), "shadowBlur": literal("8L"), "shadowSpread": literal("0L"), "transparency": literal("86L"), "color": solid_color("#39506A")}}],
        "visualHeader": [{"properties": {"show": literal("false")}}],
    }


def visual_file(name: str, visual_type: str, x: int, y: int, width: int, height: int, query_state: dict[str, Any] | None = None, title: str | None = None, tab_order: int = 0) -> dict[str, Any]:
    visual: dict[str, Any] = {"visualType": visual_type}
    if query_state:
        visual["query"] = {"queryState": query_state}
    if title:
        visual["visualContainerObjects"] = container_objects(title)
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "height": height, "width": width, "tabOrder": tab_order},
        "visual": visual,
    }


def quote_dax_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def card(name: str, measure_name: str, title: str, x: int, y: int, width: int = 240, height: int = 115, tab_order: int = 0) -> dict[str, Any]:
    result = visual_file(name, "card", x, y, width, height, {
        "Values": {"projections": [projection(field_measure("FactSellIn", measure_name), f"FactSellIn.{measure_name}")]}
    }, title, tab_order)
    result["visual"]["objects"] = {
        "labels": [{"properties": {"color": solid_color("#1F4E79"), "fontSize": literal("25D"), "bold": literal("true"), "fontFamily": literal("'Segoe UI Semibold'")}}],
        "categoryLabels": [{"properties": {"show": literal("false")}}],
    }
    return result


def slicer(name: str, entity: str, column: str, title: str, x: int, y: int, width: int, height: int, tab_order: int) -> dict[str, Any]:
    return visual_file(name, "slicer", x, y, width, height, {
        "Values": {"projections": [projection(field_column(entity, column), f"{entity}.{column}")]}
    }, title, tab_order)


def line_chart(name: str, category_entity: str, category: str, measures: list[str], title: str, x: int, y: int, width: int, height: int, tab_order: int) -> dict[str, Any]:
    return visual_file(name, "lineChart", x, y, width, height, {
        "Category": {"projections": [projection(field_column(category_entity, category), f"{category_entity}.{category}")]},
        "Y": {"projections": [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}") for item in measures]},
    }, title, tab_order)


def bar_chart(name: str, category_entity: str, category: str, measures: list[str], title: str, x: int, y: int, width: int, height: int, tab_order: int, horizontal: bool = False) -> dict[str, Any]:
    return visual_file(name, "barChart" if horizontal else "columnChart", x, y, width, height, {
        "Category": {"projections": [projection(field_column(category_entity, category), f"{category_entity}.{category}")]},
        "Y": {"projections": [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}") for item in measures]},
    }, title, tab_order)


def table_visual(name: str, fields: list[tuple[str, str, str]], measures: list[str], title: str, x: int, y: int, width: int, height: int, tab_order: int) -> dict[str, Any]:
    projections = [projection(field_column(entity, field), f"{entity}.{field}", display) for entity, field, display in fields]
    projections.extend(projection(field_measure("FactSellIn", item), f"FactSellIn.{item}") for item in measures)
    return visual_file(name, "tableEx", x, y, width, height, {"Values": {"projections": projections}}, title, tab_order)


def page_definition(page_name: str, display_name: str, visuals: list[dict[str, Any]], width: int = 1600, height: int = 900) -> tuple[dict[str, Any], dict[str, Any]]:
    page = {
        "$schema": f"{REPORT_SCHEMA}/page/2.0.0/schema.json",
        "name": page_name,
        "displayName": display_name,
        "displayOption": "FitToPage",
        "width": width,
        "height": height,
        "objects": {"background": [{"properties": {"color": solid_color("#F3F6FA"), "transparency": literal("0D")}}]},
    }
    pages = {"$schema": f"{REPORT_SCHEMA}/pagesMetadata/1.0.0/schema.json"}
    return page, {"visuals": visuals, "pages": pages}


def build_pages(report_definition_dir: Path) -> list[str]:
    page_specs: OrderedDict[str, tuple[str, list[dict[str, Any]]]] = OrderedDict()
    page_specs["CEO_TongQuan"] = ("CEO | Tổng quan", [
        slicer("slicer_year", "DimDate", "Year", "Năm", 24, 18, 150, 55, 1),
        slicer("slicer_month", "DimDate", "MonthLabel", "Kỳ báo cáo", 185, 18, 180, 55, 2),
        slicer("slicer_mien", "DimTerritory", "Mien", "Miền", 380, 18, 200, 55, 3),
        card("card_actual", "Doanh thu Sell In", "Doanh thu Sell In", 24, 100, 240, 115, 4),
        card("card_target_attainment", "Tỷ lệ đạt Target", "Đạt Target", 280, 100, 240, 115, 5),
        card("card_gap", "Khoảng cách Target", "Khoảng cách Target", 536, 100, 240, 115, 6),
        card("card_yoy", "Tăng trưởng YoY", "Tăng trưởng YoY", 792, 100, 240, 115, 7),
        card("card_forecast", "Run-rate dự báo tháng", "Run-rate dự báo", 1048, 100, 240, 115, 8),
        card("card_forecast_target", "Dự báo đạt Target", "Dự báo đạt Target", 1304, 100, 240, 115, 9),
        line_chart("trend_actual_target", "DimDate", "MonthLabel", ["Doanh thu Sell In", "Target", "Doanh thu LY"], "Xu hướng doanh thu vs kế hoạch (VND)", 24, 245, 760, 300, 10),
        bar_chart("region_attainment", "DimTerritory", "Vung", ["Doanh thu Sell In", "Target"], "Thực tế và Target theo vùng (VND)", 808, 245, 736, 300, 11, horizontal=True),
        table_visual("at_risk_regions", [("DimTerritory", "Mien", "Miền"), ("DimTerritory", "Vung", "Vùng")], ["Doanh thu Sell In", "Target", "Tỷ lệ đạt Target", "Khoảng cách Target", "Tăng trưởng YoY"], "Bảng điều hành vùng cần ưu tiên", 24, 570, 1520, 285, 12),
    ])
    page_specs["KeHoach_DuBao"] = ("Kế hoạch & dự báo", [
        slicer("slicer_plan_year", "DimDate", "Year", "Năm", 24, 18, 150, 55, 1),
        slicer("slicer_plan_mien", "DimTerritory", "Mien", "Miền", 185, 18, 200, 55, 2),
        card("plan_target", "Target", "Target kỳ hiện tại", 24, 100, 260, 115, 3),
        card("plan_actual", "Doanh thu Sell In", "Actual kỳ hiện tại", 300, 100, 260, 115, 4),
        card("plan_gap", "Còn thiếu để đạt Target", "Còn thiếu", 576, 100, 260, 115, 5),
        card("plan_need_day", "Cần doanh thu mỗi ngày", "Cần mỗi ngày còn lại", 852, 100, 260, 115, 6),
        card("plan_3m", "Target 3 tháng tới", "Target 3 tháng tới", 1128, 100, 260, 115, 7),
        line_chart("plan_trend", "DimDate", "MonthLabel", ["Doanh thu Sell In", "Target", "Run-rate dự báo tháng"], "Actual – Target – Run-rate theo tháng", 24, 245, 900, 320, 8),
        table_visual("plan_monthly", [("DimDate", "MonthLabel", "Tháng"), ("DimDate", "Year", "Năm")], ["Doanh thu Sell In", "Target", "Khoảng cách Target", "Tỷ lệ đạt Target", "Run-rate dự báo tháng"], "Bảng kế hoạch và dự báo", 948, 245, 596, 320, 9),
        table_visual("plan_actions", [("DimTerritory", "Mien", "Miền"), ("DimTerritory", "Vung", "Vùng")], ["Còn thiếu để đạt Target", "Cần doanh thu mỗi ngày", "Dự báo đạt Target"], "Ưu tiên hành động", 24, 590, 1520, 265, 10),
    ])
    page_specs["Vung_Mien"] = ("Vùng & miền", [
        slicer("slicer_area_year", "DimDate", "Year", "Năm", 24, 18, 150, 55, 1),
        slicer("slicer_area_period", "DimDate", "MonthLabel", "Kỳ", 185, 18, 180, 55, 2),
        bar_chart("area_actual_target", "DimTerritory", "Mien", ["Doanh thu Sell In", "Target"], "Actual và Target theo miền (VND)", 24, 100, 750, 320, 3),
        bar_chart("area_yoy", "DimTerritory", "Mien", ["Tăng trưởng YoY", "Tăng trưởng MoM"], "Động lực tăng trưởng theo miền", 800, 100, 744, 320, 4),
        table_visual("area_detail", [("DimTerritory", "Mien", "Miền"), ("DimTerritory", "Vung", "Vùng")], ["Doanh thu Sell In", "Target", "Tỷ lệ đạt Target", "Đạt Target Vikoda", "Doanh thu KDT", "Tăng trưởng YoY"], "Chi tiết quản trị theo vùng", 24, 455, 1520, 395, 5),
    ])
    page_specs["KhachHang_SanPham"] = ("Khách hàng & sản phẩm", [
        slicer("slicer_detail_year", "DimDate", "Year", "Năm", 24, 18, 150, 55, 1),
        slicer("slicer_detail_mien", "DimTerritory", "Mien", "Miền", 185, 18, 200, 55, 2),
        card("detail_customers", "Số khách hàng", "Số khách hàng", 24, 100, 250, 115, 3),
        card("detail_vikoda_mix", "Tỷ trọng Vikoda", "Tỷ trọng Vikoda", 290, 100, 250, 115, 4),
        card("detail_returns", "Tỷ lệ trả hàng", "Tỷ lệ trả hàng", 556, 100, 250, 115, 5),
        card("detail_kdt", "Doanh thu KDT", "Doanh thu KDT", 822, 100, 250, 115, 6),
        bar_chart("top_customers", "DimCustomer", "CustomerName", ["Doanh thu Sell In"], "Top khách hàng theo doanh thu", 24, 245, 760, 300, 7, horizontal=True),
        bar_chart("top_products", "DimProduct", "ProductName", ["Doanh thu Sell In"], "Top sản phẩm theo doanh thu", 808, 245, 736, 300, 8, horizontal=True),
        table_visual("customer_product_detail", [("DimCustomer", "CustomerName", "Khách hàng"), ("DimProduct", "ProductName", "Sản phẩm"), ("DimTerritory", "Vung", "Vùng")], ["Doanh thu Sell In", "Sản lượng", "Tăng trưởng YoY", "Tỷ trọng Vikoda"], "Drill-down khách hàng → sản phẩm", 24, 570, 1520, 280, 9),
    ])

    page_names = list(page_specs.keys())
    for page_name, (display_name, visuals) in page_specs.items():
        page_folder = report_definition_dir / "pages" / page_name
        page_folder.mkdir(parents=True, exist_ok=True)
        write_json(page_folder / "page.json", {
            "$schema": f"{REPORT_SCHEMA}/page/2.0.0/schema.json",
            "name": page_name,
            "displayName": display_name,
            "displayOption": "FitToPage",
            "width": 1600,
            "height": 900,
            "objects": {"background": [{"properties": {"color": solid_color("#F3F6FA"), "transparency": literal("0D")}}]},
        })
        visuals_dir = page_folder / "visuals"
        for visual in visuals:
            visual_dir = visuals_dir / visual["name"]
            write_json(visual_dir / "visual.json", visual)
    write_json(report_definition_dir / "pages" / "pages.json", {
        "$schema": f"{REPORT_SCHEMA}/pagesMetadata/1.0.0/schema.json",
        "pageOrder": page_names,
        "activePageName": page_names[0],
    })
    return page_names


EXECUTIVE_COLORS = {
    "navy": "#102A43",
    "navy_2": "#163A5F",
    "blue": "#2563EB",
    "cyan": "#0891B2",
    "teal": "#0F766E",
    "green": "#16A34A",
    "amber": "#D97706",
    "red": "#DC2626",
    "purple": "#7C3AED",
    "ink": "#172033",
    "muted": "#64748B",
    "line": "#DCE3ED",
    "canvas": "#F3F6FA",
    "surface": "#FFFFFF",
    "white": "#FFFFFF",
}


MATRIX_LABELS = {
    "Doanh thu Sell In": "Actual",
    "Target": "Target",
    "Tỷ lệ đạt Target": "% đạt",
    "Khoảng cách Target": "Gap",
    "Còn thiếu để đạt Target": "Còn thiếu",
    "Cần doanh thu mỗi ngày": "Cần/ngày",
    "Run-rate dự báo tháng": "Dự báo",
    "Dự báo EOM Mùa vụ": "Dự báo mùa vụ",
    "Dự báo đạt Target": "Dự báo %",
    "Nhịp độ bán hàng": "Pacing %",
    "% Thời gian tháng đã qua": "Thời gian %",
    "Doanh thu ngày bình quân MTD": "DT/ngày MTD",
    "Hệ số gia tốc cần đạt": "Gia tốc req",
    "Doanh thu Vikoda": "Vikoda",
    "Target Vikoda": "Target VKD",
    "Đạt Target Vikoda": "% đạt VKD",
    "Doanh thu KDT": "KDT",
    "Doanh thu LY": "Cùng kỳ LY",
    "Tăng trưởng YoY": "YoY %",
    "Tăng trưởng MoM": "MoM %",
    "Sản lượng": "Sản lượng",
    "Số khách hàng": "Khách hàng",
    "Tỷ lệ trả hàng": "% trả hàng",
    "Tỷ trọng Vikoda": "Tỷ trọng VKD",
    "Sản lượng quy đổi KTB": "SL quy đổi K/T/B",
    "SL Két (K)": "SL Két (K)",
    "SL Thùng (T)": "SL Thùng (T)",
    "SL Bình (B)": "SL Bình (B)",
    "Tỷ trọng Két %": "Két %",
    "Tỷ trọng Thùng %": "Thùng %",
    "Tỷ trọng Bình %": "Bình %",
    "SL Két (K) LY": "Két LY",
    "SL Thùng (T) LY": "Thùng LY",
    "SL Bình (B) LY": "Bình LY",
    "Tăng trưởng Két (K)": "Két YoY %",
    "Tăng trưởng Thùng (T)": "Thùng YoY %",
    "Tăng trưởng Bình (B)": "Bình YoY %",
    "SL chưa quy đổi": "SL chưa đổi",
    "Doanh thu Vikoda LY": "Vikoda LY",
    "Tăng trưởng Vikoda": "Vikoda YoY %",
    "Doanh thu KDT LY": "KDT LY",
    "Tăng trưởng KDT": "KDT YoY %",
    "Tỷ trọng KDT": "Tỷ trọng KDT",
    "Số SKU": "Số SKU",
    "Số khách hàng LY": "KH cùng kỳ LY",
    "Tăng trưởng số khách hàng": "KH YoY %",
    "Doanh thu bình quân KH": "DT/KH",
    "Doanh thu / Điểm bán Active": "Drop Size",
    "Khách hàng mới": "KH mới",
    "Khách hàng ngừng mua": "KH ngừng mua",
    "Doanh thu hệ thống MT": "DT hệ thống MT",
    "Xếp hạng SKU": "Hạng",
    "Đóng góp doanh thu": "% đóng góp",
}

MONEY_MEASURES = {
    "Doanh thu Sell In", "Doanh thu bán", "Doanh thu trả hàng", "Target", "Target Vikoda",
    "Khoảng cách Target", "Doanh thu LY", "Doanh thu tháng trước", "Doanh thu Vikoda",
    "Khoảng cách Vikoda", "Doanh thu KDT", "Run-rate dự báo tháng", "Dự báo EOM Mùa vụ",
    "Còn thiếu để đạt Target", "Cần doanh thu mỗi ngày", "Doanh thu ngày bình quân MTD",
    "Target 3 tháng tới", "Doanh thu Vikoda LY", "Doanh thu KDT LY", "Doanh thu bình quân KH",
    "Doanh thu / Điểm bán Active", "Doanh thu hệ thống MT",
}
PERCENT_MEASURES = {
    "Tỷ lệ đạt Target", "Đạt Target Vikoda", "Tăng trưởng YoY", "Tăng trưởng MoM",
    "Dự báo đạt Target", "Nhịp độ bán hàng", "% Thời gian tháng đã qua", "Tỷ lệ trả hàng",
    "Tỷ trọng Vikoda", "Tăng trưởng Két (K)", "Tăng trưởng Thùng (T)", "Tăng trưởng Bình (B)",
    "Tỷ trọng Két %", "Tỷ trọng Thùng %", "Tỷ trọng Bình %", "Tăng trưởng Vikoda",
    "Tăng trưởng KDT", "Tỷ trọng KDT", "Tăng trưởng số khách hàng", "Đóng góp doanh thu",
}
INTEGER_MEASURES = {
    "Số khách hàng", "Số SKU", "Số khách hàng LY", "Khách hàng mới",
    "Khách hàng ngừng mua", "Xếp hạng SKU",
}
QUANTITY_MEASURES = {
    "Sản lượng", "Sản lượng quy đổi KTB", "SL Két (K)", "SL Thùng (T)", "SL Bình (B)",
    "SL Két (K) LY", "SL Thùng (T) LY", "SL Bình (B) LY", "SL chưa quy đổi",
}


def _axis_title(entity: str, column: str) -> str:
    titles = {
        "MonthLabel": "Kỳ báo cáo", "MonthAxis": "Kỳ YY/MM", "MonthStart": "Kỳ báo cáo", "Year": "Năm", "Mien": "Miền", "Vung": "Vùng",
        "CustomerName": "Khách hàng", "ProductShortName": "Sản phẩm", "ProductAxisLabel": "Sản phẩm", "PackUnit": "ĐVT",
        "Channel": "Kênh", "CustomerType": "Loại khách hàng", "SystemMT": "Hệ thống MT",
        "Province": "Tỉnh/Thành", "ProductGroup": "Nhóm sản phẩm", "CoBrand": "Thương hiệu",
        "PackagingType": "Kiểu bao bì", "ProductType": "Loại hàng",
    }
    return titles.get(column, column)


def _value_axis_title(measures: list[str]) -> str:
    if measures and all(item in PERCENT_MEASURES for item in measures):
        return "%"
    if any(item in MONEY_MEASURES for item in measures):
        return "Triệu đồng"
    if measures and all(item in INTEGER_MEASURES for item in measures):
        return "Số lượng"
    return "Sản lượng quy đổi"


def executive_nav_panel(name: str, active: str, tab_order: int = 1) -> dict[str, Any]:
    """ERP-style rail background; interactive page buttons are layered above it."""
    paragraphs = [{"textRuns": [{"value": "MENU ĐIỀU HÀNH", "textStyle": {
        "fontFamily": "Segoe UI Semibold", "fontSize": "11pt", "fontWeight": "600", "color": EXECUTIVE_COLORS["white"],
    }}]}]
    paragraphs.append({"textRuns": [{"value": "CHUYỂN TRANG  ·  BỘ LỌC", "textStyle": {
        "fontFamily": "Segoe UI", "fontSize": "8pt", "fontWeight": "400", "color": "#9FB6CF",
    }}]})
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": 0, "y": 72, "z": tab_order, "width": 220, "height": 648, "tabOrder": tab_order},
        "visual": {
            "visualType": "textbox",
            "objects": {"general": [{"properties": {"paragraphs": paragraphs}}]},
            "visualContainerObjects": {
                "title": [{"properties": {"show": literal("false")}}],
                "background": [{"properties": {"show": literal("true"), "color": solid_color(EXECUTIVE_COLORS["navy_2"]), "transparency": literal("0D")}}],
                "border": [{"properties": {"show": literal("false")}}],
                "dropShadow": [{"properties": {"show": literal("false")}}],
                "padding": [{"properties": {"top": literal("18D"), "bottom": literal("10D"), "left": literal("16D"), "right": literal("12D")}}],
                "visualHeader": [{"properties": {"show": literal("false")}}],
            },
            "drillFilterOtherVisuals": False,
        },
    }


def executive_nav_button(
    name: str,
    label: str,
    target_page: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    active: bool,
) -> dict[str, Any]:
    """Clickable Power BI page-navigation button with an explicit active state."""
    fill = "#0EA5E9" if active else "#1B456A"
    font = EXECUTIVE_COLORS["white"] if active else "#DCEAF7"
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "actionButton",
            "objects": {
                "icon": [
                    {"properties": {"shapeType": literal("'blank'")}, "selector": {"id": "default"}},
                    {"properties": {"show": literal("false")}},
                ],
                "text": [
                    {"properties": {"show": literal("true")}},
                    {"properties": {
                        "text": literal(quote_dax_literal(label)),
                        "fontColor": solid_color(font),
                        "fontSize": literal("9D"),
                        "bold": literal("true" if active else "false"),
                        "fontFamily": literal("'Segoe UI Semibold'"),
                        "horizontalAlignment": literal("'left'"),
                        "verticalAlignment": literal("'middle'"),
                        "leftMargin": literal("12L"),
                        "rightMargin": literal("8L"),
                    }, "selector": {"id": "default"}},
                ],
                "outline": [{"properties": {"show": literal("false")}}],
                "fill": [
                    {"properties": {"show": literal("true")}},
                    {"properties": {"fillColor": solid_color(fill)}, "selector": {"id": "default"}},
                    {"properties": {"fillColor": solid_color("#2A5A7F")}, "selector": {"id": "hover"}},
                ],
                "shape": [{"properties": {"roundEdge": literal("6L")}}],
            },
            "visualContainerObjects": {
                "title": [{"properties": {"show": literal("false")}}],
                "visualLink": [{"properties": {
                    "show": literal("true"),
                    "type": literal("'PageNavigation'"),
                    "navigationSection": literal(quote_dax_literal(target_page)),
                    "showDefaultTooltip": literal("false"),
                }}],
                "background": [{"properties": {"show": literal("false")}}],
                "border": [{"properties": {"show": literal("false"), "radius": literal("6D")}}],
                "dropShadow": [{"properties": {"show": literal("false")}}],
                "padding": [{"properties": {
                    "top": literal("0D"), "bottom": literal("0D"), "left": literal("0D"), "right": literal("0D"),
                }}],
                "visualHeader": [{"properties": {"show": literal("false")}}],
            },
            "drillFilterOtherVisuals": False,
        },
    }


NAV_PAGES: tuple[tuple[str, str], ...] = (
    ("CEO_TongQuan", "01  TỔNG QUAN"),
    ("Kenh_KhachHang", "02  KÊNH & KH"),
    ("SanPham", "03  SẢN PHẨM"),
    ("Vung_Mien", "04  VÙNG MIỀN"),
    ("KhachHang_SanPham", "05  CHI TIẾT"),
    ("KeHoach_KhuyenNghi", "06  KHUYẾN NGHỊ"),
)


def executive_nav_buttons(active_page: str, start_tab_order: int = 2) -> list[dict[str, Any]]:
    return [
        executive_nav_button(
            f"nav_to_{target}", label, target, 14, NAV_BUTTON_Y + index * NAV_BUTTON_STEP, 192, NAV_BUTTON_HEIGHT,
            start_tab_order + index, active=(target == active_page),
        )
        for index, (target, label) in enumerate(NAV_PAGES)
    ]


def executive_container(
    title: str | None,
    *,
    subtitle: str | None = None,
    background: str = EXECUTIVE_COLORS["surface"],
    title_color: str = EXECUTIVE_COLORS["ink"],
    border_color: str = EXECUTIVE_COLORS["line"],
    shadow: bool = True,
    radius: int = 10,
    padding: int = 8,
    title_size: int = 11,
) -> dict[str, Any]:
    return {
        "title": [{"properties": {
            "show": literal("true" if title else "false"),
            "text": literal(quote_dax_literal(title or "")),
            "fontSize": literal(f"{title_size}D"),
            "bold": literal("true"),
            "fontColor": solid_color(title_color),
            "fontFamily": literal("'Segoe UI Semibold'"),
            "titleWrap": literal("true"),
            "alignment": literal("'left'"),
        }}],
        "subTitle": [{"properties": {
            "show": literal("true" if subtitle else "false"),
            "text": literal(quote_dax_literal(subtitle or "")),
            "fontSize": literal("9D"),
            "fontColor": solid_color(EXECUTIVE_COLORS["muted"]),
            "fontFamily": literal("'Segoe UI'"),
            "alignment": literal("'left'"),
        }}],
        "background": [{"properties": {
            "show": literal("true"),
            "color": solid_color(background),
            "transparency": literal("0D"),
        }}],
        "border": [{"properties": {
            "show": literal("true"),
            "color": solid_color(border_color),
            "radius": literal(f"{radius}D"),
            "width": literal("1D"),
        }}],
        "padding": [{"properties": {
            "top": literal(f"{padding}D"),
            "bottom": literal(f"{padding}D"),
            "left": literal(f"{padding}D"),
            "right": literal(f"{padding}D"),
        }}],
        "dropShadow": [{"properties": {
            "show": literal("true" if shadow else "false"),
            "angle": literal("90L"),
            "shadowDistance": literal("2L"),
            "shadowBlur": literal("8L"),
            "shadowSpread": literal("0L"),
            "transparency": literal("86L"),
            "color": solid_color("#39506A"),
        }}],
        "visualHeader": [{"properties": {"show": literal("false")}}],
    }


def executive_header(name: str, title: str, subtitle: str, tab_order: int = 0) -> dict[str, Any]:
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": 0, "y": 0, "z": tab_order, "width": 1280, "height": 72, "tabOrder": tab_order},
        "visual": {
            "visualType": "textbox",
            "objects": {"general": [{"properties": {"paragraphs": [
                {"textRuns": [{"value": title, "textStyle": {
                    "fontFamily": "Segoe UI Semibold", "fontSize": "21pt", "fontWeight": "600", "color": EXECUTIVE_COLORS["white"],
                }}]},
                {"textRuns": [{"value": subtitle, "textStyle": {
                    "fontFamily": "Segoe UI", "fontSize": "9pt", "fontWeight": "400", "color": "#C8D7E8",
                }}]},
            ]}}]},
            "visualContainerObjects": {
                "title": [{"properties": {"show": literal("false")}}],
                "background": [{"properties": {"show": literal("true"), "color": solid_color(EXECUTIVE_COLORS["navy"]), "transparency": literal("0D")}}],
                "border": [{"properties": {"show": literal("false")}}],
                "dropShadow": [{"properties": {"show": literal("false")}}],
                "padding": [{"properties": {"top": literal("9D"), "bottom": literal("6D"), "left": literal("24D"), "right": literal("24D")}}],
                "visualHeader": [{"properties": {"show": literal("false")}}],
            },
            "drillFilterOtherVisuals": False,
        },
    }


def executive_card(
    name: str,
    measure_name: str,
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    accent: str,
    tint: str,
) -> dict[str, Any]:
    precision = 0 if measure_name in MONEY_MEASURES or measure_name in INTEGER_MEASURES or measure_name in QUANTITY_MEASURES else 1
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "card",
            "query": {"queryState": {"Values": {"projections": [projection(
                field_measure("FactSellIn", measure_name), f"FactSellIn.{measure_name}", title
            )]}}},
            "objects": {
                # Chiều cao khả dụng của thẻ chỉ còn ~42 px sau tiêu đề và
                # padding. Chữ 22pt cao hơn khoảng đó nên phần chân số bị cắt;
                # 17pt vừa khít mà vẫn đọc rõ từ xa.
                "labels": [{"properties": {
                    "color": solid_color(accent), "fontSize": literal("17D"), "bold": literal("true"),
                    "fontFamily": literal("'Segoe UI Semibold'"),
                    "labelDisplayUnits": literal("1D"), "labelPrecision": literal(f"{precision}L"),
                    "horizontalAlignment": literal("'left'"),
                }}],
                "categoryLabels": [{"properties": {"show": literal("false")}}],
                "wordWrap": [{"properties": {"show": literal("false")}}],
            },
            "visualContainerObjects": executive_container(
                title, background=tint, title_color=EXECUTIVE_COLORS["muted"], border_color=accent,
                padding=6, title_size=8, radius=8, shadow=False,
            ),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_date_range_slicer(
    name: str,
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    entity: str = "DimDate",
    column: str = "Date",
) -> dict[str, Any]:
    card_bg = "#FFFFFF"
    card_border = "#93C5FD"
    title_color = "#0369A1"
    result = {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "slicer",
            "query": {"queryState": {"Values": {"projections": [projection(
                field_column(entity, column), f"{entity}.{column}", title
            )]}}},
            "objects": {
                "data": [{"properties": {"mode": literal("'Between'")}}],
                "header": [{"properties": {"show": literal("false")}}],
                "slider": [{"properties": {
                    "show": literal("true"),
                    "color": solid_color("#0284C7"),
                }}],
                "input": [{"properties": {
                    "fontColor": solid_color("#0F172A"),
                    "backColor": solid_color("#FFFFFF"),
                    "fontSize": literal("9D"),
                    "fontFamily": literal("'Segoe UI Semibold'"),
                    "outlineColor": solid_color("#CBD5E1"),
                    "outlineWeight": literal("1D"),
                }}],
            },
            "visualContainerObjects": executive_container(
                title,
                background=card_bg,
                title_color=title_color,
                border_color=card_border,
                shadow=False,
                radius=8,
                padding=6,
                title_size=8,
            ),
            "syncGroup": {"groupName": f"sync_{entity}_{column}", "fieldChanges": True, "filterChanges": True},
            "drillFilterOtherVisuals": True,
        },
    }
    return result


def executive_slicer(
    name: str,
    entity: str,
    column: str,
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    single_select: bool = False,
) -> dict[str, Any]:
    rail_surface = "#123A5D"
    rail_border = "#557A98"
    result = {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "slicer",
            "query": {"queryState": {"Values": {"projections": [projection(
                field_column(entity, column), f"{entity}.{column}", title
            )]}}},
            "objects": {
                "data": [{"properties": {"mode": literal("'Dropdown'")}}],
                "selection": [{"properties": {"singleSelect": literal("true" if single_select else "false")}}],
                "header": [{"properties": {"show": literal("false")}}],
                "items": [{"properties": {
                    "fontColor": solid_color("#7DD3FC"), "backColor": solid_color(rail_surface),
                    "fontSize": literal("10D"), "fontFamily": literal("'Segoe UI Semibold'"),
                    "outlineColor": solid_color(rail_border), "outlineWeight": literal("1D"),
                }}],
            },
            "visualContainerObjects": executive_container(
                title,
                background=rail_surface,
                title_color="#BFD2E5",
                border_color=rail_border,
                shadow=False,
                radius=6,
                padding=5,
                title_size=8,
            ),
            "syncGroup": {"groupName": f"sync_{entity}_{column}", "fieldChanges": True, "filterChanges": True},
            "drillFilterOtherVisuals": True,
        },
    }
    return result


def executive_bar(
    name: str,
    category_entity: str,
    category: str,
    measures: list[str],
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    palette: list[str],
    horizontal: bool = True,
    show_labels: bool = True,
    sort_ascending: bool = False,
    subtitle: str | None = None,
    long_labels: bool = False,
) -> dict[str, Any]:
    measure_projections = [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}", MATRIX_LABELS.get(item, item)) for item in measures]
    percent_only = bool(measures) and all(item in PERCENT_MEASURES for item in measures)
    precision = 1 if percent_only else 0
    label_properties: dict[str, Any] = {
        "show": literal("true" if show_labels else "false"),
        "fontSize": literal("9D"),
        "color": solid_color(EXECUTIVE_COLORS["muted"]),
        "labelPrecision": literal(f"{precision}L"),
    }
    if not percent_only:
        label_properties["labelDisplayUnits"] = literal("1D")
    data_points = []
    for index, item in enumerate(measure_projections):
        point: dict[str, Any] = {"properties": {"fill": solid_color(palette[index % len(palette)])}}
        if len(measure_projections) > 1:
            point["selector"] = {"metadata": item["queryRef"]}
        data_points.append(point)
    axis_note = f"X · {_axis_title(category_entity, category)}   |   Y · {_value_axis_title(measures)}"
    # Power BI chỉ dành 25% chiều rộng visual cho nhãn trục phân loại. Tên SKU
    # và tên khách hàng dài 40-50 ký tự nên bị cắt thành dấu ba chấm. Nới hạn
    # mức này và giảm cỡ chữ để nhãn hiện trọn.
    category_font = 9 if long_labels else 10
    category_axis_properties: dict[str, Any] = {
        "show": literal("true"), "fontSize": literal(f"{category_font}D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
        "titleShow": literal("true"), "titleText": literal(quote_dax_literal(_axis_title(category_entity, category))),
        "showAxisTitle": literal("true"), "labelColor": solid_color(EXECUTIVE_COLORS["ink"]),
        "titleFontSize": literal("10D"), "titleFontColor": solid_color(EXECUTIVE_COLORS["ink"]),
    }
    if long_labels:
        category_axis_properties["maxMarginFactor"] = literal("48D")
        category_axis_properties["concatenateLabels"] = literal("false")
        category_axis_properties["innerPadding"] = literal("18D")
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "barChart" if horizontal else "clusteredColumnChart",
            "query": {
                "queryState": {
                    "Category": {"projections": [projection(field_column(category_entity, category), f"{category_entity}.{category}", category)]},
                    "Y": {"projections": measure_projections},
                },
                "sortDefinition": {"sort": [{"field": measure_projections[0]["field"], "direction": "Ascending" if sort_ascending else "Descending"}], "isDefaultSort": True},
            },
            "objects": {
                "legend": [{"properties": {"show": literal("true" if len(measures) > 1 else "false"), "position": literal("'Top'")}}],
                "dataPoint": data_points,
                "labels": [{"properties": label_properties}],
                "categoryAxis": [{"properties": category_axis_properties}],
                "valueAxis": [{"properties": {
                    "show": literal("true"), "labelDisplayUnits": literal("1D"), "labelPrecision": literal(f"{precision}L"),
                    "fontSize": literal("10D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "titleShow": literal("true"), "titleText": literal(quote_dax_literal(_value_axis_title(measures))),
                    "showAxisTitle": literal("true"), "labelColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "titleFontSize": literal("10D"), "titleFontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "gridlineShow": literal("true"), "gridlineColor": solid_color("#D4DCE6"),
                }}],
            },
            "visualContainerObjects": executive_container(title, subtitle=subtitle or axis_note, shadow=False),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_combo(
    name: str,
    category_entity: str,
    category: str,
    column_measures: list[str],
    line_measures: list[str],
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    column_palette: list[str],
    line_palette: list[str],
    show_labels: bool = True,
    scalar_axis: bool = False,
    sort_by_category: bool = False,
) -> dict[str, Any]:
    """DaTaxan-inspired vertical columns plus a highlighted trend line."""
    columns = [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}", MATRIX_LABELS.get(item, item)) for item in column_measures]
    lines = [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}", MATRIX_LABELS.get(item, item)) for item in line_measures]
    data_points: list[dict[str, Any]] = []
    for index, item in enumerate(columns):
        data_points.append({
            "properties": {"fill": solid_color(column_palette[index % len(column_palette)])},
            "selector": {"metadata": item["queryRef"]},
        })
    for index, item in enumerate(lines):
        data_points.append({
            "properties": {"fill": solid_color(line_palette[index % len(line_palette)])},
            "selector": {"metadata": item["queryRef"]},
        })
    line_axis_title = " / ".join(MATRIX_LABELS.get(item, item) for item in line_measures)
    axis_note = f"X · {_axis_title(category_entity, category)}   |   CỘT · {_value_axis_title(column_measures)}   |   ĐƯỜNG · {line_axis_title}"
    category_projection = projection(
        field_column(category_entity, category),
        f"{category_entity}.{category}",
        _axis_title(category_entity, category),
    )
    sort_projection = category_projection if sort_by_category else columns[0]
    sort_direction = "Ascending" if sort_by_category else "Descending"
    category_axis_properties: dict[str, Any] = {
        "show": literal("true"), "showAxisTitle": literal("true"), "titleShow": literal("true"),
        "titleText": literal(quote_dax_literal(_axis_title(category_entity, category))),
        "fontSize": literal("11D"), "textSize": literal("11D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
        "labelColor": solid_color(EXECUTIVE_COLORS["ink"]), "titleFontSize": literal("11D"),
        "titleFontColor": solid_color(EXECUTIVE_COLORS["ink"]), "gridlineShow": literal("false"),
        "labelDensity": literal("50D" if scalar_axis else "100D"), "concatenateLabels": literal("false"),
        "wordWrap": literal("false"),
    }
    category_axis_properties["axisType"] = literal("'Scalar'" if scalar_axis else "'Categorical'")
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "lineClusteredColumnComboChart",
            "query": {
                "queryState": {
                    "Category": {"projections": [category_projection]},
                    # Native combo charts use Y for columns and Y2 for the
                    # secondary line. ColumnY/LineY renders a blank plot in
                    # current Power BI Desktop builds.
                    "Y": {"projections": columns},
                    "Y2": {"projections": lines},
                },
                "sortDefinition": {"sort": [{"field": sort_projection["field"], "direction": sort_direction}], "isDefaultSort": True},
            },
            "objects": {
                "legend": [{"properties": {
                    "show": literal("true"), "position": literal("'Top'"), "fontSize": literal("9D"),
                    "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                }}],
                "dataPoint": data_points,
                "labels": [{"properties": {
                    "show": literal("true" if show_labels else "false"), "fontSize": literal("9D"), "color": solid_color(EXECUTIVE_COLORS["ink"]),
                    "labelDisplayUnits": literal("1D"), "labelPrecision": literal("0L"),
                }}],
                "lineStyles": [{"properties": {
                    "strokeWidth": literal("3D"), "lineChartType": literal("'smooth'"), "showMarker": literal("true"),
                }}],
                "markers": [{"properties": {
                    "show": literal("true"), "markerSize": literal("7D"), "borderShow": literal("true"),
                    "borderColor": solid_color(EXECUTIVE_COLORS["white"]),
                }}],
                "categoryAxis": [{"properties": category_axis_properties}],
                "valueAxis": [{"properties": {
                    "show": literal("true"), "showAxisTitle": literal("true"), "titleShow": literal("true"),
                    "titleText": literal(quote_dax_literal(_value_axis_title(column_measures))),
                    "labelDisplayUnits": literal("1D"), "labelPrecision": literal("0L"),
                    "fontSize": literal("11D"), "textSize": literal("11D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "labelColor": solid_color(EXECUTIVE_COLORS["ink"]), "titleFontSize": literal("11D"),
                    "titleFontColor": solid_color(EXECUTIVE_COLORS["ink"]), "gridlineShow": literal("true"),
                    "gridlineColor": solid_color("#D4DCE6"), "secShow": literal("true"),
                    "secShowAxisTitle": literal("true"), "secTitleShow": literal("true"),
                    "secTitleText": literal(quote_dax_literal(line_axis_title)), "secLabelDisplayUnits": literal("1D"),
                    "secLabelPrecision": literal("1L"), "secFontSize": literal("11D"),
                    "secFontColor": solid_color(EXECUTIVE_COLORS["purple"]), "secGridlineShow": literal("false"),
                }}],
                "plotArea": [{"properties": {
                    "show": literal("true"), "backgroundColor": solid_color("#F8FBFF"), "transparency": literal("0D"),
                }}],
            },
            "visualContainerObjects": executive_container(title, subtitle=axis_note, background="#F8FBFF", border_color="#B8C7DB", shadow=False),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_line(
    name: str,
    category_entity: str,
    category: str,
    measures: list[str],
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    palette: list[str],
    show_labels: bool = True,
    scalar_axis: bool = False,
    sort_by_category: bool = True,
) -> dict[str, Any]:
    measure_projections = [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}", MATRIX_LABELS.get(item, item)) for item in measures]
    percent_only = bool(measures) and all(item in PERCENT_MEASURES for item in measures)
    precision = 1 if percent_only else 0
    axis_note = f"X · {_axis_title(category_entity, category)}   |   Y · {_value_axis_title(measures)}"
    line_label_properties: dict[str, Any] = {
        "show": literal("true" if show_labels else "false"), "fontSize": literal("9D"), "bold": literal("true"),
        "color": solid_color(EXECUTIVE_COLORS["ink"]), "labelPrecision": literal(f"{precision}L"),
        "background": solid_color("#FFFFFF"), "transparency": literal("18D"), "labelDensity": literal("35D"),
    }
    if not percent_only:
        line_label_properties["labelDisplayUnits"] = literal("1D")
    category_projection = projection(
        field_column(category_entity, category),
        f"{category_entity}.{category}",
        _axis_title(category_entity, category),
    )
    query: dict[str, Any] = {"queryState": {
        "Category": {"projections": [category_projection]},
        "Y": {"projections": measure_projections},
    }}
    if sort_by_category:
        query["sortDefinition"] = {
            "sort": [{"field": category_projection["field"], "direction": "Ascending"}],
            "isDefaultSort": True,
        }
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "lineChart",
            "query": query,
            "objects": {
                "legend": [{"properties": {"show": literal("true"), "position": literal("'Top'")}}],
                "dataPoint": [{"properties": {"fill": solid_color(palette[index % len(palette)])}, "selector": {"metadata": item["queryRef"]}} for index, item in enumerate(measure_projections)],
                "labels": [{"properties": line_label_properties}],
                "markers": [{"properties": {"show": literal("true"), "markerSize": literal("6D"), "borderShow": literal("false")}}],
                "lineStyles": [{"properties": {"strokeWidth": literal("3D"), "lineChartType": literal("'smooth'"), "showMarker": literal("true")}}],
                "categoryAxis": [{"properties": {
                    "show": literal("true"), "fontSize": literal("11D"), "textSize": literal("11D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "axisType": literal("'Scalar'" if scalar_axis else "'Categorical'"),
                    "titleShow": literal("true"), "titleText": literal(quote_dax_literal(_axis_title(category_entity, category))),
                    "showAxisTitle": literal("true"), "labelColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "titleFontSize": literal("11D"), "titleFontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "labelDensity": literal("50D" if scalar_axis else "100D"), "concatenateLabels": literal("false"), "wordWrap": literal("false"),
                    "gridlineShow": literal("false"),
                }}],
                "valueAxis": [{"properties": {
                    "show": literal("true"), "labelDisplayUnits": literal("1D"), "labelPrecision": literal(f"{precision}L"),
                    "fontSize": literal("11D"), "textSize": literal("11D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "titleShow": literal("true"), "titleText": literal(quote_dax_literal(_value_axis_title(measures))),
                    "showAxisTitle": literal("true"), "labelColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "titleFontSize": literal("11D"), "titleFontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "gridlineShow": literal("true"), "gridlineColor": solid_color("#CBD5E1"),
                }}],
                "plotArea": [{"properties": {
                    "show": literal("true"), "backgroundColor": solid_color("#F8FBFF"), "transparency": literal("0D"),
                }}],
            },
            "visualContainerObjects": executive_container(title, subtitle=axis_note, background="#F8FBFF", border_color="#B8C7DB", shadow=False),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_donut(
    name: str,
    category_entity: str,
    category: str,
    measure_name: str,
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
) -> dict[str, Any]:
    """Compact composition visual used only where part-to-whole adds meaning."""
    category_projection = projection(
        field_column(category_entity, category),
        f"{category_entity}.{category}",
        _axis_title(category_entity, category),
    )
    value_projection = projection(
        field_measure("FactSellIn", measure_name),
        f"FactSellIn.{measure_name}",
        MATRIX_LABELS.get(measure_name, measure_name),
    )
    precision = 1 if measure_name in PERCENT_MEASURES else 0
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "donutChart",
            "query": {
                "queryState": {
                    "Category": {"projections": [category_projection]},
                    "Y": {"projections": [value_projection]},
                },
                "sortDefinition": {"sort": [{"field": value_projection["field"], "direction": "Descending"}], "isDefaultSort": True},
            },
            "objects": {
                "legend": [{"properties": {
                    "show": literal("true"), "position": literal("'Right'"), "fontSize": literal("10D"),
                    "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                }}],
                "labels": [{"properties": {
                    "show": literal("true"), "fontSize": literal("10D"), "bold": literal("true"),
                    "color": solid_color(EXECUTIVE_COLORS["ink"]), "labelDisplayUnits": literal("1D"),
                    "labelPrecision": literal(f"{precision}L"), "enableBackground": literal("true"),
                    "backgroundColor": solid_color(EXECUTIVE_COLORS["white"]), "backgroundTransparency": literal("18D"),
                }}],
                "categoryLabels": [{"properties": {
                    "show": literal("true"), "fontSize": literal("10D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                }}],
                "dataPoint": [{"properties": {
                    "borderShow": literal("true"), "borderColor": solid_color(EXECUTIVE_COLORS["white"]),
                    "borderSize": literal("2D"),
                }}],
            },
            "visualContainerObjects": executive_container(
                title,
                subtitle=f"Cơ cấu theo {_axis_title(category_entity, category).lower()}",
                background="#F8FBFF",
                border_color="#B8C7DB",
                shadow=False,
            ),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_waterfall(
    name: str,
    category_entity: str,
    category: str,
    measure_name: str,
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
) -> dict[str, Any]:
    """Positive/negative contribution chart for management exceptions."""
    category_projection = projection(
        field_column(category_entity, category),
        f"{category_entity}.{category}",
        _axis_title(category_entity, category),
    )
    value_projection = projection(
        field_measure("FactSellIn", measure_name),
        f"FactSellIn.{measure_name}",
        MATRIX_LABELS.get(measure_name, measure_name),
    )
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "waterfallChart",
            "query": {
                "queryState": {
                    "Category": {"projections": [category_projection]},
                    "Y": {"projections": [value_projection]},
                },
                "sortDefinition": {"sort": [{"field": value_projection["field"], "direction": "Descending"}], "isDefaultSort": True},
            },
            "objects": {
                "legend": [{"properties": {"show": literal("false")}}],
                "labels": [{"properties": {
                    "show": literal("true"), "fontSize": literal("9D"), "bold": literal("true"),
                    "color": solid_color(EXECUTIVE_COLORS["ink"]), "labelDisplayUnits": literal("1D"),
                    "labelPrecision": literal("0L"), "enableBackground": literal("true"),
                    "backgroundColor": solid_color(EXECUTIVE_COLORS["white"]), "backgroundTransparency": literal("15D"),
                }}],
                "categoryAxis": [{"properties": {
                    "show": literal("true"), "showAxisTitle": literal("true"), "titleShow": literal("true"),
                    "titleText": literal(quote_dax_literal(_axis_title(category_entity, category))),
                    "fontSize": literal("9D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "labelColor": solid_color(EXECUTIVE_COLORS["ink"]), "titleFontSize": literal("10D"),
                    # Tên miền dài hơn 25% chiều rộng mặc định dành cho nhãn trục.
                    "maxMarginFactor": literal("48D"), "concatenateLabels": literal("false"),
                }}],
                "valueAxis": [{"properties": {
                    "show": literal("true"), "showAxisTitle": literal("true"), "titleShow": literal("true"),
                    "titleText": literal(quote_dax_literal(_value_axis_title([measure_name]))),
                    "labelDisplayUnits": literal("1D"), "labelPrecision": literal("0L"),
                    "fontSize": literal("10D"), "fontColor": solid_color(EXECUTIVE_COLORS["ink"]),
                    "gridlineShow": literal("true"), "gridlineColor": solid_color("#D4DCE6"),
                    "totalsEnabled": literal("true"),
                }}],
                "plotArea": [{"properties": {
                    "show": literal("true"), "backgroundColor": solid_color("#F8FBFF"), "transparency": literal("0D"),
                }}],
            },
            "visualContainerObjects": executive_container(
                title,
                subtitle="Đóng góp dương / âm so với Target",
                background="#F8FBFF",
                border_color="#B8C7DB",
                shadow=False,
            ),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_treemap(
    name: str,
    group_entity: str,
    group_column: str,
    detail_entity: str,
    detail_column: str,
    measure_name: str,
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    subtitle: str | None = None,
) -> dict[str, Any]:
    """Bản đồ tỷ trọng hai cấp — thay cho bản đồ địa lý khi cột Tỉnh/Thành còn trống."""
    group_projection = projection(
        field_column(group_entity, group_column), f"{group_entity}.{group_column}", _axis_title(group_entity, group_column)
    )
    detail_projection = projection(
        field_column(detail_entity, detail_column), f"{detail_entity}.{detail_column}", _axis_title(detail_entity, detail_column)
    )
    value_projection = projection(
        field_measure("FactSellIn", measure_name), f"FactSellIn.{measure_name}", MATRIX_LABELS.get(measure_name, measure_name)
    )
    precision = 1 if measure_name in PERCENT_MEASURES else 0
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "treemap",
            "query": {
                # Treemap dùng bộ vai riêng: Group / Details / Values. Đặt nhầm
                # thành Category / Y như biểu đồ cột sẽ ra một khung trắng.
                "queryState": {
                    "Group": {"projections": [group_projection]},
                    "Details": {"projections": [detail_projection]},
                    "Values": {"projections": [value_projection]},
                },
                "sortDefinition": {"sort": [{"field": value_projection["field"], "direction": "Descending"}], "isDefaultSort": True},
            },
            "objects": {
                "legend": [{"properties": {
                    "show": literal("true"), "position": literal("'Top'"), "fontSize": literal("9D"),
                    "fontColor": solid_color(EXECUTIVE_COLORS["muted"]),
                }}],
                "labels": [{"properties": {
                    "show": literal("true"), "fontSize": literal("9D"), "bold": literal("true"),
                    "color": solid_color(EXECUTIVE_COLORS["white"]),
                    "labelDisplayUnits": literal("1D"), "labelPrecision": literal(f"{precision}L"),
                }}],
                "categoryLabels": [{"properties": {
                    "show": literal("true"), "fontSize": literal("9D"), "color": solid_color(EXECUTIVE_COLORS["white"]),
                }}],
                "dataPoint": [{"properties": {
                    "borderShow": literal("true"), "borderColor": solid_color(EXECUTIVE_COLORS["white"]), "borderSize": literal("2D"),
                }}],
            },
            "visualContainerObjects": executive_container(
                title,
                subtitle=subtitle or f"Ô lớn = đóng góp lớn · {_axis_title(group_entity, group_column)} → {_axis_title(detail_entity, detail_column)}",
                background="#F8FBFF",
                border_color="#B8C7DB",
                shadow=False,
            ),
            "drillFilterOtherVisuals": True,
        },
    }


def executive_insight_panel(
    name: str,
    title: str,
    bullets: list[tuple[str, str]],
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
) -> dict[str, Any]:
    """Khối chữ tĩnh nêu cách đọc số và việc cần làm — chốt lại mạch kể chuyện."""
    paragraphs: list[dict[str, Any]] = []
    for heading, body in bullets:
        paragraphs.append({"textRuns": [{"value": heading, "textStyle": {
            "fontFamily": "Segoe UI Semibold", "fontSize": "10pt", "fontWeight": "600", "color": EXECUTIVE_COLORS["navy"],
        }}]})
        paragraphs.append({"textRuns": [{"value": body, "textStyle": {
            "fontFamily": "Segoe UI", "fontSize": "9pt", "fontWeight": "400", "color": "#41556B",
        }}]})
        paragraphs.append({"textRuns": [{"value": " ", "textStyle": {"fontFamily": "Segoe UI", "fontSize": "4pt", "color": "#FFFFFF"}}]})
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "textbox",
            "objects": {"general": [{"properties": {"paragraphs": paragraphs}}]},
            "visualContainerObjects": executive_container(
                title, subtitle="Đọc số theo thứ tự này rồi giao việc", background="#FFFBEB",
                border_color="#D97706", shadow=False, padding=12,
            ),
            "drillFilterOtherVisuals": False,
        },
    }


def executive_matrix(
    name: str,
    rows: list[tuple[str, str, str]],
    measures: list[str],
    title: str,
    x: int,
    y: int,
    width: int,
    height: int,
    tab_order: int,
    *,
    sort_measure: str | None = None,
    sort_ascending: bool = False,
    subtitle: str | None = None,
) -> dict[str, Any]:
    row_projections = [projection(field_column(entity, field), f"{entity}.{field}", display) for entity, field, display in rows]
    value_projections = [projection(field_measure("FactSellIn", item), f"FactSellIn.{item}", MATRIX_LABELS.get(item, item)) for item in measures]
    all_projections = row_projections + value_projections
    sort_index = measures.index(sort_measure) if sort_measure in measures else 0
    return {
        "$schema": f"{REPORT_SCHEMA}/visualContainer/2.8.0/schema.json",
        "name": name,
        "position": {"x": x, "y": y, "z": tab_order, "width": width, "height": height, "tabOrder": tab_order},
        "visual": {
            "visualType": "tableEx",
            "query": {
                "queryState": {"Values": {"projections": all_projections}},
                "sortDefinition": {"sort": [{"field": value_projections[sort_index]["field"], "direction": "Ascending" if sort_ascending else "Descending"}], "isDefaultSort": True},
            },
            "objects": {
                "columnHeaders": [{"properties": {
                    "fontColor": solid_color(EXECUTIVE_COLORS["white"]), "backColor": solid_color(EXECUTIVE_COLORS["navy"]),
                    "bold": literal("true"), "fontSize": literal("10D"), "fontFamily": literal("'Segoe UI Semibold'"),
                    "alignment": literal("'Center'"), "wordWrap": literal("true"),
                    "autoSizeColumnWidth": literal("true"), "columnAdjustment": literal("'growToFit'"),
                    "outline": literal("'BottomOnly'"),
                }}],
                "values": [{"properties": {
                    "fontColorPrimary": solid_color(EXECUTIVE_COLORS["ink"]), "fontColorSecondary": solid_color(EXECUTIVE_COLORS["ink"]),
                    "backColorPrimary": solid_color("#FFFFFF"), "backColorSecondary": solid_color("#EAF2FB"),
                    # Tên khách hàng dài tới 51 ký tự: cho xuống dòng thay vì cắt bằng dấu ba chấm.
                    "fontSize": literal("9D"), "fontFamily": literal("'Segoe UI'"), "wordWrap": literal("true"), "outline": literal("'BottomOnly'"),
                }}],
                "grid": [{"properties": {
                    "gridVertical": literal("true"), "gridVerticalColor": solid_color("#D7E0EA"), "gridVerticalWeight": literal("1D"),
                    "gridHorizontal": literal("true"), "gridHorizontalColor": solid_color("#CBD5E1"),
                    "gridHorizontalWeight": literal("1D"), "rowPadding": literal("7D"), "outlineColor": solid_color("#94A3B8"), "outlineWeight": literal("1D"),
                }}],
                "total": [{"properties": {
                    "totals": literal("true"), "fontColor": solid_color(EXECUTIVE_COLORS["white"]), "backColor": solid_color(EXECUTIVE_COLORS["navy"]),
                    "bold": literal("true"), "fontSize": literal("10D"),
                }}],
            },
            "visualContainerObjects": executive_container(title, subtitle=subtitle, background="#FCFDFF", border_color="#94A3B8", shadow=False, radius=8, padding=6, title_size=11),
            "drillFilterOtherVisuals": True,
        },
    }


def build_pages(
    report_definition_dir: Path,
    current_year: int = 0,
    through_month: int = 0,
    as_of_date: str = "",
) -> list[str]:
    period_label = f"{through_month:02d}/{current_year}" if current_year and through_month else "kỳ hiện tại"
    subtitle = f"Kỳ {period_label} · Dữ liệu đến {as_of_date or 'ngày cập nhật gần nhất'} · Doanh thu: triệu đồng, làm tròn · Sản lượng: Két/Thùng/Bình"
    c = EXECUTIVE_COLORS
    page_specs: OrderedDict[str, tuple[str, list[dict[str, Any]]]] = OrderedDict()

    def rail(
        prefix: str,
        page: str,
        title: str,
        slicers: list[tuple[str, str, str, str]],
        *,
        include_date_range: bool = True,
    ) -> list[dict[str, Any]]:
        """Header + rail điều hướng + bộ lọc dọc, dùng chung cho cả sáu trang."""
        items: list[dict[str, Any]] = [
            executive_header(f"title_{prefix}", title, subtitle),
            executive_nav_panel(f"nav_{prefix}", page, 1),
            *executive_nav_buttons(page, 2),
        ]
        slicer_y = SIDEBAR_SLICER_Y
        tab_order = 10
        if include_date_range:
            items.append(executive_date_range_slicer(
                f"{prefix}_date_range", "KHOẢNG THỜI GIAN",
                SIDEBAR_X, slicer_y, SIDEBAR_WIDTH, SIDEBAR_DATE_SLICER_HEIGHT,
                tab_order,
            ))
            slicer_y += SIDEBAR_DATE_SLICER_HEIGHT + 8
            tab_order += 1

        for index, (name, entity, column, label) in enumerate(slicers):
            items.append(executive_slicer(
                name, entity, column, label,
                SIDEBAR_X, slicer_y + index * SIDEBAR_SLICER_STEP, SIDEBAR_WIDTH, SIDEBAR_SLICER_HEIGHT,
                tab_order + index, single_select=(column in {"Year", "MonthLabel"}),
            ))
        return items

    # ------------------------------------------------------------------ 01
    # DASHBOARD: trả lời "tháng này có đạt không, tiền đến từ đâu, ai kéo lùi".
    overview: list[dict[str, Any]] = rail("overview", "CEO_TongQuan", "VIKODA SELL-IN | TỔNG QUAN ĐIỀU HÀNH", [])
    # Màu thẻ mã hóa loại thông tin, không phải trang trí: số thực tế · so với
    # kế hoạch · so với cùng kỳ · dự báo. Trước đây hai thẻ đầu cùng màu xanh
    # nên nhìn như bị trùng.
    overview_cards = [
        ("actual", "Doanh thu Sell In", "ACTUAL MTD", "#1D4ED8", "#EEF3FF"),
        ("attainment", "Tỷ lệ đạt Target", "% ĐẠT TARGET", "#0F766E", "#ECF8F6"),
        ("yoy", "Tăng trưởng YoY", "TĂNG TRƯỞNG YOY", "#6D28D9", "#F3EEFF"),
        ("forecast", "Dự báo đạt Target", "DỰ BÁO % TARGET", "#B45309", "#FEF6EC"),
    ]
    for index, (name, measure_name, title, accent, tint) in enumerate(overview_cards):
        overview.append(executive_card(f"overview_card_{name}", measure_name, title, SIDEBAR_X, SIDEBAR_CARD_Y + index * SIDEBAR_CARD_STEP, SIDEBAR_CARD_WIDTH, SIDEBAR_CARD_HEIGHT, 20 + index, accent=accent, tint=tint))
    overview.extend([
        executive_combo(
            "overview_trend", "DimDate", "MonthAxis", ["Doanh thu Sell In", "Doanh thu LY", "Target"], ["Tỷ lệ đạt Target", "Tăng trưởng YoY"],
            "1 · DOANH THU ACTUAL SO VỚI CÙNG KỲ NĂM TRƯỚC VÀ KẾ HOẠCH", CONTENT_X, CONTENT_Y, CONTENT_WIDTH, ROW1_HEIGHT, 30,
            column_palette=[c["blue"], c["amber"], "#B9C5D3"], line_palette=[c["purple"], c["green"]],
            show_labels=False, scalar_axis=False, sort_by_category=True,
        ),
        executive_donut(
            "overview_mix", "DimProduct", "ProductGroup", "Doanh thu Sell In",
            "2 · CƠ CẤU DOANH THU THEO NHÓM SẢN PHẨM", COL3_X[0], ROW2_Y, COL3_WIDTH, ROW2_HEIGHT, 31,
        ),
        executive_donut(
            "overview_channel_mix", "DimCustomer", "Channel", "Doanh thu Sell In",
            "3 · CƠ CẤU DOANH THU THEO KÊNH", COL3_X[1], ROW2_Y, COL3_WIDTH, ROW2_HEIGHT, 32,
        ),
        executive_waterfall(
            "overview_gap_region", "DimTerritory", "Mien", "Khoảng cách Target",
            "4 · MIỀN NÀO TẠO CHÊNH LỆCH TARGET?", COL3_X[2], ROW2_Y, COL3_WIDTH, ROW2_HEIGHT, 33,
        ),
    ])
    page_specs["CEO_TongQuan"] = ("01. Tổng quan điều hành", overview)

    # ------------------------------------------------------------------ 02
    # ANALYSIS: tiền đi qua kênh nào, hệ thống MT nào, khách nào giữ được.
    channel: list[dict[str, Any]] = rail("channel", "Kenh_KhachHang", "VIKODA SELL-IN | KÊNH & KHÁCH HÀNG", [
        ("channel_region", "DimTerritory", "Mien", "MIỀN"),
        ("channel_name", "DimCustomer", "Channel", "KÊNH"),
        ("channel_type", "DimCustomer", "CustomerType", "LOẠI KHÁCH HÀNG"),
    ])
    channel.extend([
        executive_combo(
            "channel_performance", "DimCustomer", "Channel", ["Doanh thu Sell In", "Doanh thu LY"], ["Tăng trưởng YoY"],
            "1 · DOANH THU VÀ TĂNG TRƯỞNG THEO KÊNH", COL2_X[0], CONTENT_Y, COL2_WIDTH, ROW1_HEIGHT, 30,
            column_palette=[c["blue"], c["amber"]], line_palette=[c["purple"]],
            show_labels=False, scalar_axis=False,
        ),
        executive_bar(
            "channel_system_mt", "DimCustomer", "SystemMT", ["Doanh thu hệ thống MT"],
            "2 · DOANH THU THEO HỆ THỐNG MT", COL2_X[1], CONTENT_Y, COL2_WIDTH, ROW1_HEIGHT, 31,
            palette=[c["cyan"]], horizontal=True, long_labels=True,
            subtitle="Chỉ khách đã gắn hệ thống · 90% doanh thu còn lại chưa khai báo trong DMKH",
        ),
        executive_matrix(
            "channel_top_customers",
            [("DimCustomer", "CustomerName", "Khách hàng"), ("DimCustomer", "Channel", "Kênh"), ("DimTerritory", "Mien", "Miền")],
            ["Doanh thu Sell In", "Doanh thu LY", "Tăng trưởng YoY", "Đóng góp doanh thu", "Doanh thu / Điểm bán Active", "Target", "Tỷ lệ đạt Target"],
            "3 · XẾP HẠNG KHÁCH HÀNG & QUY MÔ ĐƠN HÀNG (DROP SIZE)", COL2_X[0], ROW2_Y, 678, ROW2_HEIGHT, 32,
            sort_measure="Doanh thu Sell In",
            subtitle="Sắp xếp giảm dần · Xem Doanh thu bình quân/Điểm bán & Bấm YoY % để xem khách đang tụt",
        ),
        executive_bar(
            "channel_movement", "DimCustomer", "Channel", ["Khách hàng mới", "Khách hàng ngừng mua"],
            "4 · KHÁCH MỚI VÀ KHÁCH NGỪNG MUA", 922, ROW2_Y, 342, ROW2_HEIGHT, 33,
            palette=[c["green"], c["red"]], horizontal=False,
            subtitle="So với cùng kỳ năm trước",
        ),
    ])
    page_specs["Kenh_KhachHang"] = ("02. Kênh & khách hàng", channel)

    # ------------------------------------------------------------------ 03
    # ANALYSIS: hàng Vikoda so với hàng thương mại KDT, SKU nào sống, SKU nào chết.
    product: list[dict[str, Any]] = rail("product", "SanPham", "VIKODA SELL-IN | SẢN PHẨM & DANH MỤC", [
        ("product_region", "DimTerritory", "Mien", "MIỀN"),
        ("product_group", "DimProduct", "ProductGroup", "NHÓM SẢN PHẨM"),
        ("product_unit", "DimProduct", "PackUnit", "ĐƠN VỊ TÍNH"),
    ])
    product.extend([
        # Tên SKU dài tới 52 ký tự, nên hai biểu đồ SKU được xếp thành hai cột
        # rộng 510 px ở hàng dưới thay vì ba cột 338 px — đủ chỗ cho nhãn.
        executive_combo(
            "product_vikoda_kdt", "DimDate", "MonthAxis", ["Doanh thu Vikoda", "Doanh thu KDT"], ["Tỷ trọng Vikoda"],
            "1 · HÀNG VIKODA SO VỚI HÀNG THƯƠNG MẠI KDT", CONTENT_X, CONTENT_Y, 678, ROW1_HEIGHT, 30,
            column_palette=[c["blue"], c["amber"]], line_palette=[c["teal"]],
            show_labels=False, scalar_axis=False, sort_by_category=True,
        ),
        executive_donut(
            "product_brand_mix", "DimProduct", "CoBrand", "Doanh thu Sell In",
            "2 · CƠ CẤU DOANH THU THEO THƯƠNG HIỆU", 922, CONTENT_Y, 342, ROW1_HEIGHT, 31,
        ),
        executive_bar(
            "product_top_sku", "DimProduct", "ProductAxisLabel", ["Doanh thu Sell In"],
            "3 · HERO SKUS DẪN ĐẦU DOANH THU", COL2_X[0], ROW2_Y, COL2_WIDTH, ROW2_HEIGHT, 32,
            palette=[c["blue"]], horizontal=True, long_labels=True,
        ),
        executive_bar(
            "product_declining_sku", "DimProduct", "ProductAxisLabel", ["Tăng trưởng YoY"],
            "4 · SKU TỤT MẠNH NHẤT SO VỚI CÙNG KỲ", COL2_X[1], ROW2_Y, COL2_WIDTH, ROW2_HEIGHT, 33,
            palette=[c["red"]], horizontal=True, sort_ascending=True, long_labels=True,
            subtitle="Sắp xếp tăng dần · SKU âm nhiều nhất nằm trên cùng",
        ),
    ])
    page_specs["SanPham"] = ("03. Sản phẩm & danh mục", product)

    # ------------------------------------------------------------------ 04
    # ANALYSIS: địa bàn và sản lượng quy đổi Két/Thùng/Bình.
    region: list[dict[str, Any]] = rail("region", "Vung_Mien", "VIKODA SELL-IN | VÙNG MIỀN & SẢN LƯỢNG", [
        ("region_name", "DimTerritory", "Mien", "MIỀN"),
        ("region_area", "DimTerritory", "Vung", "VÙNG"),
        ("region_unit", "DimProduct", "PackUnit", "ĐƠN VỊ TÍNH"),
    ])
    region.extend([
        executive_line(
            "region_volume_trend", "DimDate", "MonthAxis", ["SL Két (K)", "SL Thùng (T)", "SL Bình (B)"],
            "1 · SẢN LƯỢNG QUY ĐỔI THEO THỜI GIAN", COL2_X[0], CONTENT_Y, COL2_WIDTH, ROW1_HEIGHT, 30,
            palette=[c["cyan"], c["teal"], c["purple"]], show_labels=False, scalar_axis=False,
        ),
        executive_treemap(
            "region_map", "DimTerritory", "Mien", "DimTerritory", "Vung", "Doanh thu Sell In",
            "2 · BẢN ĐỒ TỶ TRỌNG DOANH THU MIỀN → VÙNG", COL2_X[1], CONTENT_Y, COL2_WIDTH, ROW1_HEIGHT, 31,
        ),
        executive_bar(
            "region_target_risk", "DimTerritory", "Vung", ["Tỷ lệ đạt Target"],
            "3 · VÙNG CẦN ƯU TIÊN XỬ LÝ", COL3_X[0], ROW2_Y, COL3_WIDTH, ROW2_HEIGHT, 32,
            palette=[c["amber"]], horizontal=True, sort_ascending=True, long_labels=True,
            subtitle="Sắp xếp tăng dần · vùng đạt thấp nhất nằm trên cùng",
        ),
        executive_donut(
            "region_volume_mix", "DimProduct", "PackUnit", "Sản lượng quy đổi KTB",
            "4 · CƠ CẤU BAO BÌ: KÉT · THÙNG · BÌNH (%)", COL3_X[1], ROW2_Y, COL3_WIDTH, ROW2_HEIGHT, 33,
        ),
        executive_bar(
            "region_coverage", "DimTerritory", "Vung", ["Số khách hàng", "Số SKU"],
            "5 · ĐỘ PHỦ KHÁCH HÀNG VÀ DANH MỤC", COL3_X[2], ROW2_Y, COL3_WIDTH, ROW2_HEIGHT, 34,
            palette=[c["teal"], c["cyan"]], horizontal=False, long_labels=True,
        ),
    ])
    page_specs["Vung_Mien"] = ("04. Vùng miền & sản lượng", region)

    # ------------------------------------------------------------------ 05
    # REPORTING: bảng phẳng để lọc, sort và xuất số.
    detail: list[dict[str, Any]] = rail("detail", "KhachHang_SanPham", "VIKODA SELL-IN | REPORTING CHI TIẾT", [
        ("detail_region", "DimTerritory", "Mien", "MIỀN"),
        ("detail_area", "DimTerritory", "Vung", "VÙNG"),
        ("detail_group", "DimProduct", "ProductGroup", "NHÓM SẢN PHẨM"),
    ])
    detail.extend([
        executive_matrix(
            "detail_matrix",
            [("DimCustomer", "CustomerName", "Khách hàng"), ("DimCustomer", "Channel", "Kênh"), ("DimProduct", "ProductShortName", "Sản phẩm"), ("DimProduct", "PackUnit", "ĐVT")],
            ["Doanh thu Sell In", "Doanh thu LY", "Tăng trưởng YoY", "SL Két (K)", "SL Thùng (T)", "SL Bình (B)", "Tỷ lệ trả hàng"],
            "REPORTING · CHI TIẾT KHÁCH HÀNG & SẢN PHẨM", CONTENT_X, CONTENT_Y, CONTENT_WIDTH, 612, 30,
            sort_measure="Doanh thu Sell In",
            subtitle="Bấm tiêu đề cột để đổi chiều sắp xếp · dùng bộ lọc bên trái để thu hẹp phạm vi",
        ),
    ])
    page_specs["KhachHang_SanPham"] = ("05. Chi tiết KH & SP", detail)

    # ------------------------------------------------------------------ 06
    # CLOSER: còn thiếu bao nhiêu, ai phải bù, và đọc báo cáo theo thứ tự nào.
    plan: list[dict[str, Any]] = rail("plan", "KeHoach_KhuyenNghi", "VIKODA SELL-IN | KẾ HOẠCH & KHUYẾN NGHỊ", [
        ("plan_region", "DimTerritory", "Mien", "MIỀN"),
        ("plan_area", "DimTerritory", "Vung", "VÙNG"),
    ])
    plan.extend([
        executive_line(
            "plan_trend", "DimDate", "MonthAxis", ["Doanh thu Sell In", "Target", "Run-rate dự báo tháng"],
            "1 · XU HƯỚNG ACTUAL · TARGET · RUN-RATE", COL2_X[0], CONTENT_Y, 678, ROW1_HEIGHT, 30,
            palette=[c["blue"], "#B9C5D3", c["teal"]], show_labels=False, scalar_axis=False,
        ),
        executive_bar(
            "plan_forecast", "DimTerritory", "Mien", ["Dự báo đạt Target"],
            "2 · DỰ BÁO KHẢ NĂNG HOÀN THÀNH", 922, CONTENT_Y, 342, ROW1_HEIGHT, 31,
            palette=[c["teal"]], horizontal=False, sort_ascending=True, long_labels=True,
        ),
        executive_bar(
            "plan_shortfall", "DimTerritory", "Vung", ["Còn thiếu để đạt Target", "Cần doanh thu mỗi ngày"],
            "3 · GAP CẦN BÙ VÀ ÁP LỰC MỖI NGÀY THEO VÙNG", COL2_X[0], ROW2_Y, COL2_WIDTH, ROW2_HEIGHT, 32,
            palette=[c["red"], c["amber"]], horizontal=True, long_labels=True,
        ),
        executive_insight_panel(
            "plan_recommendation",
            "4 · HƯỚNG DẪN ĐIỀU HÀNH 4 BƯỚC (BEVERAGE ACTION PLAYBOOK)",
            [
                ("Bước 1 — Đánh giá Nhịp độ & Dự báo Mùa vụ", "Trang 01 & 06: Xem 'Nhịp độ bán hàng (Pacing %)' và 'Dự báo EOM Mùa vụ'. Nếu Pacing < 100% kết hợp Hệ số gia tốc > 1.3x thì vùng đang chậm nhịp nghiêm trọng, cần kích hoạt chương trình Trade Promotion."),
                ("Bước 2 — Truy nguyên nhân theo Kênh & Bao bì", "Trang 02 xem kênh GT/MT/KA nào tụt YoY; trang 03 rà soát Hero SKUs (Vikoda 500ml, 1.5L, Bình 19L); trang 04 kiểm tra cơ cấu Két - Thùng - Bình và vòng quay cọc vỏ."),
                ("Bước 3 — Giao chỉ tiêu & Áp lực ngày", "Biểu đồ 3 bên trái chỉ rõ Gap doanh thu cần bù và Mức doanh thu mỗi ngày phải đạt cho từng Quản lý vùng (RSM) để kịp cán đích trước khi đóng sổ tháng."),
                ("Bước 4 — Bảo vệ tệp Điểm bán & Khách hàng", "Rà soát danh sách NPP ngừng mua ở trang 02 để sales rep tiếp cận ngay, đồng thời theo dõi Quy mô đơn hàng (Drop Size) để nâng giá trị đơn trên mỗi khách hàng active."),
            ],
            COL2_X[1], ROW2_Y, COL2_WIDTH, ROW2_HEIGHT, 33,
        ),
    ])
    page_specs["KeHoach_KhuyenNghi"] = ("06. Kế hoạch & khuyến nghị", plan)

    page_names = list(page_specs.keys())
    for page_name, (display_name, visuals) in page_specs.items():
        page_folder = report_definition_dir / "pages" / page_name
        page_folder.mkdir(parents=True, exist_ok=True)
        write_json(page_folder / "page.json", {
            "$schema": f"{REPORT_SCHEMA}/page/2.0.0/schema.json",
            "name": page_name,
            "displayName": display_name,
            "displayOption": "FitToPage",
            "width": 1280,
            "height": 720,
            "objects": {"background": [{"properties": {"color": solid_color(c["canvas"]), "transparency": literal("0D")}}]},
        })
        for visual in visuals:
            write_json(page_folder / "visuals" / visual["name"] / "visual.json", visual)
    write_json(report_definition_dir / "pages" / "pages.json", {
        "$schema": f"{REPORT_SCHEMA}/pagesMetadata/1.0.0/schema.json",
        "pageOrder": page_names,
        "activePageName": page_names[0],
    })
    return page_names


def theme_json() -> dict[str, Any]:
    return {
        "name": "Vikoda Executive Pro 2026",
        "dataColors": ["#2563EB", "#B9C5D3", "#0F766E", "#D97706", "#7C3AED", "#16A34A", "#DC2626", "#0891B2"],
        "background": "#F3F6FA",
        "foreground": "#172033",
        "tableAccent": "#102A43",
        "good": "#16A34A",
        "neutral": "#D97706",
        "bad": "#DC2626",
        "visualStyles": {
            "*": {
                "*": {
                    "title": [{"show": True, "fontFamily": "Segoe UI Semibold", "fontSize": 11, "fontColor": {"solid": {"color": "#172033"}}}],
                    "background": [{"show": True, "color": {"solid": {"color": "#FFFFFF"}}, "transparency": 0}],
                    "border": [{"show": False}],
                    "visualHeader": [{"show": False}],
                    "legend": [{"show": True, "position": "Top", "fontFamily": "Segoe UI", "fontSize": 9, "fontColor": {"solid": {"color": "#64748B"}}}],
                }
            },
            "card": {
                "*": {
                    "categoryLabels": [{"show": True, "fontFamily": "Segoe UI", "fontSize": 10, "color": {"solid": {"color": "#64748B"}}}],
                    "calloutValue": [{"fontFamily": "Segoe UI Semibold", "fontSize": 25, "color": {"solid": {"color": "#2563EB"}}}],
                }
            },
            "slicer": {"*": {"header": [{"show": False}], "items": [{
                "fontFamily": "Segoe UI Semibold", "fontSize": 10,
                "fontColor": {"solid": {"color": "#7DD3FC"}},
                "backColor": {"solid": {"color": "#123A5D"}},
            }] }},
            "pivotTable": {"*": {"grid": [{"gridVertical": False, "gridHorizontal": True, "rowPadding": 6}] }},
        },
    }


def dax_measure_file() -> str:
    lines = [
        "// Vikoda Sell In — measures generated from build_powerbi_package.py",
        "// Refresh rule: rerun the report launcher, then open the .pbip and Refresh.",
        "",
    ]
    for item in build_measures():
        lines.append(f"{item['name']} = {item['expression']}")
        lines.append("")
    return "\n".join(lines)


def manifest_payload(output_dir: Path, current_year: int, through_month: int, as_of_date: str, counts: dict[str, int], totals: dict[str, float], pages: list[str]) -> dict[str, Any]:
    return {
        "package_version": PACKAGE_VERSION,
        "generated_at": datetime.now().astimezone().isoformat(),
        "as_of_date": as_of_date,
        "period": f"{through_month:02d}/{current_year}",
        "project": f"{REPORT_NAME}.pbip",
        "refresh": {
            "power_bi_desktop": "Open the .pbip, choose Refresh, and save.",
            "source_of_truth": "Data/File bao cao/PowerBI/Data/*.csv",
            "staging_source": "Data/Work/bao_cao/*/staging/*.json",
        },
        "tables": counts,
        "totals_vnd": totals,
        "pages": pages,
        "dataxan_reference": "https://dataxan.com/vi/mau-bao-cao-power-bi/",
    }


def build_package(
    sell_in_path: Path,
    target_path: Path,
    dmkh_path: Path,
    output_dir: Path,
    product_catalog_path: Path | None = None,
) -> dict[str, Any]:
    sell_in = load_json(sell_in_path)
    target = load_json(target_path)
    dmkh = load_json(dmkh_path)
    product_catalog = load_product_catalog(product_catalog_path)
    dim_date, dim_customer, dim_product, dim_territory, fact_rows = build_dimensions(sell_in, target, dmkh, product_catalog)
    fact_sell_in, fact_target = fact_rows
    data_dir = output_dir / "Data"
    if data_dir.exists():
        # Only generated output is removed; source/staging directories are untouched.
        shutil.rmtree(data_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    columns = {
        "DimDate": list(dim_date[0].keys()) if dim_date else [],
        "DimCustomer": list(dim_customer[0].keys()) if dim_customer else [],
        "DimProduct": list(dim_product[0].keys()) if dim_product else [],
        "DimTerritory": list(dim_territory[0].keys()) if dim_territory else [],
        "FactSellIn": list(fact_sell_in[0].keys()) if fact_sell_in else [],
        "FactTarget": list(fact_target[0].keys()) if fact_target else [],
    }
    for name, rows in [("DimDate", dim_date), ("DimCustomer", dim_customer), ("DimProduct", dim_product), ("DimTerritory", dim_territory), ("FactSellIn", fact_sell_in), ("FactTarget", fact_target)]:
        write_csv(data_dir / f"{name}.csv", rows, columns[name])

    report_dir = output_dir / f"{REPORT_NAME}.Report"
    model_dir = output_dir / f"{REPORT_NAME}.SemanticModel"

    # Thư mục `.pbi` giữ chữ ký cho phép Power BI Desktop đọc file CSV trên máy
    # này. Xóa đi thì mỗi lần mở lại phải bấm đồng ý quyền truy cập. Giữ lại
    # phần cấu hình, nhưng KHÔNG giữ `cache.abf`: cache là số liệu của lần
    # trước, để lại sẽ hiện số cũ trước khi kịp Refresh.
    preserved_settings = preserve_local_settings(report_dir, model_dir)

    for path in (report_dir, model_dir):
        if path.exists():
            shutil.rmtree(path)
    definition_dir = report_dir / "definition"
    definition_dir.mkdir(parents=True, exist_ok=True)
    report_static = report_dir / "StaticResources" / "RegisteredResources"
    report_static.mkdir(parents=True, exist_ok=True)
    write_json(report_static / "VikodaTheme.json", theme_json())
    write_json(report_dir / ".platform", platform_payload("Report", REPORT_NAME))
    write_json(report_dir / "definition.pbir", {
        "$schema": "https://developer.microsoft.com/json-schemas/fabric/item/report/definitionProperties/2.0.0/schema.json",
        "version": "4.0",
        "datasetReference": {"byPath": {"path": f"../{REPORT_NAME}.SemanticModel"}},
    })
    write_json(definition_dir / "version.json", {
        "$schema": f"{REPORT_SCHEMA}/versionMetadata/1.0.0/schema.json",
        "version": "2.0.0",
    })
    write_json(definition_dir / "report.json", {
        "$schema": f"{REPORT_SCHEMA}/report/3.0.0/schema.json",
        "themeCollection": {},
        "objects": {
            "section": [{"properties": {"verticalAlignment": {"expr": {"Literal": {"Value": "'Top'"}}}}}],
            "outspacePane": [{"properties": {"expanded": {"expr": {"Literal": {"Value": "false"}}}}}],
        },
        "settings": {"useStylableVisualContainerHeader": True, "exportDataMode": "AllowSummarized", "defaultDrillFilterOtherVisuals": True, "allowChangeFilterTypes": True, "useEnhancedTooltips": True, "useDefaultAggregateDisplayName": True},
    })
    pages = build_pages(
        definition_dir,
        int(sell_in.get("current_year", 0)),
        int(sell_in.get("through_month", 0)),
        str(sell_in.get("as_of_date") or ""),
    )
    model_dir.mkdir(parents=True, exist_ok=True)
    write_json(model_dir / ".platform", platform_payload("SemanticModel", REPORT_NAME))
    model_bim = model_definition(data_dir, {name: len(rows) for name, rows in [("DimDate", dim_date), ("DimCustomer", dim_customer), ("DimProduct", dim_product), ("DimTerritory", dim_territory), ("FactSellIn", fact_sell_in), ("FactTarget", fact_target)]})
    write_json(model_dir / "definition.pbism", {
        "$schema": f"{SEMANTIC_SCHEMA}/definitionProperties/1.0.0/schema.json",
        "version": "1.0",
    })
    write_json(model_dir / "model.bim", model_bim)
    write_json(output_dir / f"{REPORT_NAME}.pbip", {
        "$schema": "https://developer.microsoft.com/json-schemas/fabric/pbip/pbipProperties/1.0.0/schema.json",
        "version": "1.0",
        "artifacts": [{"report": {"path": f"{REPORT_NAME}.Report"}}],
        "settings": {"enableAutoRecovery": True},
    })
    write_text(output_dir / "DAX_Measures.dax", dax_measure_file())
    write_text(output_dir / "README.md", f"""# Vikoda Sell In — Dashboard Power BI

Ngày dữ liệu mới nhất: {sell_in.get('as_of_date', '')}. Kỳ báo cáo: {int(sell_in.get('through_month', 0)):02d}/{sell_in.get('current_year', '')}.

## Mở và làm mới

1. Nhấp đúp `Chay CT\\Bao cao Power BI.cmd`. Script tự kiểm tra ERP/Target/DMKH, chạy lại các chặng bị cũ, dựng lại gói này rồi mở Power BI Desktop.
2. Trong Power BI Desktop, chọn **Refresh** để semantic model đọc các file CSV UTF-8 trong `Data/`.
3. Xong. Không cần **Save As**, không cần nhớ thứ tự chạy file nào trước.

`{REPORT_NAME}.pbip` là nguồn chuẩn duy nhất, dạng text, sinh lại tự động. Dự án cố tình không giữ file `.pbix`: bản chụp đó không tự cập nhật theo dữ liệu nên chỉ tạo ra rủi ro đọc nhầm số cũ. Cần chia sẻ cho người khác thì **Publish** lên Power BI Service.

## Các trang báo cáo

Sáu trang kể một mạch liền: đạt hay hụt → vì kênh nào → vì sản phẩm nào → vì địa
bàn nào → số chi tiết → làm gì tiếp.

1. **Tổng quan điều hành** — bốn KPI, Actual–cùng kỳ LY–Target theo tháng kèm `% đạt` và `YoY %`, cơ cấu doanh thu theo nhóm sản phẩm và theo kênh, waterfall miền tạo chênh lệch.
2. **Kênh & khách hàng** — doanh thu và tăng trưởng theo kênh, doanh thu theo hệ thống MT, bảng xếp hạng khách hàng, khách mới so với khách ngừng mua.
3. **Sản phẩm & danh mục** — hàng Vikoda so với hàng thương mại KDT theo tháng, SKU dẫn đầu, SKU tụt mạnh nhất, cơ cấu theo thương hiệu.
4. **Vùng miền & sản lượng** — xu hướng Két/Thùng/Bình, treemap tỷ trọng Miền → Vùng, vùng đạt Target thấp nhất, cơ cấu sản lượng, độ phủ khách hàng và danh mục.
5. **Chi tiết KH & SP** — trang Reporting dạng bảng đầy đủ để lọc, sắp xếp và xuất số.
6. **Kế hoạch & khuyến nghị** — run-rate so với Target, dự báo hoàn thành theo miền, gap cần bù và áp lực doanh thu mỗi ngày theo vùng, kèm khối hướng dẫn đọc số và giao việc.

Giao diện dùng canvas executive 1280×720 theo phương pháp DAR của Datapot: Dashboard → Analysis → Reporting. KPI chỉ nằm ở sidebar trang Tổng quan; các trang sau dành toàn bộ vùng nội dung cho một chủ đề phân tích. Sidebar có sáu nút Page Navigation có thể bấm, tô cyan cho trang hiện tại. Slicer dùng nền navy, giá trị chọn màu cyan và cùng lưới 196×60 px; dropdown hiện trọn chữ, cách đều 8 px và tách khỏi KPI. Trục X/Y có tiêu đề, gridline và mật độ nhãn rõ; trục tháng dùng nhãn phân loại `YY/MM` để loại khoảng trắng giữa các kỳ. Bảng Reporting tự giãn gần kín trang, dùng zebra fill và dòng tổng navy. Doanh thu theo triệu đồng và sản lượng quản trị đều hiển thị không có phần thập phân.

Package gồm sáu bảng theo mô hình hình sao, 51 measure DAX và bảy quan hệ. Sản lượng K/T/B lấy `Số lượng ÷ Quy cách` từ DMSP, có cùng kỳ và tăng trưởng YoY riêng cho từng đơn vị. Trang Tổng quan so sánh Actual, cùng kỳ LY và Target bằng cột dọc, đồng thời hiển thị `% đạt` và `YoY %`; donut chỉ dùng cho cơ cấu, waterfall dùng để chỉ ra đóng góp dương/âm vào gap, treemap thay cho bản đồ địa lý.

Chưa dựng được bản đồ tỉnh/thành vì cột `Tỉnh/Thành` trong DMKH gần như bỏ trống
(hiện chỉ vài khách có giá trị). Khi DMKH khai báo đủ tỉnh, đổi treemap trang 04
sang `filledMap` là dùng được ngay, không phải sửa mô hình.

Sau khi thay ERP/Target/DMKH/DMSP, chạy `Chay CT\\Bao cao Power BI.cmd` hoặc `Chay CT\\Bao cao Target.cmd` — cả hai đều tự phát hiện chặng nào bị cũ và chạy lại đúng phần cần thiết. Không chỉnh tay các file trong `Data/` vì lần chạy sau sẽ ghi lại.

Hướng visual tham khảo DaTaxan và phương pháp DAR của Datapot: https://dataxan.com/vi/mau-bao-cao-power-bi/ · https://datapot.vn/xay-dung-bao-cao-power-bi-theo-phuong-phap-dar/
""")
    current_period = f"{int(sell_in['current_year'])}{int(sell_in['through_month']):02d}"
    prior_year, prior_month = previous_month(int(sell_in["current_year"]), int(sell_in["through_month"]))
    prior_period = f"{prior_year}{prior_month:02d}"
    last_year_period = f"{int(sell_in['current_year']) - 1}{int(sell_in['through_month']):02d}"
    totals = {
        "actual": sum(row["RevenueVND"] for row in fact_sell_in if row["PeriodKey"] == current_period),
        "target": sum(row["TargetTotalVND"] for row in fact_target if row["PeriodKey"] == current_period),
        "vikoda": sum(row["RevenueVND"] for row in fact_sell_in if row["PeriodKey"] == current_period and row["IsVikoda"]),
        "target_vikoda": sum(row["TargetVikodaVND"] for row in fact_target if row["PeriodKey"] == current_period),
        "last_year": sum(row["RevenueVND"] for row in fact_sell_in if row["PeriodKey"] == last_year_period),
        "previous_month": sum(row["RevenueVND"] for row in fact_sell_in if row["PeriodKey"] == prior_period),
    }
    counts = {"DimDate": len(dim_date), "DimCustomer": len(dim_customer), "DimProduct": len(dim_product), "DimTerritory": len(dim_territory), "FactSellIn": len(fact_sell_in), "FactTarget": len(fact_target)}
    manifest = manifest_payload(output_dir, int(sell_in["current_year"]), int(sell_in["through_month"]), str(sell_in.get("as_of_date") or ""), counts, totals, pages)
    unconverted_products = [row for row in dim_product if not optional_positive_float(row.get("PackSize"))]
    current_converted = {
        unit: sum(
            safe_float(row.get("ConvertedQuantity"))
            for row in fact_sell_in
            if row["PeriodKey"] == current_period and row.get("PackUnit") == unit and row.get("ConvertedQuantity") is not None
        )
        for unit in ("Két", "Thùng", "Bình")
    }
    manifest["product_catalog"] = {
        "source": str(product_catalog_path or ""),
        "catalog_codes": len(product_catalog),
        "model_products": len(dim_product),
        "matched_products": sum(1 for row in dim_product if row.get("ProductCode") in product_catalog),
        "convertible_products": len(dim_product) - len(unconverted_products),
        "unconverted_products": [{"code": row.get("ProductCode"), "name": row.get("ProductName")} for row in unconverted_products],
        "current_period_converted_quantity": current_converted,
        "current_period_unconverted_quantity": sum(
            safe_float(row.get("Quantity"))
            for row in fact_sell_in
            if row["PeriodKey"] == current_period and row.get("ConvertedQuantity") is None
        ),
    }
    manifest["display_unit"] = "million_vnd"
    restore_local_settings(preserved_settings)
    removed_pbix = remove_stale_pbix(output_dir)
    if removed_pbix:
        print(
            f"Da xoa {removed_pbix}: quy trinh moi chi dung PBIP lam nguon duy nhat.",
            file=sys.stderr,
        )
    manifest["standalone_pbix"] = False
    write_json(output_dir / "refresh_manifest.json", manifest)
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sell-in-data-file", type=Path, required=True)
    parser.add_argument("--target-data-file", type=Path, required=True)
    parser.add_argument("--dmkh-data-file", type=Path, required=True)
    parser.add_argument("--product-catalog-file", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = build_package(
        args.sell_in_data_file,
        args.target_data_file,
        args.dmkh_data_file,
        args.output_dir,
        args.product_catalog_file,
    )
    print(json.dumps({"output_dir": str(args.output_dir), "tables": manifest["tables"], "pages": manifest["pages"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
