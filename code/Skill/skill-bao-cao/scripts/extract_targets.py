from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

TARGET_SOURCE_POLICY = "ANNUAL_TWO_SOURCE_VBA_MTKA"


def normalized_key(value: Any) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.replace("\u00a0", " ").split())
    text = unicodedata.normalize("NFKD", text)
    text = "".join(character for character in text if not unicodedata.combining(character))
    return text.replace("Đ", "D").replace("đ", "d").upper()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


REPORTING_REGIONS = {
    "BAC MIEN TRUNG": ("Miền Bắc", "Bắc Miền Trung"),
    "DONG BAC": ("Miền Bắc", "Đông Bắc"),
    "HA NOI": ("Miền Bắc", "Hà Nội"),
    "TAY BAC": ("Miền Bắc", "Tây Bắc"),
    "MIEN DONG": ("Miền Nam", "Miền Đông"),
    "MIEN TAY": ("Miền Nam", "Miền Tây"),
    "TP. HCM 1": ("Miền Nam", "TP. HCM 1"),
    "TP HCM 1": ("Miền Nam", "TP. HCM 1"),
    "TP. HCM 2": ("Miền Nam", "TP. HCM 2"),
    "TP HCM 2": ("Miền Nam", "TP. HCM 2"),
    "MIEN TRUNG 1A": ("Miền Trung 1", "Miền Trung 1A"),
    "MIEN TRUNG 1B": ("Miền Trung 1", "Miền Trung 1B"),
    "TAY NGUYEN": ("Miền Trung 1", "Tây Nguyên"),
    "MIEN TRUNG 2A": ("Miền Trung 2", "Miền Trung 2A"),
    "MIEN TRUNG 2B": ("Miền Trung 2", "Miền Trung 2B"),
    "KA MIEN BAC": ("KA", "KA Miền Bắc"),
    "KA MIEN TRUNG 1": ("KA", "KA Miền Trung 1"),
    "KA MIEN TRUNG 2": ("KA", "KA Miền Trung 2"),
    "KA MIEN NAM": ("KA", "KA Miền Nam"),
    "MT": ("MT", "MT"),
    "B2C": ("B2C", "B2C"),
    "OTHER": ("Other", "Other"),
    "XK": ("Other", "Other"),
}


def normalize_reporting_pair(area: Any, region: Any) -> tuple[str, str]:
    area_text = clean_text(area)
    region_text = clean_text(region)
    region_key = normalized_key(region_text)
    area_key = normalized_key(area_text)

    if area_key == "KA" and region_key in {"MIEN TRUNG 1", "MIEN TRUNG 1A"}:
        region_key = "KA MIEN TRUNG 1"
    elif area_key == "KA" and region_key in {"MIEN TRUNG 2", "MIEN TRUNG 2A"}:
        region_key = "KA MIEN TRUNG 2"
    elif area_key == "KA" and region_key == "MIEN BAC":
        region_key = "KA MIEN BAC"
    elif area_key == "KA" and region_key == "MIEN NAM":
        region_key = "KA MIEN NAM"

    if region_key in REPORTING_REGIONS:
        return REPORTING_REGIONS[region_key]
    if area_key in REPORTING_REGIONS:
        return REPORTING_REGIONS[area_key]
    return ("Other", "Other")


def parse_year_from_name(path: Path) -> int | None:
    match = re.search(r"(20\d{2})", path.stem)
    return int(match.group(1)) if match else None


def as_decimal(value: Any, *, blank_is_zero: bool = True) -> Decimal:
    if value is None or clean_text(value) == "":
        return Decimal(0) if blank_is_zero else Decimal("NaN")
    if isinstance(value, bool):
        raise ValueError("Giá trị Boolean không hợp lệ cho Target")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = clean_text(value)
    if text.startswith("#"):
        raise ValueError(f"Lỗi Excel trong ô Target: {text}")
    try:
        return Decimal(text.replace(",", ""))
    except InvalidOperation as error:
        raise ValueError(f"Target không phải số: {text}") from error


def millions_to_vnd(value: Any) -> int:
    amount = as_decimal(value) * Decimal(1_000_000)
    return int(amount.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def vnd_value(value: Any) -> int:
    amount = as_decimal(value, blank_is_zero=False)
    if amount.is_nan():
        raise ValueError("Target chuẩn không được để trống")
    return int(amount.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def iso_file_time(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).astimezone().isoformat()


def find_header_row(ws, required_keys: set[str], max_rows: int = 20) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(max_rows, ws.max_row) + 1):
        mapping: dict[str, int] = {}
        for cell in ws[row_number]:
            key = normalized_key(cell.value)
            if key:
                mapping.setdefault(key, cell.column)
        if all(any(required in key for key in mapping) for required in required_keys):
            return row_number, mapping
    raise ValueError(
        f"Không tìm thấy hàng tiêu đề chứa: {', '.join(sorted(required_keys))} "
        f"trong sheet {ws.title}"
    )


def find_column(mapping: dict[str, int], predicate) -> int:
    matches = [column for key, column in mapping.items() if predicate(key)]
    if not matches:
        raise ValueError("Không tìm thấy cột bắt buộc")
    return matches[0]


def add_record(
    accumulator: dict[tuple[str, str, str], dict[str, Any]],
    *,
    year: int,
    month: int,
    code: str,
    name: str,
    target_vikoda: int,
    target_total: int,
    source_file: str,
    source_updated: str,
    reporting_area: str = "",
    reporting_region: str = "",
) -> None:
    period = f"{year:04d}{month:02d}"
    key = (period, code, name)
    if key not in accumulator:
        accumulator[key] = {
            "Ky": period,
            "Nam": year,
            "Thang": month,
            "MaKhachHangMoi": code,
            "TenKhachHang": name,
            "TargetVikoda": 0,
            "TargetTong": 0,
            "_sources": set(),
            "_updated": [],
            "_reporting_region_totals": defaultdict(
                lambda: {"target_total": 0, "target_vikoda": 0, "records": 0}
            ),
        }
    record = accumulator[key]
    record["TargetVikoda"] += target_vikoda
    record["TargetTong"] += target_total
    record["_sources"].add(source_file)
    record["_updated"].append(source_updated)
    area, region = normalize_reporting_pair(reporting_area, reporting_region)
    region_item = record["_reporting_region_totals"][(area, region)]
    region_item["target_total"] += target_total
    region_item["target_vikoda"] += target_vikoda
    region_item["records"] += 1


def extract_npp(
    path: Path,
    accumulator: dict[tuple[str, str, str], dict[str, Any]],
    problems: list[str],
    warnings: list[str],
) -> dict[str, Any]:
    year = parse_year_from_name(path)
    if year is None:
        raise ValueError(f"Không xác định được năm từ tên file: {path.name}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = next(
            (
                sheet
                for sheet in workbook.worksheets
                if "SELL IN" in normalized_key(sheet["A2"].value)
                or "TINH KDT" in normalized_key(sheet.title)
            ),
            workbook.worksheets[0],
        )
        header_row, mapping = find_header_row(
            worksheet,
            {"MA NPP", "TEN NPP"},
            max_rows=15,
        )
        code_col = find_column(mapping, lambda key: key == "MA NPP")
        name_col = find_column(mapping, lambda key: key == "TEN NPP")
        area_col = find_column(mapping, lambda key: key == "MIEN")
        region_col = find_column(mapping, lambda key: key == "VUNG")
        month_row = header_row - 1
        label_row = header_row
        month_columns: dict[int, tuple[int, int]] = {}
        for column in range(1, worksheet.max_column + 1):
            month = MONTHS.get(normalized_key(worksheet.cell(month_row, column).value))
            if not month:
                continue
            block = {
                normalized_key(worksheet.cell(label_row, candidate).value): candidate
                for candidate in range(column, min(column + 5, worksheet.max_column + 1))
            }
            vikoda_col = find_column(block, lambda key: key == "VIKODA")
            total_col = find_column(block, lambda key: key == "TONG CONG")
            month_columns[month] = (vikoda_col, total_col)
        if set(month_columns) != set(range(1, 13)):
            raise ValueError(
                f"{path.name}: không tìm đủ 12 nhóm cột tháng, tìm thấy {sorted(month_columns)}"
            )

        data_start = header_row + 2
        max_needed_col = max(
            [code_col, name_col, area_col, region_col]
            + [column for pair in month_columns.values() for column in pair]
        )
        customer_rows = 0
        normalized_rows = 0
        special_rows: list[dict[str, Any]] = []
        found_special_types: set[str] = set()
        updated = iso_file_time(path)
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=data_start,
                min_col=1,
                max_col=max_needed_col,
                values_only=True,
            ),
            start=data_start,
        ):
            code = clean_text(values[code_col - 1])
            name = clean_text(values[name_col - 1])
            if not code and not name:
                continue
            reporting_area = clean_text(values[area_col - 1])
            reporting_region = clean_text(values[region_col - 1])
            name_key = normalized_key(name)
            area_key = normalized_key(reporting_area)
            region_key = normalized_key(reporting_region)
            special_type = ""
            if not code:
                if name_key == "B2C" and (
                    area_key == "B2C" or region_key == "B2C"
                ):
                    code = "B2C"
                    special_type = "B2C"
                elif name_key == "OTHER" and (
                    area_key == "OTHER" or region_key == "OTHER"
                ):
                    code = "0"
                    special_type = "OTHER"
                else:
                    continue
            if not name:
                problems.append(f"{path.name}!R{row_number}: thiếu TÊN NPP")
                continue
            if special_type:
                found_special_types.add(special_type)
                special_rows.append(
                    {
                        "row": row_number,
                        "type": special_type,
                        "code": code,
                        "name": name,
                    }
                )
            customer_rows += 1
            for month, (vikoda_col, total_col) in month_columns.items():
                try:
                    target_vikoda = millions_to_vnd(values[vikoda_col - 1])
                    target_total = millions_to_vnd(values[total_col - 1])
                except ValueError as error:
                    problems.append(f"{path.name}!R{row_number}: {error}")
                    continue
                if target_vikoda < 0 or target_total < 0:
                    warnings.append(
                        f"{path.name}!R{row_number} kỳ {year}{month:02d}: "
                        "Target âm được chuẩn hóa thành 0"
                    )
                    target_vikoda = max(0, target_vikoda)
                    target_total = max(0, target_total)
                add_record(
                    accumulator,
                    year=year,
                    month=month,
                    code=code,
                    name=name,
                    target_vikoda=target_vikoda,
                    target_total=target_total,
                    source_file=path.name,
                    source_updated=updated,
                    reporting_area=reporting_area,
                    reporting_region=reporting_region,
                )
                normalized_rows += 1
        missing_special_types = {"B2C", "OTHER"} - found_special_types
        if missing_special_types:
            problems.append(
                f"{path.name}: thiếu dòng Target đặc biệt "
                + ", ".join(sorted(missing_special_types))
            )
        return {
            "file": path.name,
            "year": year,
            "customer_rows": customer_rows,
            "normalized_rows": normalized_rows,
            "special_rows": special_rows,
        }
    finally:
        workbook.close()


def exact_period_value(value: Any) -> str | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        number = int(Decimal(str(value)))
        text = str(number)
    else:
        text = clean_text(value)
        try:
            text = str(int(Decimal(text.replace(",", ""))))
        except InvalidOperation:
            return None
    if re.fullmatch(r"20\d{2}(0[1-9]|1[0-2])", text):
        return text
    return None


def extract_mtka_vba_data(
    path: Path,
    periods: set[str],
    accumulator: dict[tuple[str, str, str], dict[str, Any]],
    problems: list[str],
    warnings: list[str],
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "DATA" not in workbook.sheetnames:
            raise ValueError(f"{path.name} không có sheet DATA")
        worksheet = workbook["DATA"]
        groups: dict[tuple[str, str], dict[str, Any]] = {}
        source_rows: dict[str, int] = defaultdict(int)

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=21,
                values_only=True,
            ),
            start=2,
        ):
            area = clean_text(values[0])
            area_key = normalized_key(area)
            if area_key not in {"KA", "MT"}:
                continue
            period = exact_period_value(values[17])
            if period not in periods:
                continue
            code = clean_text(values[3])
            if not code:
                continue

            name = clean_text(values[4])
            product_group = normalized_key(values[6])
            try:
                target_value = as_decimal(
                    values[20],
                    blank_is_zero=False,
                )
            except ValueError as error:
                problems.append(f"{path.name}!R{row_number}: {error}")
                continue

            key = (period, code)
            item = groups.setdefault(
                key,
                {
                    "name": "",
                    "area": area,
                    "region": clean_text(values[1]),
                    "target_vikoda": Decimal(0),
                    "target_total": Decimal(0),
                    "source_rows": 0,
                },
            )
            if not item["name"] and name:
                item["name"] = name
            item["source_rows"] += 1
            source_rows[period] += 1
            if target_value.is_nan():
                continue
            item["target_total"] += target_value
            if "VIKODA" in product_group:
                item["target_vikoda"] += target_value

        updated = iso_file_time(path)
        grouped_rows: dict[str, int] = defaultdict(int)
        for (period, code), item in groups.items():
            target_vikoda = int(
                item["target_vikoda"].quantize(
                    Decimal("1"),
                    rounding=ROUND_HALF_UP,
                )
            )
            target_total = int(
                item["target_total"].quantize(
                    Decimal("1"),
                    rounding=ROUND_HALF_UP,
                )
            )
            if target_vikoda < 0 or target_total < 0:
                warnings.append(
                    f"{path.name} kỳ {period}, mã {code}: "
                    "Target âm được chuẩn hóa thành 0"
                )
                target_vikoda = max(0, target_vikoda)
                target_total = max(0, target_total)
            add_record(
                accumulator,
                year=int(period[:4]),
                month=int(period[4:]),
                code=code,
                name=item["name"],
                target_vikoda=target_vikoda,
                target_total=target_total,
                source_file=path.name,
                source_updated=updated,
                reporting_area=item["area"],
                reporting_region=item["region"],
            )
            grouped_rows[period] += 1

        return [
            {
                "file": path.name,
                "period": period,
                "source_rows": source_rows.get(period, 0),
                "grouped_customers": grouped_rows.get(period, 0),
                "algorithm": (
                    "VBA Final V6: DATA A/D/E/G/R/U; "
                    "lọc KA/MT, nhóm theo mã, cộng U trực tiếp"
                ),
            }
            for period in sorted(periods)
        ]
    finally:
        workbook.close()


def target_record_counter(records: list[dict[str, Any]]) -> Counter:
    return Counter(
        (
            record["MaKhachHangMoi"],
            record["TenKhachHang"],
            int(record["TargetVikoda"]),
            int(record["TargetTong"]),
        )
        for record in records
    )


def extract_canonical(path: Path, problems: list[str]) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        if "DATA_TARGET" not in workbook.sheetnames:
            raise ValueError(f"{path.name} không có sheet DATA_TARGET")
        worksheet = workbook["DATA_TARGET"]
        period = None
        updated = None
        for row in worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True):
            for value in row:
                text = clean_text(value)
                period_match = re.search(r"Period:\s*(20\d{2})(0[1-9]|1[0-2])", text, re.IGNORECASE)
                if period_match:
                    period = f"{period_match.group(1)}{period_match.group(2)}"
                updated_match = re.search(
                    r"Updated:\s*(\d{1,2})/(\d{1,2})/(20\d{2})\s+(\d{1,2}):(\d{2}):(\d{2})",
                    text,
                    re.IGNORECASE,
                )
                if updated_match:
                    day, month, year, hour, minute, second = map(int, updated_match.groups())
                    updated = datetime(year, month, day, hour, minute, second).astimezone().isoformat()
        if period is None:
            raise ValueError(f"{path.name}: không tìm thấy Period: YYYYMM")
        if updated is None:
            updated = iso_file_time(path)

        header_row, mapping = find_header_row(
            worksheet,
            {"MAKHACHHANGMOI", "TENKHACHHANG", "TARGET VIKODA", "TARGET TONG"},
            max_rows=20,
        )
        code_col = find_column(mapping, lambda key: key == "MAKHACHHANGMOI")
        name_col = find_column(mapping, lambda key: key == "TENKHACHHANG")
        vikoda_col = find_column(mapping, lambda key: key == "TARGET VIKODA")
        total_col = find_column(mapping, lambda key: key == "TARGET TONG")
        max_needed_col = max(code_col, name_col, vikoda_col, total_col)
        year = int(period[:4])
        month = int(period[4:])
        records: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=header_row + 1,
                min_col=1,
                max_col=max_needed_col,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            code = clean_text(values[code_col - 1])
            name = clean_text(values[name_col - 1])
            if not code and not name:
                continue
            if not code or not name:
                problems.append(f"{path.name}!R{row_number}: thiếu mã hoặc tên khách hàng")
                continue
            try:
                target_vikoda = vnd_value(values[vikoda_col - 1])
                target_total = vnd_value(values[total_col - 1])
            except ValueError as error:
                problems.append(f"{path.name}!R{row_number}: {error}")
                continue
            records.append(
                {
                    "Ky": period,
                    "Nam": year,
                    "Thang": month,
                    "MaKhachHangMoi": code,
                    "TenKhachHang": name,
                    "TargetVikoda": target_vikoda,
                    "TargetTong": target_total,
                    "NgayCapNhatNguon": updated,
                    "NguonFile": path.name,
                }
            )
        return period, records, {
            "file": path.name,
            "period": period,
            "records": len(records),
            "updated": updated,
        }
    finally:
        workbook.close()


def summarize_periods(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, int]] = defaultdict(
        lambda: {"records": 0, "target_vikoda": 0, "target_total": 0}
    )
    for record in records:
        item = summary[record["Ky"]]
        item["records"] += 1
        item["target_vikoda"] += record["TargetVikoda"]
        item["target_total"] += record["TargetTong"]
    return [
        {"period": period, **summary[period]}
        for period in sorted(summary)
    ]


def finalize_accumulator(accumulator: dict[tuple[str, str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for key in sorted(accumulator):
        raw = accumulator[key]
        updated = max(raw["_updated"]) if raw["_updated"] else None
        reporting_area, reporting_region = max(
            raw["_reporting_region_totals"],
            key=lambda pair: (
                raw["_reporting_region_totals"][pair]["target_total"],
                raw["_reporting_region_totals"][pair]["target_vikoda"],
                raw["_reporting_region_totals"][pair]["records"],
                pair,
            ),
            default=("Other", "Other"),
        )
        records.append(
            {
                "Ky": raw["Ky"],
                "Nam": raw["Nam"],
                "Thang": raw["Thang"],
                "MaKhachHangMoi": raw["MaKhachHangMoi"],
                "TenKhachHang": raw["TenKhachHang"],
                "TargetVikoda": raw["TargetVikoda"],
                "TargetTong": raw["TargetTong"],
                "NgayCapNhatNguon": updated,
                "NguonFile": " + ".join(sorted(raw["_sources"])),
                "MienBaoCao": reporting_area,
                "VungBaoCao": reporting_region,
            }
        )
    return records


def main() -> int:
    parser = argparse.ArgumentParser(description="Chuẩn hóa Target tất cả các tháng")
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--staging-dir", required=True)
    args = parser.parse_args()

    source_dir = Path(args.source_dir).resolve()
    staging_dir = Path(args.staging_dir).resolve()
    staging_dir.mkdir(parents=True, exist_ok=True)
    data_file = staging_dir / "target_records.json"
    audit_file = staging_dir / "target_audit.json"

    problems: list[str] = []
    warnings: list[str] = []
    accumulator: dict[tuple[str, str, str], dict[str, Any]] = {}
    source_reports: list[dict[str, Any]] = []

    npp_files = sorted(
        path
        for path in source_dir.glob("Target sellin *.xlsx")
        if not path.name.startswith("~$")
    )
    if not npp_files:
        problems.append("Không tìm thấy file Target sellin YYYY.xlsx")
    years_seen: set[int] = set()
    for path in npp_files:
        try:
            year = parse_year_from_name(path)
            if year in years_seen:
                problems.append(f"Có nhiều file Target sellin cho năm {year}")
                continue
            if year is not None:
                years_seen.add(year)
            source_reports.append(
                {
                    "kind": "NPP",
                    **extract_npp(path, accumulator, problems, warnings),
                }
            )
        except Exception as error:
            problems.append(f"{path.name}: {error}")

    mtka_path = source_dir / "Target MT KA.xlsx"
    expected_periods = {
        f"{year:04d}{month:02d}"
        for year in years_seen
        for month in range(1, 13)
    }
    mtka_period_reports: list[dict[str, Any]] = []
    if mtka_path.exists():
        try:
            mtka_period_reports = extract_mtka_vba_data(
                mtka_path,
                expected_periods,
                accumulator,
                problems,
                warnings,
            )
            source_reports.extend(
                {"kind": "MT_KA_VBA_DATA", **report}
                for report in mtka_period_reports
            )
        except Exception as error:
            problems.append(f"{mtka_path.name}: {error}")
    else:
        problems.append("Không tìm thấy Target MT KA.xlsx")

    derived_records = finalize_accumulator(accumulator)
    actual_periods = {record["Ky"] for record in derived_records}
    missing_periods = sorted(expected_periods - actual_periods)
    extra_periods = sorted(actual_periods - expected_periods)
    if missing_periods:
        problems.append(
            "Thiếu kỳ Target từ hai file năm: " + ", ".join(missing_periods)
        )
    if extra_periods:
        problems.append(
            "Có kỳ Target ngoài năm của file Target sellin: "
            + ", ".join(extra_periods)
        )

    mtka_report_by_period = {
        report["period"]: report
        for report in mtka_period_reports
    }
    for period in sorted(expected_periods):
        report = mtka_report_by_period.get(period)
        if report is None or int(report.get("grouped_customers", 0)) == 0:
            problems.append(
                f"{period}: không có dữ liệu MT/KA trong "
                f"{mtka_path.name} sheet DATA"
            )

    canonical_comparison = None
    canonical_path = source_dir / "Target chuan.xlsm"
    if canonical_path.exists():
        try:
            period, canonical_records, canonical_report = extract_canonical(canonical_path, problems)
            generated_records = [
                record
                for record in derived_records
                if record["Ky"] == period
            ]
            generated_counter = target_record_counter(generated_records)
            canonical_counter = target_record_counter(canonical_records)
            missing_from_generated = sum(
                (canonical_counter - generated_counter).values()
            )
            extra_in_generated = sum(
                (generated_counter - canonical_counter).values()
            )
            generated_summary = summarize_periods(generated_records)
            reference_summary = summarize_periods(canonical_records)
            generated_target_vikoda = sum(
                record["TargetVikoda"]
                for record in generated_records
            )
            generated_target_total = sum(
                record["TargetTong"]
                for record in generated_records
            )
            reference_target_vikoda = sum(
                record["TargetVikoda"]
                for record in canonical_records
            )
            reference_target_total = sum(
                record["TargetTong"]
                for record in canonical_records
            )
            canonical_comparison = {
                "period": period,
                "status": (
                    "MATCHED"
                    if not missing_from_generated and not extra_in_generated
                    else "REFERENCE_ONLY_MISMATCH"
                ),
                "source_policy": TARGET_SOURCE_POLICY,
                "generated_from_two_annual_files": generated_summary,
                "reference_target_chuan": reference_summary,
                "delta_target_vikoda": (
                    generated_target_vikoda - reference_target_vikoda
                ),
                "delta_target_total": (
                    generated_target_total - reference_target_total
                ),
                "missing_from_generated": missing_from_generated,
                "extra_in_generated": extra_in_generated,
            }
            source_reports.append(
                {"kind": "CANONICAL_REFERENCE", **canonical_report}
            )
            if canonical_comparison["status"] != "MATCHED":
                warnings.append(
                    f"{period}: kết quả từ hai file năm khác "
                    f"{canonical_path.name}; file mẫu chỉ dùng tham chiếu, "
                    "không ghi đè dữ liệu"
                )
        except Exception as error:
            problems.append(f"{canonical_path.name}: {error}")
    else:
        warnings.append(
            "Không tìm thấy Target chuan.xlsm; bỏ qua đối chiếu tham chiếu"
        )

    derived_records.sort(
        key=lambda record: (
            record["Nam"],
            record["Thang"],
            record["MaKhachHangMoi"],
            record["TenKhachHang"],
        )
    )

    exact_keys = set()
    code_names: dict[tuple[str, str], set[str]] = defaultdict(set)
    for record in derived_records:
        key = (
            record["Ky"],
            record["MaKhachHangMoi"],
            record["TenKhachHang"],
        )
        if key in exact_keys:
            problems.append(f"Dòng trùng hoàn toàn: {' | '.join(key)}")
        exact_keys.add(key)
        code_names[(record["Ky"], record["MaKhachHangMoi"])].add(record["TenKhachHang"])
        if record["TargetVikoda"] < 0 or record["TargetTong"] < 0:
            problems.append(f"Target âm: {' | '.join(key)}")
        if record["TargetTong"] < record["TargetVikoda"]:
            problems.append(f"TargetTong nhỏ hơn TargetVikoda: {' | '.join(key)}")

    for (period, code), names in sorted(code_names.items()):
        if code == "0":
            warnings.append(f"{period}: mã khách hàng 0 có {len(names)} tên")
        elif len(names) > 1:
            warnings.append(
                f"{period}: mã {code} có nhiều tên: {', '.join(sorted(names))}"
            )

    periods = summarize_periods(derived_records)
    payload = {
        "schema_version": 1,
        "generated_at": datetime.now().astimezone().isoformat(),
        "columns": [
            "Ky",
            "Nam",
            "Thang",
            "MaKhachHangMoi",
            "TenKhachHang",
            "TargetVikoda",
            "TargetTong",
            "NgayCapNhatNguon",
            "NguonFile",
        ],
        "records": derived_records,
    }
    audit = {
        "generated_at": payload["generated_at"],
        "source_dir": str(source_dir),
        "data_file": str(data_file),
        "source_policy": TARGET_SOURCE_POLICY,
        "authoritative_sources": [
            path.name
            for path in npp_files
        ] + ([mtka_path.name] if mtka_path.exists() else []),
        "source_reports": source_reports,
        "period_source_policy": [
            {
                "period": item["period"],
                "policy": TARGET_SOURCE_POLICY,
            }
            for item in periods
        ],
        "canonical_comparison": canonical_comparison,
        "periods": periods,
        "total_records": len(derived_records),
        "warnings": warnings,
        "problems": problems,
    }
    data_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    audit_file.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(
        {
            "data_file": str(data_file),
            "audit_file": str(audit_file),
            "records": len(derived_records),
            "periods": len(periods),
            "source_policy": TARGET_SOURCE_POLICY,
            "warnings": len(warnings),
            "problems": len(problems),
        },
        ensure_ascii=False,
    ))
    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
