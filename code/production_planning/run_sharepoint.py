"""Run production planning against exact SharePoint workbooks via Microsoft Graph.

GitHub Actions authenticates with OIDC before this script starts. No client secret is
stored here. The runner downloads the live source and target workbooks to the
ephemeral workspace, recalculates T:AZ in Python, validates totals, and uploads the
target only when the logical workbook payload changed.

The target write uses If-Match with the ETag observed before download so a concurrent
human edit cannot be silently overwritten.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SYNC_DIR = PROJECT_ROOT / "code/Skill/skill-bao-cao/scripts"
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
if str(SYNC_DIR) not in sys.path:
    sys.path.insert(0, str(SYNC_DIR))

from code.production_planning.engine import calculate_daily_plans  # noqa: E402
from code.production_planning.workbook_io import (  # noqa: E402
    TARGET_SHEET,
    read_source_records,
    validate_written_plan,
    write_target_plan,
)
from sharepoint_graph_sync import get_graph_access_token, http_request_with_retry  # noqa: E402


DEFAULT_SOURCE_ITEM_ID = "01EX3DDZDCP23YZACRFBBZXS3XIFZXDAMW"
DEFAULT_TARGET_ITEM_ID = "01EX3DDZGDKHPGY5HLMJAJKFWC3IVMHDEL"
DEFAULT_SOURCE_NAME = "File tính kế hoạch - BẢN CẢI TIẾN_V2.xlsm"
DEFAULT_TARGET_NAME = "Sắp kế hoạch.xlsx"
DEFAULT_WORK_DIR = PROJECT_ROOT / "Data/Work/production_planning"


class SharePointPlanningError(RuntimeError):
    """Raised when the exact-file SharePoint planning transaction cannot be proven safe."""


def _graph_url(drive_id: str, item_id: str, suffix: str = "") -> str:
    drive = urllib.parse.quote(drive_id, safe="!_-")
    item = urllib.parse.quote(item_id, safe="!_-")
    return f"https://graph.microsoft.com/v1.0/drives/{drive}/items/{item}{suffix}"


def _authorized_request(token: str, url: str, *, method: str = "GET", data: bytes | None = None):
    request = urllib.request.Request(url, data=data, method=method)
    request.add_header("Authorization", f"Bearer {token}")
    return request


def get_item_metadata(token: str, drive_id: str, item_id: str) -> dict[str, Any]:
    request = _authorized_request(token, _graph_url(drive_id, item_id))
    body, _ = http_request_with_retry(request)
    try:
        payload = json.loads(body.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise SharePointPlanningError("Graph item metadata is not valid JSON") from exc
    if not isinstance(payload, dict):
        raise SharePointPlanningError("Graph item metadata must be a JSON object")
    return payload


def download_item(token: str, drive_id: str, item_id: str, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    request = _authorized_request(token, _graph_url(drive_id, item_id, "/content"))
    body, status = http_request_with_retry(request)
    if status != 200 or not body:
        raise SharePointPlanningError(
            f"SharePoint download failed for {item_id}: HTTP {status}, bytes={len(body)}"
        )
    destination.write_bytes(body)
    if destination.stat().st_size <= 0:
        raise SharePointPlanningError(f"Downloaded file is empty: {destination.name}")


def upload_item_if_match(
    token: str,
    drive_id: str,
    item_id: str,
    source: Path,
    etag: str,
) -> dict[str, Any]:
    if not source.is_file() or source.stat().st_size <= 0:
        raise SharePointPlanningError(f"Upload file missing or empty: {source}")
    data = source.read_bytes()
    request = _authorized_request(
        token,
        _graph_url(drive_id, item_id, "/content"),
        method="PUT",
        data=data,
    )
    request.add_header("Content-Type", "application/octet-stream")
    if etag:
        request.add_header("If-Match", etag)
    try:
        body, status = http_request_with_retry(request)
    except urllib.error.HTTPError as exc:
        if exc.code == 412:
            raise SharePointPlanningError(
                "Target workbook changed after download; refusing to overwrite concurrent edits"
            ) from exc
        raise
    if status not in {200, 201}:
        raise SharePointPlanningError(f"SharePoint upload failed: HTTP {status}")
    if not body:
        return {}
    try:
        payload = json.loads(body.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _assert_exact_item(metadata: Mapping[str, Any], expected_name: str, item_id: str) -> str:
    actual_id = str(metadata.get("id") or "").strip()
    actual_name = str(metadata.get("name") or "").strip()
    if actual_id and actual_id != item_id:
        raise SharePointPlanningError(
            f"Graph returned unexpected item id: expected={item_id}, actual={actual_id}"
        )
    if actual_name != expected_name:
        raise SharePointPlanningError(
            f"Wrong planning workbook: expected={expected_name!r}, actual={actual_name!r}"
        )
    return str(metadata.get("eTag") or metadata.get("@odata.etag") or "").strip()


def _normalize_cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if abs(value) < 1e-9:
            return 0.0
        return round(value, 9)
    return value


def _sheet_signature(path: Path, sheet_name: str = TARGET_SHEET) -> tuple[tuple[object, ...], ...]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SharePointPlanningError(f"Target sheet {sheet_name!r} not found in {path.name}")
        sheet = workbook[sheet_name]
        max_row = max(1, sheet.max_row)
        return tuple(
            tuple(_normalize_cell(sheet.cell(row, col).value) for col in range(1, 53))
            for row in range(1, max_row + 1)
        )
    finally:
        workbook.close()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _append_summary(lines: list[str]) -> None:
    summary_path = str(os.environ.get("GITHUB_STEP_SUMMARY", "")).strip()
    if not summary_path:
        return
    with Path(summary_path).open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--drive-id",
        default=os.environ.get("SHAREPOINT_DRIVE_ID", ""),
        help="SharePoint document drive id; defaults to SHAREPOINT_DRIVE_ID",
    )
    parser.add_argument(
        "--source-item-id",
        default=os.environ.get("PLANNING_SOURCE_ITEM_ID", DEFAULT_SOURCE_ITEM_ID),
    )
    parser.add_argument(
        "--target-item-id",
        default=os.environ.get("PLANNING_TARGET_ITEM_ID", DEFAULT_TARGET_ITEM_ID),
    )
    parser.add_argument(
        "--source-name",
        default=os.environ.get("PLANNING_SOURCE_NAME", DEFAULT_SOURCE_NAME),
    )
    parser.add_argument(
        "--target-name",
        default=os.environ.get("PLANNING_TARGET_NAME", DEFAULT_TARGET_NAME),
    )
    parser.add_argument("--plan-year", type=int, default=None)
    parser.add_argument("--work-dir", type=Path, default=DEFAULT_WORK_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    drive_id = str(args.drive_id or "").strip()
    source_item_id = str(args.source_item_id or "").strip()
    target_item_id = str(args.target_item_id or "").strip()
    if not drive_id:
        print("::error::SHAREPOINT_DRIVE_ID is missing; run sharepoint_bootstrap.py first")
        return 2
    if not source_item_id or not target_item_id:
        print("::error::Planning source/target item id is missing")
        return 2
    if source_item_id == target_item_id:
        print("::error::Planning source and target item IDs must be different")
        return 2

    work_dir = args.work_dir.resolve()
    source_path = work_dir / args.source_name
    target_path = work_dir / args.target_name
    output_path = work_dir / f"updated-{args.target_name}"

    try:
        token = get_graph_access_token()
        source_meta = get_item_metadata(token, drive_id, source_item_id)
        target_meta = get_item_metadata(token, drive_id, target_item_id)
        _assert_exact_item(source_meta, args.source_name, source_item_id)
        target_etag = _assert_exact_item(target_meta, args.target_name, target_item_id)
        download_item(token, drive_id, source_item_id, source_path)
        download_item(token, drive_id, target_item_id, target_path)

        records, plan_period = read_source_records(source_path, plan_year=args.plan_year)
        plans = calculate_daily_plans(record.as_plan_row() for record in records)
        validate_written_plan(records, plans)
        before_signature = _sheet_signature(target_path)
        write_target_plan(
            target_path,
            output_path,
            records=records,
            daily_plans=plans,
            plan_period=plan_period,
        )
        after_signature = _sheet_signature(output_path)

        changed = before_signature != after_signature
        upload_result: dict[str, Any] = {}
        if changed:
            upload_result = upload_item_if_match(
                token,
                drive_id,
                target_item_id,
                output_path,
                target_etag,
            )

        plan_total = sum(sum(plans[record.source_row]) for record in records)
        status = "UPDATED" if changed else "UNCHANGED"
        print(
            f"Production planning: {status} | period={plan_period} | "
            f"products={len(records)} | total={plan_total:.2f}"
        )
        _append_summary(
            [
                "## Production planning",
                "",
                f"- Status: **{status}**",
                f"- Period: `{plan_period}`",
                f"- Products: `{len(records)}`",
                f"- Planned output: `{plan_total:.2f}`",
                f"- Source: `{args.source_name}`",
                f"- Target: `{args.target_name}` / `{TARGET_SHEET}`",
                f"- Source SHA256: `{_sha256(source_path)[:16]}…`",
                f"- Target Graph id: `{upload_result.get('id', target_item_id)}`",
            ]
        )
        return 0
    except Exception as exc:
        print(f"::error::Production planning failed: {type(exc).__name__}: {exc}")
        _append_summary(
            [
                "## Production planning",
                "",
                "- Status: **FAILED**",
                f"- Error: `{type(exc).__name__}: {exc}`",
            ]
        )
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
