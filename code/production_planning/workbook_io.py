"""Workbook adapters for the Vikoda production planning engine.

The source workbook remains the operational input for columns A:S. Python replaces
only the daily allocation logic that Excel currently performs in T:AZ, then writes
the normalized result to ``Sắp kế hoạch.xlsx`` / ``Ke_hoach_SX``.
"""

from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .engine import DAYS, PlanRow, coerce_date, normalized_difference


SOURCE_SHEET = "Ke hoach SX tuan"
TARGET_SHEET = "Ke_hoach_SX"
SOURCE_FIRST_DATA_ROW = 4
TARGET_FIRST_DATA_ROW = 2


@dataclass(frozen=True)
class SourceRecord:
    source_row: int
    code: int | str
    name: str
    unit: str
    batch_qty: float
    shift_qty: float
    line: str
    group: str
    product_class: str
    setup_key: object
    shifts_per_day: float
    stock_actual: float
    stock_book: float
    fc: float
    projected_end_stock: float
    warehouse_debt: float
    required_qty: float
    planned_qty: float
    production_days: float
    start_date: date | None

    def as_plan_row(self) -> PlanRow:
        return PlanRow(
            source_row=self.source_row,
            code=self.code,
            line=self.line,
            group=self.group,
            batch_qty=self.batch_qty,
            shift_qty=self.shift_qty,
            shifts_per_day=self.shifts_per_day,
            planned_qty=self.planned_qty,
            start_date=self.start_date,
            setup_key=self.setup_key,
        )


class WorkbookContractError(RuntimeError):
    """Raised when an input workbook does not match the expected planning contract."""


def _num(value: object) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _normalize_code(value: object) -> int | str:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    return text


def _parse_plan_month(label: object) -> int:
    match = re.search(r"(\d{1,2})", str(label or ""))
    if not match:
        raise WorkbookContractError(f"Cannot parse planning month from {label!r}")
    month = int(match.group(1))
    if not 1 <= month <= 12:
        raise WorkbookContractError(f"Planning month out of range: {month}")
    return month


def _infer_plan_year(records: Iterable[SourceRecord], override: int | None) -> int:
    if override is not None:
        if override < 2000 or override > 2100:
            raise WorkbookContractError(f"Invalid planning year override: {override}")
        return override
    counts: dict[int, int] = {}
    for record in records:
        if record.start_date is not None:
            counts[record.start_date.year] = counts.get(record.start_date.year, 0) + 1
    if not counts:
        return datetime.now().year
    return max(counts, key=lambda year: (counts[year], year))


def read_source_records(
    workbook_path: str | Path,
    *,
    sheet_name: str = SOURCE_SHEET,
    plan_year: int | None = None,
) -> tuple[list[SourceRecord], str]:
    """Read cached A:S values from the planning source workbook.

    ``data_only=True`` is intentional: upstream A:S is still owned by the source
    workbook. The Python engine replaces the daily T:AZ allocation only.
    """
    path = Path(workbook_path)
    if not path.is_file() or path.stat().st_size <= 0:
        raise WorkbookContractError(f"Source workbook missing or empty: {path}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookContractError(
                f"Source sheet {sheet_name!r} not found. Available: {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        plan_month = _parse_plan_month(sheet["B2"].value)
        records: list[SourceRecord] = []
        blank_streak = 0
        for row_index in range(SOURCE_FIRST_DATA_ROW, sheet.max_row + 1):
            code_value = sheet.cell(row_index, 1).value
            if code_value in (None, ""):
                blank_streak += 1
                if records and blank_streak >= 5:
                    break
                continue
            blank_streak = 0
            code = _normalize_code(code_value)
            name = str(sheet.cell(row_index, 2).value or "").strip()
            line = str(sheet.cell(row_index, 6).value or "").strip()
            if not name or not line:
                continue
            records.append(
                SourceRecord(
                    source_row=row_index,
                    code=code,
                    name=name,
                    unit=str(sheet.cell(row_index, 3).value or "").strip(),
                    batch_qty=_num(sheet.cell(row_index, 4).value),
                    shift_qty=_num(sheet.cell(row_index, 5).value),
                    line=line,
                    group=str(sheet.cell(row_index, 7).value or "").strip(),
                    product_class=str(sheet.cell(row_index, 8).value or "").strip(),
                    setup_key=sheet.cell(row_index, 9).value,
                    shifts_per_day=_num(sheet.cell(row_index, 10).value),
                    stock_actual=_num(sheet.cell(row_index, 11).value),
                    stock_book=_num(sheet.cell(row_index, 12).value),
                    fc=_num(sheet.cell(row_index, 13).value),
                    projected_end_stock=_num(sheet.cell(row_index, 14).value),
                    warehouse_debt=_num(sheet.cell(row_index, 15).value),
                    required_qty=_num(sheet.cell(row_index, 16).value),
                    planned_qty=_num(sheet.cell(row_index, 17).value),
                    production_days=_num(sheet.cell(row_index, 18).value),
                    start_date=coerce_date(sheet.cell(row_index, 19).value),
                )
            )
        if not records:
            raise WorkbookContractError(f"No production rows found in {sheet_name!r}")
        seen: set[int | str] = set()
        duplicates: set[str] = set()
        for record in records:
            if record.code in seen:
                duplicates.add(str(record.code))
            seen.add(record.code)
        if duplicates:
            raise WorkbookContractError(
                "Duplicate product codes in source plan: " + ", ".join(sorted(duplicates))
            )
        year = _infer_plan_year(records, plan_year)
        return records, f"{year:04d}-{plan_month:02d}"
    finally:
        workbook.close()


def _clone_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _prepare_target_styles(sheet: Worksheet, last_row: int) -> None:
    """Extend the existing target visual style without redesigning the workbook."""
    header_template = sheet["L1"]
    data_template = sheet["L2"] if sheet.max_row >= 2 else None
    for column in range(13, 53):
        _clone_style(header_template, sheet.cell(1, column))
        if data_template is not None:
            for row in range(TARGET_FIRST_DATA_ROW, last_row + 1):
                _clone_style(data_template, sheet.cell(row, column))
    for row in range(TARGET_FIRST_DATA_ROW, last_row + 1):
        sheet.cell(row, 18).number_format = "dd/mm/yyyy"
        for column in range(20, 53):
            sheet.cell(row, column).number_format = "#,##0.##"


def _headers() -> dict[int, str]:
    result = {
        1: "Mã Sản Phẩm",
        2: "Tên Sản Phẩm",
        3: "Đơn vị tính",
        4: "Số lượng /mẻ",
        5: "Số lượng/ ca",
        6: "Chuyền",
        7: "Nhóm sản phẩm",
        8: "Phân loại SP",
        9: "Số ca theo ngày",
        10: "Tồn đầu thực tế",
        11: "Tồn đầu sổ sách",
        12: "FC",
        13: "Tồn cuối dự kiến",
        14: "Nợ kho",
        15: "Số lượng cần sản xuất",
        16: "Số lượng sản xuất theo mẻ/ca",
        17: "Số ngày cần sản xuất",
        18: "Ngày bắt đầu sản xuất",
        19: "Kỳ kế hoạch",
        51: "Tổng SX",
        52: "Chênh lệch (SX-P)",
    }
    for offset in range(DAYS):
        result[20 + offset] = f"Ngày {offset + 1:02d}"
    return result


def write_target_plan(
    target_workbook: str | Path,
    output_workbook: str | Path,
    *,
    records: list[SourceRecord],
    daily_plans: dict[int, list[float]],
    plan_period: str,
    sheet_name: str = TARGET_SHEET,
) -> None:
    """Write A:AZ values to ``Ke_hoach_SX`` while preserving the workbook."""
    source = Path(target_workbook)
    output = Path(output_workbook)
    if not source.is_file() or source.stat().st_size <= 0:
        raise WorkbookContractError(f"Target workbook missing or empty: {source}")
    workbook = load_workbook(source)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookContractError(
                f"Target sheet {sheet_name!r} not found. Available: {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        last_row = TARGET_FIRST_DATA_ROW + len(records) - 1
        _prepare_target_styles(sheet, max(last_row, sheet.max_row))
        for column, title in _headers().items():
            sheet.cell(1, column).value = title
        clear_last_row = max(sheet.max_row, last_row)
        for row in range(TARGET_FIRST_DATA_ROW, clear_last_row + 1):
            for column in range(1, 53):
                sheet.cell(row, column).value = None
        for target_row, record in enumerate(records, start=TARGET_FIRST_DATA_ROW):
            plan = daily_plans.get(record.source_row)
            if plan is None or len(plan) != DAYS:
                raise WorkbookContractError(f"Missing 31-day plan for product {record.code}")
            values = {
                1: record.code,
                2: record.name,
                3: record.unit,
                4: record.batch_qty,
                5: record.shift_qty,
                6: record.line,
                7: record.group,
                8: record.product_class,
                9: record.shifts_per_day,
                10: record.stock_actual,
                11: record.stock_book,
                12: record.fc,
                13: record.projected_end_stock,
                14: record.warehouse_debt,
                15: record.required_qty,
                16: record.planned_qty,
                17: record.production_days,
                18: record.start_date,
                19: plan_period,
            }
            for column, value in values.items():
                sheet.cell(target_row, column).value = value
            for index, quantity in enumerate(plan):
                sheet.cell(target_row, 20 + index).value = quantity if abs(quantity) > 1e-9 else None
            total = sum(plan)
            if abs(total) < 1e-9:
                total = 0.0
            sheet.cell(target_row, 51).value = total
            sheet.cell(target_row, 52).value = normalized_difference(total, record.planned_qty)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:AZ{last_row}"
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()


def validate_written_plan(records: list[SourceRecord], daily_plans: dict[int, list[float]]) -> None:
    """Fail closed when a positive planned quantity is not fully allocated."""
    failures: list[str] = []
    for record in records:
        plan = daily_plans.get(record.source_row, [])
        if len(plan) != DAYS:
            failures.append(f"{record.code}: expected 31 daily cells, got {len(plan)}")
            continue
        if any(value < -1e-9 for value in plan):
            failures.append(f"{record.code}: negative daily production")
        total = sum(plan)
        if record.planned_qty > 0 and abs(total - record.planned_qty) > 1e-6:
            failures.append(f"{record.code}: daily total {total} != planned {record.planned_qty}")
        if record.planned_qty <= 0 and total > 1e-6:
            failures.append(f"{record.code}: non-positive plan produced {total}")
    if failures:
        raise WorkbookContractError("; ".join(failures))
