"""Workbook contract tests for production planning."""

from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from code.production_planning.engine import calculate_daily_plans
from code.production_planning.workbook_io import (
    read_source_records,
    validate_written_plan,
    write_target_plan,
)


class ProductionPlanningWorkbookTests(unittest.TestCase):
    def _build_source(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ke hoach SX tuan"
        ws["B2"] = "Tháng 1"
        headers = [
            "Mã SP", "Tên SP", "ĐVT", "SL/mẻ", "SL/ca", "Chuyền", "Nhóm",
            "Phân loại", "Setup", "Ca/ngày", "Tồn TT", "Tồn sổ", "FC",
            "Tồn cuối", "Nợ kho", "Cần SX", "Kế hoạch SX", "Ngày SX", "Bắt đầu",
        ]
        for column, value in enumerate(headers, 1):
            ws.cell(3, column).value = value
        values = [
            9001, "Synthetic SKU", "Thùng", 100, 100, "Galon", "Galon",
            "Không đường", 1, 2, 50, 50, 250, 0, 0, 200, 200, 1, date(2026, 1, 3),
        ]
        for column, value in enumerate(values, 1):
            ws.cell(4, column).value = value
        wb.save(path)

    def _build_target(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ke_hoach_SX"
        for column in range(1, 13):
            ws.cell(1, column).value = f"H{column}"
        ws.cell(2, 1).value = 9001
        wb.save(path)

    def test_read_calculate_write_contract(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            target = root / "target.xlsx"
            output = root / "output.xlsx"
            self._build_source(source)
            self._build_target(target)
            records, period = read_source_records(source, plan_year=2026)
            plans = calculate_daily_plans([record.as_plan_row() for record in records])
            validate_written_plan(records, plans)
            write_target_plan(
                target,
                output,
                records=records,
                daily_plans=plans,
                plan_period=period,
            )
            wb = load_workbook(output, data_only=True)
            try:
                ws = wb["Ke_hoach_SX"]
                self.assertEqual(ws["A2"].value, 9001)
                self.assertEqual(ws["I2"].value, 2)
                self.assertEqual(ws["P2"].value, 200)
                self.assertEqual(ws["R2"].value.date(), date(2026, 1, 3))
                self.assertEqual(ws["S2"].value, "2026-01")
                self.assertEqual(ws["V2"].value, 200)
                self.assertEqual(ws["AY2"].value, 200)
                self.assertEqual(ws["AZ2"].value, 0)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
