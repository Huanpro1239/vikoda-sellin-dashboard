from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from normalization import quantity_number_format


OUTPUT_COLUMNS = [
    "Vung",
    "KhuVuc",
    "NgayHoaDon",
    "MaKhachHangMoi",
    "TenKhachHang",
    "MaSanPhamMoi",
    "TenSanPham",
    "SoLuong",
    "DonGia",
    "ThanhTien",
    "LoaiDonHang",
    "GhiChu",
    "Thang",
    "Nam",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit-file", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--report-file", required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    audit = json.loads(Path(args.audit_file).read_text(encoding="utf-8"))
    output_dir = Path(args.output_dir).resolve()
    report_file = Path(args.report_file).resolve()

    expected = {
        (item["year"], item["month"]): item["output_rows"]
        for item in audit["monthly_files"]
    }
    reports = []
    problems = []

    for (year, month), expected_rows in sorted(expected.items()):
        output_path = output_dir / f"Sell in T{month:02d}_{year}.xlsx"
        if not output_path.exists():
            continue
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        try:
            ws = workbook.worksheets[0]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            row_count = 0
            invalid_dates = 0
            invalid_product_prefix = 0
            nontext_product_codes = 0
            invalid_period = 0
            formula_cells = 0
            invalid_quantity_type = 0
            invalid_quantity_number_format = 0
            trailing_quantity_text = 0
            sum_quantity = 0.0
            sum_revenue = 0.0

            for values in ws.iter_rows(min_row=2, values_only=False):
                if not any(cell.value not in (None, "") for cell in values):
                    continue
                row_count += 1
                raw_values = [cell.value for cell in values]
                invoice_date = raw_values[2]
                product_code = raw_values[5]
                if not isinstance(invoice_date, (date, datetime)):
                    invalid_dates += 1
                if not isinstance(product_code, str):
                    nontext_product_codes += 1
                if not str(product_code or "").startswith(("1", "2")):
                    invalid_product_prefix += 1
                if raw_values[12] != month or raw_values[13] != year:
                    invalid_period += 1
                quantity = raw_values[7]
                if quantity is not None and (
                    isinstance(quantity, bool)
                    or not isinstance(quantity, (int, float))
                ):
                    invalid_quantity_type += 1
                if isinstance(quantity, str) and quantity.rstrip().endswith("."):
                    trailing_quantity_text += 1
                if values[7].number_format != quantity_number_format(quantity):
                    invalid_quantity_number_format += 1
                sum_quantity += float(raw_values[7] or 0)
                sum_revenue += float(raw_values[9] or 0)
                formula_cells += sum(
                    1 for cell in values if cell.data_type == "f"
                )

            report = {
                "file": str(output_path),
                "sheet_count": len(workbook.worksheets),
                "sheet_name": ws.title,
                "headers_ok": headers == OUTPUT_COLUMNS,
                "row_count": row_count,
                "expected_rows": expected_rows,
                "invalid_dates": invalid_dates,
                "invalid_product_prefix": invalid_product_prefix,
                "nontext_product_codes": nontext_product_codes,
                "invalid_period": invalid_period,
                "formula_cells": formula_cells,
                "invalid_quantity_type": invalid_quantity_type,
                "invalid_quantity_number_format": invalid_quantity_number_format,
                "trailing_quantity_text": trailing_quantity_text,
                "sum_quantity": sum_quantity,
                "sum_revenue": sum_revenue,
            }
            reports.append(report)
            if (
                report["sheet_count"] != 1
                or report["sheet_name"] != "Sell in"
                or not report["headers_ok"]
                or report["row_count"] != report["expected_rows"]
                or invalid_dates
                or invalid_product_prefix
                or nontext_product_codes
                or invalid_period
                or formula_cells
                or invalid_quantity_type
                or invalid_quantity_number_format
                or trailing_quantity_text
            ):
                problems.append(report)
        finally:
            workbook.close()

    missing = [
        {"year": year, "month": month}
        for year, month in expected
        if not (output_dir / f"Sell in T{month:02d}_{year}.xlsx").exists()
    ]
    payload = {
        "reports": reports,
        "missing": missing,
        "problems": problems,
    }
    report_file.parent.mkdir(parents=True, exist_ok=True)
    report_file.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    if missing or problems or len(reports) != len(expected):
        raise SystemExit(1)
    print(str(report_file))


if __name__ == "__main__":
    main()
