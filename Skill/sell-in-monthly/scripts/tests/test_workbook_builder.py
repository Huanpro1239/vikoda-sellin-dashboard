"""Khóa định dạng workbook Sell In.

Luồng chính (`build_outputs.py`) và luồng chuyển giao (`portable_sell_in.py`)
cùng gọi `workbook_builder`, nên test ở đây bảo đảm file ra giống nhau và giữ
đúng định dạng của các workbook đang có trong `Data/out put`.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPT_DIR))

from extraction import OUTPUT_COLUMNS  # noqa: E402
from workbook_builder import (  # noqa: E402
    build_workbook,
    verify_workbook,
    write_monthly_workbook,
)


def make_rows(count: int = 3, quantity: float = 240) -> list[dict]:
    return [
        {
            "Vung": "MT",
            "KhuVuc": "MT",
            "NgayHoaDon": date(2026, 7, 10 + index),
            "MaKhachHangMoi": f"KH-{index:03d}",
            "TenKhachHang": "KHACH HANG",
            "MaSanPhamMoi": "130100001",
            "TenSanPham": "SAN PHAM",
            "SoLuong": quantity,
            "DonGia": 100,
            "ThanhTien": 1000,
            "LoaiDonHang": "Don hang ban",
            "GhiChu": "",
            "Thang": 7,
            "Nam": 2026,
        }
        for index in range(count)
    ]


def saved(rows: list[dict], month: int = 7, year: int = 2026):
    folder = tempfile.TemporaryDirectory()
    path = Path(folder.name) / "out.xlsx"
    workbook = build_workbook(rows, month, year)
    workbook.save(path)
    workbook.close()
    return folder, load_workbook(path)


class FormatTest(unittest.TestCase):
    def setUp(self) -> None:
        self.folder, self.wb = saved(make_rows())
        self.ws = self.wb.worksheets[0]

    def tearDown(self) -> None:
        self.wb.close()
        self.folder.cleanup()

    def test_mot_sheet_ten_sell_in(self) -> None:
        self.assertEqual(len(self.wb.worksheets), 1)
        self.assertEqual(self.ws.title, "Sell in")

    def test_dung_14_cot_dung_thu_tu(self) -> None:
        headers = [cell.value for cell in self.ws[1]]
        self.assertEqual(headers, OUTPUT_COLUMNS)
        self.assertEqual(len(headers), 14)

    def test_khoa_dong_tieu_de(self) -> None:
        """Luồng Node cũ đặt freeze rồi bị engine bỏ mất; openpyxl giữ được."""
        self.assertEqual(self.ws.freeze_panes, "A2")

    def test_an_duong_luoi(self) -> None:
        self.assertFalse(self.ws.sheet_view.showGridLines)

    def test_dinh_dang_tieu_de(self) -> None:
        cell = self.ws.cell(1, 1)
        self.assertEqual(cell.fill.fgColor.rgb, "FF0F766E")
        self.assertEqual(cell.font.color.rgb, "FFFFFFFF")
        self.assertTrue(cell.font.b)
        self.assertEqual(cell.font.name, "Aptos")
        self.assertEqual(self.ws.row_dimensions[1].height, 30)

    def test_dinh_dang_so_tung_cot(self) -> None:
        self.assertEqual(self.ws.cell(2, 3).number_format, "dd/mm/yyyy")
        self.assertEqual(self.ws.cell(2, 4).number_format, "@")
        self.assertEqual(self.ws.cell(2, 6).number_format, "@")
        self.assertEqual(self.ws.cell(2, 9).number_format, "#,##0")
        self.assertEqual(self.ws.cell(2, 10).number_format, "#,##0")

    def test_ngay_hoa_don_la_ngay_excel_that(self) -> None:
        self.assertIsInstance(self.ws.cell(2, 3).value, date)

    def test_ma_khach_hang_va_san_pham_la_chu(self) -> None:
        self.assertIsInstance(self.ws.cell(2, 4).value, str)
        self.assertIsInstance(self.ws.cell(2, 6).value, str)

    def test_bang_va_do_rong_cot(self) -> None:
        table = list(self.ws.tables.values())[0]
        self.assertEqual(table.displayName, "SellIn_2026_07")
        self.assertEqual(table.ref, "A1:N4")
        self.assertEqual(table.tableStyleInfo.name, "TableStyleMedium2")
        self.assertEqual(self.ws.column_dimensions["A"].width, 12)
        self.assertEqual(self.ws.column_dimensions["N"].width, 9)


class QuantityFormatTest(unittest.TestCase):
    def test_so_nguyen_khong_co_phan_thap_phan(self) -> None:
        folder, wb = saved(make_rows(1, quantity=240))
        ws = wb.worksheets[0]
        self.assertEqual(ws.cell(2, 8).value, 240)
        self.assertEqual(ws.cell(2, 8).number_format, "#,##0")
        wb.close(); folder.cleanup()

    def test_so_le_that_duoc_giu(self) -> None:
        folder, wb = saved(make_rows(1, quantity=1.92))
        ws = wb.worksheets[0]
        self.assertEqual(ws.cell(2, 8).value, 1.92)
        self.assertEqual(ws.cell(2, 8).number_format, "#,##0.##")
        wb.close(); folder.cleanup()


class EmptyMonthTest(unittest.TestCase):
    def test_thang_khong_co_dong_dung_auto_filter(self) -> None:
        folder, wb = saved([])
        ws = wb.worksheets[0]
        self.assertEqual(len(ws.tables), 0)
        self.assertEqual(ws.auto_filter.ref, "A1:N1")
        wb.close(); folder.cleanup()


class WriteAndVerifyTest(unittest.TestCase):
    def test_ghi_roi_kiem_tra_dat(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            rows = make_rows(5)
            report = write_monthly_workbook(Path(folder), rows, 7, 2026)
            self.assertTrue(report["ok"])
            self.assertEqual(report["row_count"], 5)
            self.assertEqual(report["expected_rows"], 5)
            self.assertEqual(report["formula_cells"], 0)
            self.assertTrue(report["file"].endswith("Sell in T07_2026.xlsx"))
            self.assertTrue(Path(report["file"]).is_file())

    def test_khong_de_lai_file_tam(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            write_monthly_workbook(Path(folder), make_rows(2), 7, 2026)
            leftovers = [p.name for p in Path(folder).iterdir()
                         if p.name.startswith(".")]
            self.assertEqual(leftovers, [])

    def test_verify_bat_duoc_sai_ky(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "sai.xlsx"
            workbook = build_workbook(make_rows(2), 7, 2026)
            workbook.save(path)
            workbook.close()
            # Kỳ trong file là 7/2026, kiểm tra theo 8/2026 phải trượt.
            report = verify_workbook(path, 8, 2026, 2)
            self.assertFalse(report["ok"])
            self.assertEqual(report["invalid_period"], 2)

    def test_verify_bat_duoc_lech_so_dong(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "lech.xlsx"
            workbook = build_workbook(make_rows(3), 7, 2026)
            workbook.save(path)
            workbook.close()
            report = verify_workbook(path, 7, 2026, 99)
            self.assertFalse(report["ok"])
            self.assertEqual(report["row_count"], 3)


if __name__ == "__main__":
    unittest.main()
