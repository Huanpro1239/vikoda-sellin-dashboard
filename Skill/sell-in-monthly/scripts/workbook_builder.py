"""Dựng workbook Sell In theo tháng bằng openpyxl.

Module dùng chung cho luồng chính (`build_outputs.py`) và luồng chuyển giao
(`portable_sell_in.py`). Trước đây luồng chính dựng file bằng Node
(`build_outputs.mjs` + `@oai/artifact-tool`) còn luồng chuyển giao dựng bằng
openpyxl, nên cùng một tháng ra hai kiểu định dạng khác nhau. Giữ đúng một bản
ở đây để mọi file đầu ra giống nhau bất kể chạy luồng nào.

Định dạng lấy theo 18/19 file đang có trong `Data/out put`, tức kiểu openpyxl:
tiêu đề Aptos 11 đậm nền xanh, thân bảng dùng font mặc định, khóa dòng tiêu đề.
"""

from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from extraction import COLUMN_WIDTHS, OUTPUT_COLUMNS, row_to_values
from normalization import quantity_number_format


SHEET_TITLE = "Sell in"

# Dùng ARGB đủ 8 ký tự để Excel không phải tự suy diễn kênh alpha.
HEADER_FILL_COLOR = "FF0F766E"
HEADER_FONT_COLOR = "FFFFFFFF"
HEADER_BORDER_COLOR = "FF0B5D56"
HEADER_FONT_NAME = "Aptos"
HEADER_FONT_SIZE = 11
HEADER_ROW_HEIGHT = 30

# Thân bảng để font mặc định của Excel. Đổi thành ("Aptos", 10) nếu muốn giống
# kiểu mà luồng Node cũ tạo ra; khi đó sửa cả tài liệu định dạng.
BODY_FONT_NAME: str | None = None
BODY_FONT_SIZE: int | None = None

TABLE_STYLE = "TableStyleMedium2"

# Chỉ số cột (1-based) trong OUTPUT_COLUMNS.
COL_NGAY_HOA_DON = 3
COL_MA_KHACH_HANG = 4
COL_MA_SAN_PHAM = 6
COL_SO_LUONG = 8
COL_DON_GIA = 9
COL_THANH_TIEN = 10
COL_THANG = 13
COL_NAM = 14


def table_name(month: int, year: int) -> str:
    return f"SellIn_{year}_{month:02d}"


def build_workbook(rows: list[dict], month: int, year: int) -> Workbook:
    """Tạo workbook một sheet `Sell in` đã định dạng đầy đủ."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = SHEET_TITLE
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.append(OUTPUT_COLUMNS)

    for row in rows:
        ws.append(row_to_values(row))

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(
        name=HEADER_FONT_NAME,
        size=HEADER_FONT_SIZE,
        bold=True,
        color=HEADER_FONT_COLOR,
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    header_border = Border(
        bottom=Side(style="medium", color=HEADER_BORDER_COLOR)
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    body_font = (
        Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE)
        if BODY_FONT_NAME
        else None
    )
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    last_row = len(rows) + 1
    for row_number in range(2, last_row + 1):
        ws.cell(row_number, COL_NGAY_HOA_DON).number_format = "dd/mm/yyyy"
        ws.cell(row_number, COL_MA_KHACH_HANG).number_format = "@"
        ws.cell(row_number, COL_MA_SAN_PHAM).number_format = "@"
        ws.cell(row_number, COL_SO_LUONG).number_format = quantity_number_format(
            ws.cell(row_number, COL_SO_LUONG).value
        )
        ws.cell(row_number, COL_DON_GIA).number_format = "#,##0"
        ws.cell(row_number, COL_THANH_TIEN).number_format = "#,##0"
        ws.cell(row_number, COL_THANG).number_format = "0"
        ws.cell(row_number, COL_NAM).number_format = "0"
        for column_number in (COL_SO_LUONG, COL_DON_GIA, COL_THANH_TIEN):
            ws.cell(row_number, column_number).alignment = right
        for column_number in (COL_THANG, COL_NAM):
            ws.cell(row_number, column_number).alignment = center
        if body_font is not None:
            for column_number in range(1, len(OUTPUT_COLUMNS) + 1):
                ws.cell(row_number, column_number).font = body_font

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    table_ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{last_row}"
    if rows:
        table = Table(displayName=table_name(month, year), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = table_ref

    return workbook


def verify_workbook(
    workbook_path: Path,
    month: int,
    year: int,
    expected_rows: int,
) -> dict:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
    )
    try:
        ws = workbook.worksheets[0]
        headers = [
            cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        row_count = 0
        invalid_dates = 0
        invalid_product_prefix = 0
        nontext_product_codes = 0
        invalid_period = 0
        formula_cells = 0
        invalid_quantity_type = 0
        invalid_quantity_number_format = 0
        trailing_quantity_text = 0
        sum_quantity = 0.0
        sum_revenue = 0.0

        for cells in ws.iter_rows(min_row=2, values_only=False):
            if not any(cell.value not in (None, "") for cell in cells):
                continue
            values = [cell.value for cell in cells]
            row_count += 1
            if not isinstance(values[2], (date, datetime)):
                invalid_dates += 1
            if not isinstance(values[5], str):
                nontext_product_codes += 1
            if not str(values[5] or "").startswith(("1", "2")):
                invalid_product_prefix += 1
            if values[12] != month or values[13] != year:
                invalid_period += 1
            quantity = values[7]
            if quantity is not None and (
                isinstance(quantity, bool)
                or not isinstance(quantity, (int, float))
            ):
                invalid_quantity_type += 1
            if isinstance(quantity, str) and quantity.rstrip().endswith("."):
                trailing_quantity_text += 1
            if cells[7].number_format != quantity_number_format(quantity):
                invalid_quantity_number_format += 1
            sum_quantity += float(values[7] or 0)
            sum_revenue += float(values[9] or 0)
            formula_cells += sum(
                1 for cell in cells if cell.data_type == "f"
            )

        report = {
            "file": str(workbook_path),
            "sheet_count": len(workbook.worksheets),
            "sheet_name": ws.title,
            "headers_ok": headers == OUTPUT_COLUMNS,
            "row_count": row_count,
            "expected_rows": expected_rows,
            "invalid_dates": invalid_dates,
            "invalid_product_prefix": invalid_product_prefix,
            "nontext_product_codes": nontext_product_codes,
            "invalid_period": invalid_period,
            "formula_cells": formula_cells,
            "invalid_quantity_type": invalid_quantity_type,
            "invalid_quantity_number_format": invalid_quantity_number_format,
            "trailing_quantity_text": trailing_quantity_text,
            "sum_quantity": sum_quantity,
            "sum_revenue": sum_revenue,
        }
        report["ok"] = (
            report["sheet_count"] == 1
            and report["sheet_name"] == "Sell in"
            and report["headers_ok"]
            and row_count == expected_rows
            and invalid_dates == 0
            and invalid_product_prefix == 0
            and nontext_product_codes == 0
            and invalid_period == 0
            and formula_cells == 0
            and invalid_quantity_type == 0
            and invalid_quantity_number_format == 0
            and trailing_quantity_text == 0
        )
        return report
    finally:
        workbook.close()


def write_monthly_workbook(
    output_dir: Path,
    rows: list[dict],
    month: int,
    year: int,
) -> dict:
    output_path = output_dir / f"Sell in T{month:02d}_{year}.xlsx"
    temp_path = output_dir / (
        f".{output_path.stem}.{os.getpid()}.tmp.xlsx"
    )
    workbook = build_workbook(rows, month, year)
    try:
        workbook.save(temp_path)
    finally:
        workbook.close()

    try:
        verification = verify_workbook(
            temp_path,
            month,
            year,
            len(rows),
        )
        if not verification["ok"]:
            raise RuntimeError(
                f"Verification failed for {output_path.name}: {verification}"
            )
        try:
            os.replace(temp_path, output_path)
        except PermissionError as exc:
            raise PermissionError(
                f"Close the open workbook before running: {output_path}"
            ) from exc
        verification["file"] = str(output_path)
        return verification
    finally:
        if temp_path.exists():
            temp_path.unlink()
