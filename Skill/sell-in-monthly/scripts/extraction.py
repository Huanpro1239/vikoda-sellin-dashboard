"""Đọc và chuẩn hóa file ERP nguồn.

Module dùng chung cho luồng chính (`extract_sources.py`) và luồng chuyển giao
(`portable_sell_in.py`). Mọi quy tắc lọc và chuẩn hóa dòng nguồn phải nằm ở
đây; không viết lại bản thứ hai cho từng luồng.

`extract_file` trả `NgayHoaDon` dạng `datetime.date` (hoặc `None`). Luồng chính
tự đổi sang chuỗi ISO khi ghi staging JSON; luồng chuyển giao ghi thẳng ngày
Excel thật vào workbook.
"""

from __future__ import annotations

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from normalization import clean_number


OUTPUT_COLUMNS = [
    "Vung",
    "KhuVuc",
    "NgayHoaDon",
    "MaKhachHangMoi",
    "TenKhachHang",
    "MaSanPhamMoi",
    "TenSanPham",
    "SoLuong",
    "DonGia",
    "ThanhTien",
    "LoaiDonHang",
    "GhiChu",
    "Thang",
    "Nam",
]

# 12 cột đầu phải có trong hàng tiêu đề nguồn; `Thang` và `Nam` lấy từ tên file.
REQUIRED_SOURCE_COLUMNS = OUTPUT_COLUMNS[:12]

COLUMN_WIDTHS = [12, 14, 13, 18, 34, 18, 56, 13, 15, 17, 18, 46, 9, 9]

FILE_PATTERN = re.compile(
    r"_(?P<company>Vikoda|Vkoda|VKD)_T(?P<month>\d{1,2})_(?P<year>\d{4})\.(?:xlsx|xlsm)$",
    re.IGNORECASE,
)

HEADER_SEARCH_ROWS = 40

DATE_TEXT_PATTERNS = ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_id(value: Any) -> str:
    """Giữ mã khách hàng và mã sản phẩm ở dạng chữ, không sinh `.0` thừa."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).strip()


def parse_invoice_date(value: Any, epoch: datetime) -> date | None:
    parsed: date | datetime | None = None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value, epoch)
        except (TypeError, ValueError, OverflowError):
            parsed = None
    elif value is not None:
        text = str(value).strip()
        for pattern in DATE_TEXT_PATTERNS:
            try:
                parsed = datetime.strptime(text, pattern)
                break
            except ValueError:
                continue
    if parsed is None:
        return None
    return parsed.date() if isinstance(parsed, datetime) else parsed


def iso_invoice_date(value: date | None) -> str | None:
    return value.isoformat() if value is not None else None


def sort_key(row: dict) -> tuple:
    """Thứ tự dòng trong workbook: theo ngày, rồi mã khách hàng, rồi mã sản phẩm."""
    return (
        row["NgayHoaDon"] or date.min,
        row["MaKhachHangMoi"],
        row["MaSanPhamMoi"],
    )


def locate_header(ws) -> tuple[int, dict[str, int]]:
    required = set(REQUIRED_SOURCE_COLUMNS)
    for row_number in range(1, min(ws.max_row, HEADER_SEARCH_ROWS) + 1):
        values = [clean_text(cell.value) for cell in ws[row_number]]
        mapping = {value: index for index, value in enumerate(values) if value}
        if required.issubset(mapping):
            return row_number, mapping
    raise ValueError(
        "Không tìm thấy hàng tiêu đề chứa đủ các cột bắt buộc trong "
        f"{HEADER_SEARCH_ROWS} hàng đầu của sheet {ws.title}."
    )


def normalize_company(raw_company: str) -> str:
    return "VKD" if raw_company.upper() == "VKD" else "Vikoda"


def parse_source_name(name: str) -> tuple[str, int, int] | None:
    """Đọc công ty, tháng, năm từ tên file nguồn. Trả `None` khi không hợp lệ."""
    match = FILE_PATTERN.search(name)
    if not match:
        return None
    month = int(match.group("month"))
    year = int(match.group("year"))
    if not 1 <= month <= 12:
        return None
    return normalize_company(match.group("company")), month, year


def is_candidate_source(file_path: Path) -> bool:
    """Bỏ file tạm của Excel và các đuôi không phải workbook."""
    return (
        not file_path.name.startswith("~$")
        and file_path.suffix.lower() in {".xlsx", ".xlsm"}
    )


def new_audit(file_path: Path, company: str, month: int, year: int, ws, header_row: int) -> dict:
    return {
        "file": file_path.name,
        "company": company,
        "month": month,
        "year": year,
        "sheet": ws.title,
        "header_row": header_row,
        "source_rows": 0,
        "excluded_blank": 0,
        "excluded_customer_vkd3": 0,
        "excluded_product_prefix": 0,
        "transformed_product_prefix": 0,
        "invalid_invoice_date": 0,
        "output_rows": 0,
        "sum_quantity": 0.0,
        "sum_revenue": 0.0,
    }


def extract_file(
    file_path: Path,
    company: str,
    month: int,
    year: int,
) -> tuple[list[dict], dict]:
    """Đọc một file ERP và trả về các dòng hợp lệ kèm nhật ký đối soát.

    Quy tắc bắt buộc:
      - Với Vikoda/Vkoda, loại `MaKhachHangMoi = VKD3`.
      - Chỉ giữ `MaSanPhamMoi` bắt đầu bằng `1` hoặc `2`.
      - Với VKD, đổi ký tự đầu `2` thành `1`.
      - `Thang`, `Nam` lấy từ tên file nguồn, không lấy từ nội dung.
    """
    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        ws = workbook.worksheets[0]
        header_row, header_map = locate_header(ws)
        rows: list[dict] = []
        audit = new_audit(file_path, company, month, year, ws, header_row)

        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(value not in (None, "") for value in values):
                audit["excluded_blank"] += 1
                continue
            audit["source_rows"] += 1

            customer_code = clean_id(values[header_map["MaKhachHangMoi"]])
            if company == "Vikoda" and customer_code.upper() == "VKD3":
                audit["excluded_customer_vkd3"] += 1
                continue

            product_code = clean_id(values[header_map["MaSanPhamMoi"]])
            if not product_code or product_code[0] not in {"1", "2"}:
                audit["excluded_product_prefix"] += 1
                continue
            if company == "VKD" and product_code.startswith("2"):
                product_code = "1" + product_code[1:]
                audit["transformed_product_prefix"] += 1

            invoice_date = parse_invoice_date(
                values[header_map["NgayHoaDon"]],
                workbook.epoch,
            )
            if invoice_date is None:
                audit["invalid_invoice_date"] += 1

            quantity = clean_number(values[header_map["SoLuong"]])
            unit_price = clean_number(values[header_map["DonGia"]])
            revenue = clean_number(values[header_map["ThanhTien"]])
            audit["sum_quantity"] += float(quantity or 0)
            audit["sum_revenue"] += float(revenue or 0)

            rows.append(
                {
                    "Vung": clean_text(values[header_map["Vung"]]),
                    "KhuVuc": clean_text(values[header_map["KhuVuc"]]),
                    "NgayHoaDon": invoice_date,
                    "MaKhachHangMoi": customer_code,
                    "TenKhachHang": clean_text(
                        values[header_map["TenKhachHang"]]
                    ),
                    "MaSanPhamMoi": product_code,
                    "TenSanPham": clean_text(values[header_map["TenSanPham"]]),
                    "SoLuong": quantity,
                    "DonGia": unit_price,
                    "ThanhTien": revenue,
                    "LoaiDonHang": clean_text(
                        values[header_map["LoaiDonHang"]]
                    ),
                    "GhiChu": clean_text(values[header_map["GhiChu"]]),
                    "Thang": month,
                    "Nam": year,
                }
            )

        audit["output_rows"] = len(rows)
        return rows, audit
    finally:
        workbook.close()


def row_to_values(row: dict) -> list[Any]:
    return [row[column] for column in OUTPUT_COLUMNS]
