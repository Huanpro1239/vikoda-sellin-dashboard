from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from formula_eval import WorkbookEvaluator


TARGET_COLUMNS = [
    "Ky",
    "Nam",
    "Thang",
    "MaKhachHangMoi",
    "TenKhachHang",
    "TargetVikoda",
    "TargetTong",
    "NgayCapNhatNguon",
    "NguonFile",
]

DATA_COLUMNS = [
    "Vung",
    "KhuVuc",
    "NgayHoaDon",
    "MaKhachHangMoi",
    "TenKhachHang",
    "MaSanPhamMoi",
    "TenSanPham",
    "SoLuong",
    "DonGia",
    "ThanhTien",
    "LoaiDonHang",
    "GhiChu",
    "Thang",
    "Nam",
]

ALLOWED_INVOICE_TYPE_KEYS = {
    "DON HANG BAN",
    "DON TRA HANG",
    "HOA DON BAN",
    "HOA DON TRA HANG",
    "NHAP HOA DON TRA HANG",
}

DMKH_COLUMNS = [
    "MaKhachHangMoi",
    "TenKhachHang",
    "TenKhachHangdaydu",
    "DiaChi",
    "KenhBanHang",
    "Loaikhachhang",
    "Hethong MT",
    "MIEN",
    "VUNG",
    "TINHTHANH",
    "QUANHUYEN",
    "CODE",
    "TENKH",
    "ThongTinBoSungNguon",
]

PVT_DATA_COLUMNS = [
    "MIEN",
    "VUNG",
    "MaKH",
    "KhachHang",
    "SanPham",
    "Actual",
    "CungKyLY",
    "ThangTruoc",
    "Vikoda",
    "TargetTong",
    "TargetVikoda",
    "KDT",
    "VikodaLY",
    "VikodaThangTruoc",
]

PIVOT_COLUMNS = [
    "Sales Region",
    "Area",
    "Total Target",
    "Vikoda Target",
    "Total MTD",
    "Total % (3) vs (1)",
    "Total Gap",
    "Vikoda MTD",
    "Vikoda % (6) vs (2)",
    "Vikoda Gap",
    "KDT MTD",
    "Total Last Year",
    "Total % (3) vs (10)",
    "Vikoda Last Year",
    "Vikoda % (6) vs (12)",
    "Total Last Month",
    "Total % (3) vs (14)",
    "Vikoda Last Month",
    "Vikoda % (6) vs (16)",
]

# Bo cuc PIVOT: 1 tieu de, 2 ky bao cao, 3 nhom cot, 4 ten cot, 5 so thu tu cot.
PIVOT_HEADER_ROW = 4
PIVOT_NUMBER_ROW = 5
PIVOT_FIRST_DATA_ROW = 6
# 20 vung + 8 dong tong mien + 1 Grand Total.
PIVOT_GRAND_TOTAL_ROW = PIVOT_FIRST_DATA_ROW + 20 + 8

BC_COLUMNS = [
    "Vùng / Khách hàng / Sản phẩm",
    "Actual",
    "Cùng kỳ LY",
    "% vs LY",
    "Tháng trước",
    "% vs TT",
    "Vikoda",
    "Target",
    "Target Vikoda",
    "% đạt Target",
]

BC_HEADER_ROW = 10
BC_FIRST_DATA_ROW = 11

# Chi tieu doi chieu giua sheet BC_ va PIVOT.
BC_FIELD_COLUMNS = {
    "actual": 2,
    "last_year": 3,
    "prior_month": 5,
    "vikoda_actual": 7,
    "target_total": 8,
    "target_vikoda": 9,
}

REPORTING_STRUCTURE = [
    ("Miền Bắc", ["Bắc Miền Trung", "Đông Bắc", "Hà Nội", "Tây Bắc"]),
    ("Miền Nam", ["Miền Đông", "Miền Tây", "TP. HCM 1", "TP. HCM 2"]),
    ("Miền Trung 1", ["Miền Trung 1A", "Miền Trung 1B", "Tây Nguyên"]),
    ("Miền Trung 2", ["Miền Trung 2A", "Miền Trung 2B"]),
    ("KA", ["KA Miền Bắc", "KA Miền Trung 1", "KA Miền Trung 2", "KA Miền Nam"]),
    ("MT", ["MT"]),
    ("B2C", ["B2C"]),
    ("Other", ["Other"]),
]


def normalized_key(value) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.replace("\u00a0", " ").split())
    text = text.replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(
        character for character in text
        if not unicodedata.combining(character)
    )
    return text.upper()


def is_revenue_invoice_type(value) -> bool:
    return normalized_key(value) in ALLOWED_INVOICE_TYPE_KEYS


def period_key(year, month) -> str:
    return f"{int(year):04d}{int(month):02d}"


def previous_period(year: int, month: int) -> str:
    if month > 1:
        return period_key(year, month - 1)
    return period_key(year - 1, 12)


def expected_pivot_totals(sell_in_payload, target_payload):
    current_year = int(sell_in_payload["current_year"])
    current_month = int(sell_in_payload["through_month"])
    current = period_key(current_year, current_month)
    last_year = period_key(current_year - 1, current_month)
    prior_month = previous_period(current_year, current_month)
    totals = {
        "target_total": 0,
        "target_vikoda": 0,
        "actual": 0,
        "vikoda_actual": 0,
        "kdt_actual": 0,
        "last_year": 0,
        "vikoda_last_year": 0,
        "prior_month": 0,
        "vikoda_prior_month": 0,
    }
    for record in target_payload["records"]:
        if str(record["Ky"]) == current:
            totals["target_total"] += int(record["TargetTong"])
            totals["target_vikoda"] += int(record["TargetVikoda"])
    for row in sell_in_payload["rows"]:
        row_period = period_key(row[13], row[12])
        amount = int(round(row[9]))
        product_key = normalized_key(row[6])
        if row_period == current:
            totals["actual"] += amount
            if "VIKODA" in product_key:
                totals["vikoda_actual"] += amount
            if "KDT" in product_key:
                totals["kdt_actual"] += amount
        elif row_period == last_year:
            totals["last_year"] += amount
            if "VIKODA" in product_key:
                totals["vikoda_last_year"] += amount
        elif row_period == prior_month:
            totals["prior_month"] += amount
            if "VIKODA" in product_key:
                totals["vikoda_prior_month"] += amount
    return {
        "periods": {
            "current": current,
            "last_year": last_year,
            "prior_month": prior_month,
        },
        "totals": totals,
    }


def close_enough(actual, expected, tolerance: float = 0.5) -> bool:
    if not isinstance(actual, (int, float)):
        return False
    return abs(float(actual) - float(expected)) <= tolerance


def summarize_target(records):
    result = defaultdict(lambda: {"records": 0, "target_vikoda": 0, "target_total": 0})
    for record in records:
        item = result[str(record["Ky"])]
        item["records"] += 1
        item["target_vikoda"] += int(record["TargetVikoda"])
        item["target_total"] += int(record["TargetTong"])
    return dict(result)


def normalized_decimal(value: Decimal) -> str:
    return str(value.normalize())


def expected_sell_in_summary(audit):
    result = {}
    for report in audit["periods"]:
        result[str(report["period"])] = {
            "records": int(report["rows"]),
            "quantity": normalized_decimal(Decimal(str(report["quantity"]))),
            "amount": int(report["amount"]),
        }
    return result


def add_problem(problems: list[str], message: str) -> None:
    if len(problems) < 200:
        problems.append(message)


def main() -> int:
    parser = argparse.ArgumentParser(description="Kiem tra Bao_Cao_Sell_in.xlsx")
    parser.add_argument("--data-file", required=True)
    parser.add_argument("--sell-in-data-file", required=True)
    parser.add_argument("--sell-in-audit-file", required=True)
    parser.add_argument("--dmkh-data-file", required=True)
    parser.add_argument("--dmkh-audit-file", required=True)
    parser.add_argument("--output-file", required=True)
    parser.add_argument("--report-file", required=True)
    args = parser.parse_args()

    target_data_file = Path(args.data_file).resolve()
    sell_in_data_file = Path(args.sell_in_data_file).resolve()
    sell_in_audit_file = Path(args.sell_in_audit_file).resolve()
    dmkh_data_file = Path(args.dmkh_data_file).resolve()
    dmkh_audit_file = Path(args.dmkh_audit_file).resolve()
    output_file = Path(args.output_file).resolve()
    report_file = Path(args.report_file).resolve()
    report_file.parent.mkdir(parents=True, exist_ok=True)

    target_payload = json.loads(target_data_file.read_text(encoding="utf-8"))
    sell_in_payload = json.loads(sell_in_data_file.read_text(encoding="utf-8"))
    sell_in_audit = json.loads(sell_in_audit_file.read_text(encoding="utf-8"))
    dmkh_payload = json.loads(dmkh_data_file.read_text(encoding="utf-8"))
    dmkh_audit = json.loads(dmkh_audit_file.read_text(encoding="utf-8"))
    expected_target_summary = summarize_target(target_payload["records"])
    expected_data_summary = expected_sell_in_summary(sell_in_audit)
    pivot_expected = expected_pivot_totals(sell_in_payload, target_payload)

    problems: list[str] = []
    filter_definition = sell_in_audit.get("invoice_type_filter", {})
    if filter_definition.get("source_column") != "LoaiDonHang":
        add_problem(
            problems,
            "Audit Sell In thiếu quy tắc lọc Loại hóa đơn từ cột LoaiDonHang",
        )

    audited_source_rows = 0
    audited_included_rows = 0
    audited_excluded_rows = 0
    audited_excluded_amount = 0
    for period_report in sell_in_audit.get("periods", []):
        source_rows = int(period_report.get("source_rows", -1))
        included_rows = int(period_report.get("rows", -1))
        excluded_rows = int(period_report.get("excluded_rows", -1))
        if source_rows != included_rows + excluded_rows:
            add_problem(
                problems,
                f"Audit Sell In kỳ {period_report.get('period')}: "
                "source_rows không bằng rows + excluded_rows",
            )

        type_source_rows = 0
        type_included_rows = 0
        type_excluded_rows = 0
        type_excluded_amount = 0
        for item in period_report.get("invoice_types", []):
            item_rows = int(item.get("rows", 0))
            item_amount = int(item.get("amount", 0))
            expected_included = is_revenue_invoice_type(
                item.get("invoice_type")
            )
            if bool(item.get("included")) != expected_included:
                add_problem(
                    problems,
                    f"Audit Sell In kỳ {period_report.get('period')}: "
                    f"phân loại sai {item.get('invoice_type')!r}",
                )
            type_source_rows += item_rows
            if expected_included:
                type_included_rows += item_rows
            else:
                type_excluded_rows += item_rows
                type_excluded_amount += item_amount

        if (
            type_source_rows != source_rows
            or type_included_rows != included_rows
            or type_excluded_rows != excluded_rows
            or type_excluded_amount
            != int(period_report.get("excluded_amount", 0))
        ):
            add_problem(
                problems,
                f"Audit loại hóa đơn kỳ {period_report.get('period')} "
                "không khớp tổng kỳ",
            )

        audited_source_rows += source_rows
        audited_included_rows += included_rows
        audited_excluded_rows += excluded_rows
        audited_excluded_amount += int(
            period_report.get("excluded_amount", 0)
        )

    if (
        audited_source_rows != int(sell_in_audit.get("total_source_rows", -1))
        or audited_included_rows != int(sell_in_audit.get("total_rows", -1))
        or audited_excluded_rows
        != int(sell_in_audit.get("total_excluded_rows", -1))
        or audited_excluded_amount
        != int(sell_in_audit.get("total_excluded_amount", 0))
    ):
        add_problem(
            problems,
            "Tổng audit lọc loại hóa đơn không khớp chi tiết từng kỳ",
        )

    for row_number, row in enumerate(sell_in_payload["rows"], start=2):
        if not is_revenue_invoice_type(row[10]):
            add_problem(
                problems,
                f"Staging Sell In còn loại hóa đơn không hợp lệ tại "
                f"hàng {row_number}: {row[10]!r}",
            )

    actual_target_records: list[dict] = []
    actual_data_summary_raw = defaultdict(
        lambda: {"records": 0, "quantity": Decimal(0), "amount": 0}
    )
    actual_data_rows = 0
    actual_dmkh_rows = 0
    actual_pvt_data_rows = 0
    actual_pivot_totals: dict[str, int | float | None] = {}

    if not output_file.exists():
        add_problem(problems, f"Khong tim thay file dau ra: {output_file}")
    else:
        # openpyxl khong ghi kem gia tri da tinh cho cong thuc, nen bo kiem tra
        # tu tinh lai bang WorkbookEvaluator thay vi doc gia tri Excel dem san.
        workbook = load_workbook(output_file, read_only=False, data_only=False)
        try:
            expected_sheet_names = [
                "Target",
                "Data",
                "DMKH",
                "PIVOT",
                "PVT_DATA",
            ] + [f"BC_{area}" for area, _ in REPORTING_STRUCTURE]
            if workbook.sheetnames != expected_sheet_names:
                add_problem(
                    problems,
                    f"Sheet khong dung: {workbook.sheetnames}; "
                    f"can {expected_sheet_names}",
                )

            if "Target" in workbook.sheetnames:
                worksheet = workbook["Target"]
                headers = [worksheet.cell(1, column).value for column in range(1, 10)]
                if headers != TARGET_COLUMNS:
                    add_problem(problems, f"Tieu de Target khong dung: {headers}")
                if worksheet.freeze_panes != "A2":
                    add_problem(
                        problems, f"Freeze panes Target khong dung: {worksheet.freeze_panes}"
                    )
                if "tblTarget" not in worksheet.tables:
                    add_problem(problems, "Khong co Excel Table tblTarget")

                exact_keys = set()
                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        min_col=1,
                        max_col=9,
                        values_only=True,
                    ),
                    start=2,
                ):
                    if not any(value is not None for value in values):
                        continue
                    period, year, month, code, name, vikoda, total, _, _ = values
                    key = (str(period), str(code), str(name))
                    if key in exact_keys:
                        add_problem(
                            problems, f"Dong Target trung tai hang {row_number}: {key}"
                        )
                    exact_keys.add(key)
                    if not isinstance(period, str) or not period.isdigit() or len(period) != 6:
                        add_problem(
                            problems, f"Ky Target khong dung tai hang {row_number}: {period!r}"
                        )
                    if not isinstance(year, int) or not isinstance(month, int):
                        add_problem(
                            problems,
                            f"Nam/Thang Target khong phai so nguyen tai hang {row_number}",
                        )
                    if not isinstance(code, str):
                        add_problem(
                            problems,
                            f"Ma khach hang Target khong phai chu tai hang {row_number}",
                        )
                    if not isinstance(vikoda, (int, float)) or not isinstance(
                        total, (int, float)
                    ):
                        add_problem(
                            problems, f"Target khong phai so tai hang {row_number}"
                        )
                        continue
                    if vikoda < 0 or total < 0 or total < vikoda:
                        add_problem(
                            problems, f"Target khong hop le tai hang {row_number}"
                        )
                    actual_target_records.append(
                        {
                            "Ky": str(period),
                            "TargetVikoda": int(round(vikoda)),
                            "TargetTong": int(round(total)),
                        }
                    )
                    for column, value in enumerate(values, start=1):
                        if isinstance(value, str) and value.startswith("="):
                            add_problem(
                                problems,
                                f"Khong duoc co cong thuc tai Target!"
                                f"{worksheet.cell(row_number, column).coordinate}",
                            )
                if worksheet.max_row - 1 != len(target_payload["records"]):
                    add_problem(
                        problems,
                        f"So dong Target khong khop: output={worksheet.max_row - 1}, "
                        f"staging={len(target_payload['records'])}",
                    )

            if "Data" in workbook.sheetnames:
                worksheet = workbook["Data"]
                headers = [worksheet.cell(1, column).value for column in range(1, 15)]
                if headers != DATA_COLUMNS:
                    add_problem(problems, f"Tieu de Data khong dung: {headers}")
                if worksheet.freeze_panes != "A2":
                    add_problem(
                        problems, f"Freeze panes Data khong dung: {worksheet.freeze_panes}"
                    )
                if "tblDataSellIn" not in worksheet.tables:
                    add_problem(problems, "Khong co Excel Table tblDataSellIn")

                expected_formats = {
                    "C2": "dd/mm/yyyy",
                    "D2": "@",
                    "F2": "@",
                    "I2": "#,##0",
                    "J2": "#,##0",
                    "M2": "0",
                    "N2": "0",
                }
                for coordinate, expected_format in expected_formats.items():
                    actual_format = worksheet[coordinate].number_format
                    if actual_format != expected_format:
                        add_problem(
                            problems,
                            f"Dinh dang {coordinate}={actual_format!r}, can {expected_format!r}",
                        )

                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        min_col=1,
                        max_col=14,
                        values_only=True,
                    ),
                    start=2,
                ):
                    if not any(value is not None for value in values):
                        continue
                    actual_data_rows += 1
                    (
                        _,
                        _,
                        invoice_date,
                        customer_code,
                        _,
                        product_code,
                        _,
                        quantity,
                        unit_price,
                        amount,
                        invoice_type,
                        _,
                        month,
                        year,
                    ) = values
                    if not is_revenue_invoice_type(invoice_type):
                        add_problem(
                            problems,
                            f"Data còn loại hóa đơn không hợp lệ tại "
                            f"hàng {row_number}: {invoice_type!r}",
                        )
                    if not isinstance(invoice_date, (date, datetime)):
                        add_problem(
                            problems,
                            f"NgayHoaDon khong phai ngay Excel tai hang {row_number}",
                        )
                    if not isinstance(customer_code, str):
                        add_problem(
                            problems,
                            f"MaKhachHangMoi khong phai chu tai hang {row_number}",
                        )
                    if not isinstance(product_code, str):
                        add_problem(
                            problems,
                            f"MaSanPhamMoi khong phai chu tai hang {row_number}",
                        )
                    if not all(
                        isinstance(value, (int, float))
                        for value in (quantity, unit_price, amount)
                    ):
                        add_problem(
                            problems,
                            f"SoLuong/DonGia/ThanhTien khong phai so tai hang {row_number}",
                        )
                        continue
                    expected_quantity_format = (
                        "#,##0"
                        if Decimal(str(quantity))
                        == Decimal(str(quantity)).to_integral_value()
                        else "#,##0.##"
                    )
                    actual_quantity_format = worksheet.cell(
                        row_number, 8
                    ).number_format
                    if actual_quantity_format != expected_quantity_format:
                        add_problem(
                            problems,
                            f"Dinh dang SoLuong tai H{row_number}="
                            f"{actual_quantity_format!r}, can "
                            f"{expected_quantity_format!r}",
                        )
                    if not isinstance(month, int) or not isinstance(year, int):
                        add_problem(
                            problems,
                            f"Thang/Nam Data khong phai so nguyen tai hang {row_number}",
                        )
                        continue
                    period = f"{year}{month:02d}"
                    if period not in expected_data_summary:
                        add_problem(
                            problems,
                            f"Ky Data ngoai pham vi tai hang {row_number}: {period}",
                        )
                    summary = actual_data_summary_raw[period]
                    summary["records"] += 1
                    summary["quantity"] += Decimal(str(quantity))
                    summary["amount"] += int(round(amount))
                    for column, value in enumerate(values, start=1):
                        if isinstance(value, str) and value.startswith("="):
                            add_problem(
                                problems,
                                f"Khong duoc co cong thuc tai Data!"
                                f"{worksheet.cell(row_number, column).coordinate}",
                            )
                if worksheet.max_row - 1 != len(sell_in_payload["rows"]):
                    add_problem(
                        problems,
                        f"So dong Data khong khop: output={worksheet.max_row - 1}, "
                        f"staging={len(sell_in_payload['rows'])}",
                    )

            if "DMKH" in workbook.sheetnames:
                worksheet = workbook["DMKH"]
                headers = [worksheet.cell(1, column).value for column in range(1, 15)]
                if headers != DMKH_COLUMNS:
                    add_problem(problems, f"Tieu de DMKH khong dung: {headers}")
                if worksheet.freeze_panes != "A2":
                    add_problem(
                        problems, f"Freeze panes DMKH khong dung: {worksheet.freeze_panes}"
                    )
                if "tblDMKH" not in worksheet.tables:
                    add_problem(problems, "Khong co Excel Table tblDMKH")

                for row_number, expected_row in enumerate(
                    dmkh_payload["rows"],
                    start=2,
                ):
                    actual_row = [
                        worksheet.cell(row_number, column).value
                        for column in range(1, 15)
                    ]
                    normalized_actual = [
                        "" if value is None else str(value) for value in actual_row
                    ]
                    normalized_expected = [
                        "" if value is None else str(value) for value in expected_row
                    ]
                    if normalized_actual != normalized_expected:
                        add_problem(
                            problems,
                            f"Du lieu DMKH khong khop staging tai hang {row_number}",
                        )
                    for column, value in enumerate(actual_row, start=1):
                        cell = worksheet.cell(row_number, column)
                        if cell.number_format != "@":
                            add_problem(
                                problems,
                                f"Dinh dang DMKH {cell.coordinate}="
                                f"{cell.number_format!r}, can '@'",
                            )
                        if isinstance(value, str) and value.startswith("="):
                            add_problem(
                                problems,
                                f"Khong duoc co cong thuc tai DMKH!{cell.coordinate}",
                            )
                    actual_dmkh_rows += 1

                if worksheet.max_row - 1 != len(dmkh_payload["rows"]):
                    add_problem(
                        problems,
                        f"So dong DMKH khong khop: output={worksheet.max_row - 1}, "
                        f"staging={len(dmkh_payload['rows'])}",
                    )

            model_by_region = defaultdict(
                lambda: {
                    "target_total": 0,
                    "target_vikoda": 0,
                    "actual": 0,
                    "vikoda_actual": 0,
                    "kdt_actual": 0,
                    "last_year": 0,
                    "vikoda_last_year": 0,
                    "prior_month": 0,
                    "vikoda_prior_month": 0,
                }
            )
            model_totals = {
                key: 0
                for key in pivot_expected["totals"]
            }
            valid_reporting_pairs = {
                (area, region)
                for area, regions in REPORTING_STRUCTURE
                for region in regions
            }
            if "PVT_DATA" in workbook.sheetnames:
                worksheet = workbook["PVT_DATA"]
                headers = [
                    worksheet.cell(1, column).value
                    for column in range(1, 15)
                ]
                if headers != PVT_DATA_COLUMNS:
                    add_problem(
                        problems,
                        f"Tieu de PVT_DATA khong dung: {headers}",
                    )
                if worksheet.sheet_state != "hidden":
                    add_problem(
                        problems,
                        f"PVT_DATA phai an, hien tai={worksheet.sheet_state!r}",
                    )
                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        min_col=1,
                        max_col=14,
                        values_only=True,
                    ),
                    start=2,
                ):
                    if not any(value is not None for value in values):
                        continue
                    actual_pvt_data_rows += 1
                    (
                        area,
                        region,
                        customer_code,
                        customer_name,
                        product,
                        actual,
                        last_year,
                        prior_month,
                        vikoda_actual,
                        target_total,
                        target_vikoda,
                        kdt_actual,
                        vikoda_last_year,
                        vikoda_prior_month,
                    ) = values
                    pair = (str(area), str(region))
                    if pair not in valid_reporting_pairs:
                        add_problem(
                            problems,
                            f"Mien/Vung PVT_DATA khong hop le tai hang "
                            f"{row_number}: {pair}",
                        )
                    if (
                        (
                            normalized_key(customer_code) == "B2C"
                            or normalized_key(customer_name).startswith("B2C")
                        )
                        and (target_total or target_vikoda)
                        and pair != ("B2C", "B2C")
                    ):
                        add_problem(
                            problems,
                            f"Target B2C bị ánh xạ sai tại PVT_DATA hàng "
                            f"{row_number}: {pair}",
                        )
                    if (
                        normalized_key(customer_name) == "OTHER"
                        and (target_total or target_vikoda)
                        and pair != ("Other", "Other")
                    ):
                        add_problem(
                            problems,
                            f"Target Other bị ánh xạ sai tại PVT_DATA hàng "
                            f"{row_number}: {pair}",
                        )
                    numbers = [
                        actual,
                        last_year,
                        prior_month,
                        vikoda_actual,
                        target_total,
                        target_vikoda,
                        kdt_actual,
                        vikoda_last_year,
                        vikoda_prior_month,
                    ]
                    if not all(
                        isinstance(value, (int, float))
                        for value in numbers
                    ):
                        add_problem(
                            problems,
                            f"So lieu PVT_DATA khong phai so tai hang "
                            f"{row_number}",
                        )
                        continue
                    if any(
                        worksheet.cell(row_number, column).data_type == "f"
                        for column in range(1, 15)
                    ):
                        add_problem(
                            problems,
                            f"PVT_DATA khong duoc co cong thuc tai hang "
                            f"{row_number}",
                        )

                    product_key = normalized_key(product)
                    metrics = model_by_region[pair]
                    metrics["actual"] += int(round(actual))
                    metrics["last_year"] += int(round(last_year))
                    metrics["prior_month"] += int(round(prior_month))
                    metrics["vikoda_actual"] += int(round(vikoda_actual))
                    metrics["target_total"] += int(round(target_total))
                    metrics["target_vikoda"] += int(round(target_vikoda))
                    metrics["kdt_actual"] += int(round(kdt_actual))
                    metrics["vikoda_last_year"] += int(round(vikoda_last_year))
                    metrics["vikoda_prior_month"] += int(
                        round(vikoda_prior_month)
                    )
                    expected_kdt = actual if "KDT" in product_key else 0
                    expected_vikoda_last_year = (
                        last_year if "VIKODA" in product_key else 0
                    )
                    expected_vikoda_prior_month = (
                        prior_month if "VIKODA" in product_key else 0
                    )
                    for label, actual_value, expected_value in (
                        ("KDT", kdt_actual, expected_kdt),
                        (
                            "VikodaLY",
                            vikoda_last_year,
                            expected_vikoda_last_year,
                        ),
                        (
                            "VikodaThangTruoc",
                            vikoda_prior_month,
                            expected_vikoda_prior_month,
                        ),
                    ):
                        if not close_enough(actual_value, expected_value):
                            add_problem(
                                problems,
                                f"{label} PVT_DATA sai tai hang {row_number}",
                            )

                for metrics in model_by_region.values():
                    for key, value in metrics.items():
                        model_totals[key] += value
                for key, expected in pivot_expected["totals"].items():
                    if not close_enough(model_totals[key], expected):
                        add_problem(
                            problems,
                            f"Tong PVT_DATA {key} khong khop nguon: "
                            f"output={model_totals[key]}, expected={expected}",
                        )

            if "PIVOT" in workbook.sheetnames:
                evaluator = WorkbookEvaluator(workbook)
                worksheet = workbook["PIVOT"]
                headers = [
                    worksheet.cell(PIVOT_HEADER_ROW, column).value
                    for column in range(1, len(PIVOT_COLUMNS) + 1)
                ]
                if headers != PIVOT_COLUMNS:
                    add_problem(problems, f"Tieu de PIVOT khong dung: {headers}")
                if worksheet.sheet_state != "visible":
                    add_problem(
                        problems,
                        f"PIVOT phai hien, hien tai={worksheet.sheet_state!r}",
                    )
                if worksheet.freeze_panes != "C6":
                    add_problem(
                        problems,
                        f"Freeze panes PIVOT khong dung: {worksheet.freeze_panes}",
                    )

                # Dong so thu tu cot (1)...(17) de ten cot phan tram tham chieu duoc.
                for column in range(3, len(PIVOT_COLUMNS) + 1):
                    value = str(
                        worksheet.cell(PIVOT_NUMBER_ROW, column).value or ""
                    )
                    if not value.endswith(f"({column - 2})"):
                        add_problem(
                            problems,
                            f"Thieu so thu tu cot ({column - 2}) tai PIVOT hang "
                            f"{PIVOT_NUMBER_ROW}",
                        )

                expected_formula_cells = 29 * 17
                formula_cells = sum(
                    1
                    for row in worksheet.iter_rows(
                        min_row=PIVOT_FIRST_DATA_ROW,
                        max_row=PIVOT_GRAND_TOTAL_ROW,
                        min_col=3,
                        max_col=len(PIVOT_COLUMNS),
                    )
                    for cell in row
                    if cell.data_type == "f"
                )
                if formula_cells != expected_formula_cells:
                    add_problem(
                        problems,
                        f"So cong thuc PIVOT khong dung: "
                        f"{formula_cells}, can {expected_formula_cells}",
                    )

                metric_columns = {
                    "target_total": 3,
                    "target_vikoda": 4,
                    "actual": 5,
                    "vikoda_actual": 8,
                    "kdt_actual": 11,
                    "last_year": 12,
                    "vikoda_last_year": 14,
                    "prior_month": 16,
                    "vikoda_prior_month": 18,
                }
                row_number = PIVOT_FIRST_DATA_ROW
                for area, regions in REPORTING_STRUCTURE:
                    for region in regions:
                        if (
                            worksheet.cell(row_number, 1).value != area
                            or worksheet.cell(row_number, 2).value != region
                        ):
                            add_problem(
                                problems,
                                f"Nhan PIVOT sai tai hang {row_number}",
                            )
                        metrics = model_by_region[(area, region)]
                        for key, column in metric_columns.items():
                            actual_value = evaluator.cell_value(
                                "PIVOT", row_number, column
                            )
                            if not close_enough(actual_value, metrics[key]):
                                add_problem(
                                    problems,
                                    f"PIVOT {area}/{region} {key} khong khop: "
                                    f"output={actual_value}, "
                                    f"expected={metrics[key]}",
                                )
                        row_number += 1
                    if worksheet.cell(row_number, 1).value != f"{area} Total":
                        add_problem(
                            problems,
                            f"Nhan tong PIVOT sai tai hang {row_number}",
                        )
                    row_number += 1

                grand_total_row = PIVOT_GRAND_TOTAL_ROW
                if worksheet.cell(grand_total_row, 1).value != "Grand Total":
                    add_problem(
                        problems,
                        f"Khong tim thay Grand Total tai PIVOT!A{grand_total_row}",
                    )
                for key, column in metric_columns.items():
                    actual_value = evaluator.cell_value(
                        "PIVOT", grand_total_row, column
                    )
                    actual_pivot_totals[key] = actual_value
                    expected_value = pivot_expected["totals"][key]
                    if not close_enough(actual_value, expected_value):
                        add_problem(
                            problems,
                            f"PIVOT Grand Total {key} khong khop: "
                            f"output={actual_value}, expected={expected_value}",
                        )

                totals = pivot_expected["totals"]

                def ratio(numerator: str, denominator: str) -> float:
                    return (
                        totals[numerator] / totals[denominator]
                        if totals[denominator]
                        else 0
                    )

                expected_ratios = {
                    6: ratio("actual", "target_total"),
                    9: ratio("vikoda_actual", "target_vikoda"),
                    13: ratio("actual", "last_year"),
                    15: ratio("vikoda_actual", "vikoda_last_year"),
                    17: ratio("actual", "prior_month"),
                    19: ratio("vikoda_actual", "vikoda_prior_month"),
                }
                for column, expected_value in expected_ratios.items():
                    actual_value = evaluator.cell_value(
                        "PIVOT", grand_total_row, column
                    )
                    if not close_enough(
                        actual_value, expected_value, tolerance=1e-9
                    ):
                        add_problem(
                            problems,
                            f"Ty le PIVOT tai cot {column} hang "
                            f"{grand_total_row} khong khop: "
                            f"output={actual_value}, expected={expected_value}",
                        )

            # Cac sheet bao cao chi tiet theo mien.
            bc_sheet_names = [
                name for name in workbook.sheetnames if name.startswith("BC_")
            ]
            expected_bc_names = [f"BC_{area}" for area, _ in REPORTING_STRUCTURE]
            if bc_sheet_names != expected_bc_names:
                add_problem(
                    problems,
                    f"Sheet bao cao mien khong dung: {bc_sheet_names}; "
                    f"can {expected_bc_names}",
                )
            else:
                evaluator = WorkbookEvaluator(workbook)
                bc_totals = {key: 0.0 for key in BC_FIELD_COLUMNS}
                for area, regions in REPORTING_STRUCTURE:
                    ws = workbook[f"BC_{area}"]
                    headers = [
                        ws.cell(BC_HEADER_ROW, column).value
                        for column in range(1, len(BC_COLUMNS) + 1)
                    ]
                    if headers != BC_COLUMNS:
                        add_problem(
                            problems,
                            f"Tieu de BC_{area} khong dung: {headers}",
                        )
                    if ws.sheet_state != "visible":
                        add_problem(problems, f"BC_{area} phai hien")

                    region_rows: list[int] = []
                    customer_rows: list[int] = []
                    product_rows: list[int] = []
                    grand_row = None
                    row_number = BC_FIRST_DATA_ROW
                    while True:
                        label = ws.cell(row_number, 1).value
                        if label is None:
                            break
                        if label == "Grand Total":
                            grand_row = row_number
                            break
                        level = ws.row_dimensions[row_number].outlineLevel or 0
                        if level == 0:
                            region_rows.append(row_number)
                        elif level == 1:
                            customer_rows.append(row_number)
                        else:
                            product_rows.append(row_number)
                        row_number += 1

                    if grand_row is None:
                        add_problem(problems, f"BC_{area} thieu dong Grand Total")
                        continue
                    if [ws.cell(item, 1).value for item in region_rows] != regions:
                        add_problem(
                            problems,
                            f"BC_{area} thieu hoac sai thu tu vung ban hang",
                        )

                    # Moi khach hang phai bang tong cac dong san pham cua no.
                    current_customer = None
                    children: dict[int, list[int]] = {}
                    for item in sorted(customer_rows + product_rows):
                        if item in customer_rows:
                            current_customer = item
                            children[item] = []
                        elif current_customer is not None:
                            children[current_customer].append(item)
                    for customer_row, product_list in children.items():
                        for column in BC_FIELD_COLUMNS.values():
                            parent = evaluator.cell_value(
                                f"BC_{area}", customer_row, column
                            )
                            child_total = sum(
                                float(ws.cell(item, column).value or 0)
                                for item in product_list
                            )
                            if not close_enough(parent, child_total):
                                add_problem(
                                    problems,
                                    f"BC_{area} hang {customer_row} cot {column}: "
                                    f"khach hang={parent} khac tong san pham="
                                    f"{child_total}",
                                )

                    for key, column in BC_FIELD_COLUMNS.items():
                        value = evaluator.cell_value(f"BC_{area}", grand_row, column)
                        bc_totals[key] += value
                        expected_value = sum(
                            model_by_region[(area, region)][key]
                            for region in regions
                        )
                        if not close_enough(value, expected_value):
                            add_problem(
                                problems,
                                f"BC_{area} Grand Total {key} khong khop PIVOT: "
                                f"output={value}, expected={expected_value}",
                            )

                for key, column in BC_FIELD_COLUMNS.items():
                    expected_value = pivot_expected["totals"][key]
                    if not close_enough(bc_totals[key], expected_value):
                        add_problem(
                            problems,
                            f"Tong 8 sheet BC_ cho {key} khong khop PIVOT: "
                            f"output={bc_totals[key]}, expected={expected_value}",
                        )
        finally:
            workbook.close()

    actual_target_summary = summarize_target(actual_target_records)
    if actual_target_summary != expected_target_summary:
        add_problem(
            problems, "Tong so dong hoac tong Target theo ky khong khop staging"
        )

    actual_data_summary = {
        period: {
            "records": values["records"],
            "quantity": normalized_decimal(values["quantity"]),
            "amount": values["amount"],
        }
        for period, values in sorted(actual_data_summary_raw.items())
    }
    if actual_data_summary != expected_data_summary:
        add_problem(
            problems,
            "So dong, SoLuong hoac ThanhTien Sell In theo ky khong khop staging",
        )

    report = {
        "target_data_file": str(target_data_file),
        "sell_in_data_file": str(sell_in_data_file),
        "sell_in_audit_file": str(sell_in_audit_file),
        "dmkh_data_file": str(dmkh_data_file),
        "dmkh_audit_file": str(dmkh_audit_file),
        "output_file": str(output_file),
        "target": {
            "expected_records": len(target_payload["records"]),
            "actual_records": len(actual_target_records),
            "expected_periods": expected_target_summary,
            "actual_periods": actual_target_summary,
        },
        "data": {
            "expected_records": len(sell_in_payload["rows"]),
            "actual_records": actual_data_rows,
            "source_records": sell_in_audit.get("total_source_rows"),
            "excluded_records": sell_in_audit.get("total_excluded_rows"),
            "excluded_amount": sell_in_audit.get("total_excluded_amount"),
            "invoice_type_filter": sell_in_audit.get("invoice_type_filter"),
            "expected_periods": expected_data_summary,
            "actual_periods": actual_data_summary,
        },
        "dmkh": {
            "expected_records": len(dmkh_payload["rows"]),
            "actual_records": actual_dmkh_rows,
            "unique_nonblank_customer_codes": dmkh_audit.get(
                "unique_nonblank_customer_codes"
            ),
            "duplicate_customer_codes": len(
                dmkh_audit.get("duplicate_customer_codes", {})
            ),
            "warnings": dmkh_audit.get("warnings", []),
        },
        "pivot": {
            "periods": pivot_expected["periods"],
            "expected_totals": pivot_expected["totals"],
            "actual_totals": actual_pivot_totals,
            "pvt_data_records": actual_pvt_data_rows,
        },
        "problems": problems,
    }
    report_file.write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "report_file": str(report_file),
                "target_records": len(actual_target_records),
                "data_records": actual_data_rows,
                "dmkh_records": actual_dmkh_rows,
                "pvt_data_records": actual_pvt_data_rows,
                "target_periods": len(actual_target_summary),
                "data_periods": len(actual_data_summary),
                "problems": len(problems),
            },
            ensure_ascii=False,
        )
    )
    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
