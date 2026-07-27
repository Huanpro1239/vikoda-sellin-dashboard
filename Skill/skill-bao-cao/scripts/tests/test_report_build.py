"""Khóa mô hình số liệu, bộ tính công thức và bố cục PIVOT / BC_.

Dựng một bộ dữ liệu nhỏ rồi kiểm tra workbook ra đúng bố cục file mẫu và mọi
con số cộng khớp nhau, để lần sửa sau không phải mở Excel mới biết sai.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path


SCRIPT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(SCRIPT_DIR / "vendor"))

from openpyxl import Workbook, load_workbook  # noqa: E402

import build_pivot_sheet as builder  # noqa: E402
from formula_eval import FormulaError, WorkbookEvaluator  # noqa: E402
from report_model import (  # noqa: E402
    REPORTING_STRUCTURE,
    build_model,
    normalize_reporting_pair,
    period_key,
    previous_period,
)


# Cột Sell In staging: [Vung, KhuVuc, NgayHoaDon, MaKH, TenKH, MaSP, TenSP,
#                       SoLuong, DonGia, ThanhTien, LoaiDonHang, GhiChu,
#                       Thang, Nam]
def sell_in_row(code, product, amount, month, year, name="KHACH HANG"):
    return [
        "", "", "2026-07-10", code, name, "130100001", product,
        1, amount, amount, "Đơn hàng bán", "", month, year,
    ]


def dmkh_row(code, mien, vung):
    return [code, "TEN", "", "", "", "", "", mien, vung, "", "", "", "", ""]


def payloads():
    sell_in = {
        "current_year": 2026,
        "through_month": 7,
        "as_of_date": "2026-07-27",
        "rows": [
            sell_in_row("KH1", "Vikoda 1.5L", 100, 7, 2026),
            sell_in_row("KH1", "KDT 500ml", 40, 7, 2026),
            sell_in_row("KH1", "Vikoda 1.5L", 60, 7, 2025),
            sell_in_row("KH1", "Vikoda 1.5L", 80, 6, 2026),
            sell_in_row("KH2", "Nuoc suoi", 30, 7, 2026),
            sell_in_row("KH9", "Vikoda 1.5L", 25, 7, 2026),  # không có trong DMKH
        ],
    }
    target = {
        "records": [
            {
                "Ky": "202607", "MaKhachHangMoi": "KH1",
                "TenKhachHang": "KHACH HANG 1",
                "TargetTong": 500, "TargetVikoda": 300,
                "MienBaoCao": "", "VungBaoCao": "",
            },
            {
                "Ky": "202607", "MaKhachHangMoi": "B2C",
                "TenKhachHang": "B2C",
                "TargetTong": 200, "TargetVikoda": 100,
                "MienBaoCao": "", "VungBaoCao": "",
            },
            {   # kỳ khác, phải bị bỏ qua
                "Ky": "202606", "MaKhachHangMoi": "KH1",
                "TenKhachHang": "KHACH HANG 1",
                "TargetTong": 999, "TargetVikoda": 999,
                "MienBaoCao": "", "VungBaoCao": "",
            },
        ]
    }
    dmkh = {
        "rows": [
            dmkh_row("KH1", "Miền Bắc", "Hà Nội"),
            dmkh_row("KH2", "Miền Bắc", "Hà Nội"),
        ]
    }
    return sell_in, target, dmkh


class ModelTest(unittest.TestCase):
    def setUp(self) -> None:
        self.model = build_model(*payloads())
        self.rows = self.model.rows()

    def test_ky_hien_tai_cung_ky_va_thang_truoc(self) -> None:
        self.assertEqual(self.model.current_period, "202607")
        self.assertEqual(self.model.last_year_period, "202507")
        self.assertEqual(self.model.prior_month_period, "202606")

    def test_thang_1_lui_ve_thang_12_nam_truoc(self) -> None:
        self.assertEqual(previous_period(2026, 1), "202512")
        self.assertEqual(period_key(2026, 7), "202607")

    def test_vikoda_va_kdt_nhan_dien_theo_ten_san_pham(self) -> None:
        by_product = {row[4]: row for row in self.rows if row[2] == "KH1"}
        self.assertEqual(by_product["Vikoda 1.5L"][8], 100)   # cột Vikoda
        self.assertEqual(by_product["Vikoda 1.5L"][11], 0)    # cột KDT
        self.assertEqual(by_product["KDT 500ml"][8], 0)
        self.assertEqual(by_product["KDT 500ml"][11], 40)

    def test_cung_ky_va_thang_truoc_tach_dung_cot(self) -> None:
        row = next(
            r for r in self.rows
            if r[2] == "KH1" and r[4] == "Vikoda 1.5L"
        )
        self.assertEqual(row[5], 100)   # Actual
        self.assertEqual(row[6], 60)    # CungKyLY
        self.assertEqual(row[7], 80)    # ThangTruoc
        self.assertEqual(row[12], 60)   # VikodaLY
        self.assertEqual(row[13], 80)   # VikodaThangTruoc

    def test_khach_hang_ngoai_dmkh_roi_vao_other(self) -> None:
        row = next(r for r in self.rows if r[2] == "KH9")
        self.assertEqual((row[0], row[1]), ("Other", "Other"))
        self.assertEqual(self.model.mapping_stats["data_fallback_other"], 1)

    def test_target_b2c_luon_vao_vung_b2c(self) -> None:
        row = next(r for r in self.rows if r[2] == "B2C")
        self.assertEqual((row[0], row[1]), ("B2C", "B2C"))

    def test_chi_lay_target_ky_hien_tai(self) -> None:
        total = sum(row[9] for row in self.rows)
        self.assertEqual(total, 700)

    def test_ma_trung_mien_vung_mau_thuan_bi_chan(self) -> None:
        sell_in, target, dmkh = payloads()
        dmkh["rows"].append(dmkh_row("KH1", "Miền Nam", "Miền Tây"))
        with self.assertRaises(ValueError):
            build_model(sell_in, target, dmkh)


class ReportingPairTest(unittest.TestCase):
    def test_20_vung_ban_hang(self) -> None:
        pairs = [(a, r) for a, regions in REPORTING_STRUCTURE for r in regions]
        self.assertEqual(len(pairs), 20)
        self.assertEqual(len(REPORTING_STRUCTURE), 8)

    def test_ka_tach_rieng_khoi_vung_npp(self) -> None:
        self.assertEqual(
            normalize_reporting_pair("KA", "Miền Trung 1"),
            ("KA", "KA Miền Trung 1"),
        )


class FormulaEvalTest(unittest.TestCase):
    def setUp(self) -> None:
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "PVT_DATA"
        ws.append(["MIEN", "VUNG", "Val"])
        for mien, vung, val in [
            ("Miền Bắc", "Hà Nội", 10),
            ("Miền Bắc", "Hà Nội", 5),
            ("Miền Bắc", "Tây Bắc", 7),
            ("Miền Nam", "Miền Tây", 3),
        ]:
            ws.append([mien, vung, val])
        self.sheet = self.wb.create_sheet("S")
        self.ev = WorkbookEvaluator(self.wb)

    def test_sumifs_mot_dieu_kien(self) -> None:
        self.sheet["A1"] = "=SUMIFS('PVT_DATA'!$C$2:$C$5,'PVT_DATA'!$A$2:$A$5,\"Miền Bắc\")"
        self.assertEqual(self.ev.cell_value("S", 1, 1), 22)

    def test_sumifs_hai_dieu_kien(self) -> None:
        self.sheet["A1"] = (
            "=SUMIFS('PVT_DATA'!$C$2:$C$5,'PVT_DATA'!$A$2:$A$5,\"Miền Bắc\","
            "'PVT_DATA'!$B$2:$B$5,\"Hà Nội\")"
        )
        self.assertEqual(self.ev.cell_value("S", 1, 1), 15)

    def test_sumifs_dieu_kien_tham_chieu_o(self) -> None:
        self.sheet["B1"] = "Miền Nam"
        self.sheet["A1"] = "=SUMIFS('PVT_DATA'!$C$2:$C$5,'PVT_DATA'!$A$2:$A$5,$B1)"
        self.assertEqual(self.ev.cell_value("S", 1, 1), 3)

    def test_sum_dai_va_danh_sach(self) -> None:
        self.sheet["A1"], self.sheet["A2"], self.sheet["A3"] = 1, 2, 4
        self.sheet["B1"] = "=SUM(A1:A3)"
        self.sheet["B2"] = "=SUM(A1,A3)"
        self.assertEqual(self.ev.cell_value("S", 1, 2), 7)
        self.assertEqual(self.ev.cell_value("S", 2, 2), 5)

    def test_iferror_khi_chia_cho_khong(self) -> None:
        self.sheet["A1"], self.sheet["A2"] = 10, 0
        self.sheet["B1"] = "=IFERROR(A1/A2,0)"
        self.assertEqual(self.ev.cell_value("S", 1, 2), 0)

    def test_chia_va_tru(self) -> None:
        self.sheet["A1"], self.sheet["A2"] = 10, 4
        self.sheet["B1"] = "=IFERROR(A1/A2,0)"
        self.sheet["B2"] = "=A1-A2"
        self.assertEqual(self.ev.cell_value("S", 1, 2), 2.5)
        self.assertEqual(self.ev.cell_value("S", 2, 2), 6)

    def test_cong_thuc_la_bi_bao_loi(self) -> None:
        self.sheet["A1"] = "=VLOOKUP(A2,B:C,2,FALSE)"
        with self.assertRaises(FormulaError):
            self.ev.cell_value("S", 1, 1)


class WorkbookLayoutTest(unittest.TestCase):
    """Dựng workbook thật từ dữ liệu nhỏ rồi soi bố cục và số liệu."""

    @classmethod
    def setUpClass(cls) -> None:
        sell_in, target, dmkh = payloads()
        model = build_model(sell_in, target, dmkh)
        rows = model.rows()

        wb = Workbook()
        for name in ("Target", "Data", "DMKH"):
            wb.create_sheet(name)
        del wb[wb.sheetnames[0]]
        pivot_ws = wb.create_sheet("PIVOT")
        pvt_ws = wb.create_sheet("PVT_DATA")
        last = builder.build_pvt_data_sheet(pvt_ws, rows)
        builder.build_pivot_sheet(pivot_ws, last, "07/2026", "2026-07-27")
        tree = builder.order_tree(builder.build_tree(rows))
        for area, _ in REPORTING_STRUCTURE:
            builder.build_bc_sheet(
                wb.create_sheet(builder.sheet_title_for(area)),
                area, tree[area], last, "07/2026", "2026-07-27",
            )
        cls.folder = tempfile.TemporaryDirectory()
        path = Path(cls.folder.name) / "report.xlsx"
        wb.save(path)
        cls.wb = load_workbook(path)
        cls.ev = WorkbookEvaluator(cls.wb)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.wb.close()
        cls.folder.cleanup()

    def test_du_13_sheet(self) -> None:
        expected = ["Target", "Data", "DMKH", "PIVOT", "PVT_DATA"] + [
            f"BC_{area}" for area, _ in REPORTING_STRUCTURE
        ]
        self.assertEqual(self.wb.sheetnames, expected)

    def test_bo_cuc_pivot_theo_file_mau(self) -> None:
        ws = self.wb["PIVOT"]
        self.assertEqual(ws["A1"].value, "DAILY REVENUE SELL IN REPORT: VIKODA")
        self.assertEqual(ws["E2"].value, "Report for:")
        self.assertEqual(ws["A3"].value, "MARKET STRUCTURE")
        self.assertEqual(ws["A4"].value, "Sales Region")
        self.assertEqual(ws["B4"].value, "Area")
        self.assertEqual(ws["F4"].value, "Total % (3) vs (1)")
        self.assertEqual(ws["A6"].value, "Miền Bắc")
        self.assertEqual(ws.freeze_panes, "C6")

    def test_dong_so_thu_tu_cot(self) -> None:
        ws = self.wb["PIVOT"]
        for column in range(3, 20):
            self.assertTrue(
                str(ws.cell(5, column).value).endswith(f"({column - 2})"),
                f"cột {column}",
            )

    def test_pivot_du_493_cong_thuc(self) -> None:
        ws = self.wb["PIVOT"]
        count = sum(
            1
            for row in ws.iter_rows(min_row=6, max_row=34, min_col=3, max_col=19)
            for cell in row
            if cell.data_type == "f"
        )
        self.assertEqual(count, 493)

    def test_pivot_du_dong_vung_tong_mien_va_grand_total(self) -> None:
        ws = self.wb["PIVOT"]
        self.assertEqual(ws["A34"].value, "Grand Total")
        labels = [ws.cell(r, 1).value for r in range(6, 34)]
        subtotals = [f"{area} Total" for area, _ in REPORTING_STRUCTURE]
        self.assertEqual([x for x in labels if x in subtotals], subtotals)
        regions = [ws.cell(r, 2).value for r in range(6, 35)]
        self.assertEqual(sum(1 for x in regions if x), 20)

    def test_tien_hien_thi_theo_trieu_dong(self) -> None:
        ws = self.wb["PIVOT"]
        self.assertIn(",,", ws.cell(6, 3).number_format)
        self.assertIn("%", ws.cell(6, 6).number_format)

    def test_grand_total_pivot_khop_nguon(self) -> None:
        # Target 500 + 200; Actual 100+40+30+25
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 3), 700)
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 5), 195)
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 8), 125)   # Vikoda
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 11), 40)   # KDT
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 12), 60)   # cùng kỳ
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 16), 80)   # tháng trước

    def test_gap_va_ty_le_la_cong_thuc(self) -> None:
        self.assertEqual(self.ev.cell_value("PIVOT", 34, 7), 195 - 700)
        self.assertAlmostEqual(self.ev.cell_value("PIVOT", 34, 6), 195 / 700)

    def test_pvt_data_du_cot_va_du_dong(self) -> None:
        """Việc ẩn PVT_DATA do main() làm; ở đây chỉ kiểm tra nội dung."""
        ws = self.wb["PVT_DATA"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[:5], ["MIEN", "VUNG", "MaKH", "KhachHang", "SanPham"])
        self.assertEqual(len(headers), 14)
        self.assertGreater(ws.max_row, 1)

    def test_bc_co_ba_cap_va_thu_gon_duoc(self) -> None:
        ws = self.wb["BC_Miền Bắc"]
        self.assertEqual(ws.cell(10, 1).value, "Vùng / Khách hàng / Sản phẩm")
        self.assertFalse(ws.sheet_properties.outlinePr.summaryBelow)
        levels = {
            ws.row_dimensions[r].outlineLevel or 0
            for r in range(11, ws.max_row)
            if ws.cell(r, 1).value and ws.cell(r, 1).value != "Grand Total"
        }
        self.assertEqual(levels, {0, 1, 2})

    def test_bc_du_vung_cua_mien(self) -> None:
        ws = self.wb["BC_Miền Bắc"]
        regions = [
            ws.cell(r, 1).value
            for r in range(11, ws.max_row)
            if (ws.row_dimensions[r].outlineLevel or 0) == 0
            and ws.cell(r, 1).value
            and ws.cell(r, 1).value != "Grand Total"
        ]
        self.assertEqual(
            regions, ["Bắc Miền Trung", "Đông Bắc", "Hà Nội", "Tây Bắc"]
        )

    def test_bc_khach_hang_bang_tong_san_pham(self) -> None:
        ws = self.wb["BC_Miền Bắc"]
        customer_row = next(
            r for r in range(11, ws.max_row)
            if (ws.row_dimensions[r].outlineLevel or 0) == 1
        )
        children = []
        row = customer_row + 1
        while (ws.row_dimensions[row].outlineLevel or 0) == 2:
            children.append(row)
            row += 1
        self.assertTrue(children)
        for column in (2, 3, 5, 7, 8, 9):
            parent = self.ev.cell_value("BC_Miền Bắc", customer_row, column)
            total = sum(float(ws.cell(r, column).value or 0) for r in children)
            self.assertAlmostEqual(parent, total, msg=f"cột {column}")

    def test_bc_grand_total_khop_pivot(self) -> None:
        """Tổng 8 sheet BC_ phải bằng Grand Total của PIVOT."""
        for pivot_column, bc_column in ((5, 2), (12, 3), (16, 5), (8, 7), (3, 8)):
            bc_total = sum(
                self.ev.cell_value(
                    f"BC_{area}",
                    self._grand_row(f"BC_{area}"),
                    bc_column,
                )
                for area, _ in REPORTING_STRUCTURE
            )
            pivot_total = self.ev.cell_value("PIVOT", 34, pivot_column)
            self.assertAlmostEqual(
                bc_total, pivot_total, msg=f"cột PIVOT {pivot_column}"
            )

    def _grand_row(self, sheet_name: str) -> int:
        ws = self.wb[sheet_name]
        for row in range(11, ws.max_row + 1):
            if ws.cell(row, 1).value == "Grand Total":
                return row
        raise AssertionError(f"{sheet_name} thiếu Grand Total")

    def test_kpi_dau_trang_tro_toi_grand_total(self) -> None:
        ws = self.wb["BC_Miền Bắc"]
        grand = self._grand_row("BC_Miền Bắc")
        self.assertEqual(ws.cell(6, 2).value, "ACTUAL")
        self.assertEqual(ws.cell(7, 2).value, f"=B{grand}")
        self.assertEqual(
            self.ev.cell_value("BC_Miền Bắc", 7, 2),
            self.ev.cell_value("BC_Miền Bắc", grand, 2),
        )


if __name__ == "__main__":
    unittest.main()
