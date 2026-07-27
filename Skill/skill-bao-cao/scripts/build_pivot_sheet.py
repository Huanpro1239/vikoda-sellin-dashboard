"""Dựng PVT_DATA, PIVOT và 8 sheet báo cáo chi tiết theo miền.

Thay `build_pivot_sheet.mjs` + `merge_pivot_workbook.py`: ghi thẳng vào workbook
báo cáo bằng openpyxl nên không còn cần Node, `@oai/artifact-tool`, cũng không
phải ghép file bằng thao tác XML.

Bố cục PIVOT theo file mẫu `Sell_in_report_chuan.xlsb`: tiêu đề tiếng Anh, dòng
nhóm, và dòng đánh số thứ tự cột (1)…(17) để tên cột phần trăm tham chiếu được
tới cột gốc.

Tiền hiển thị theo triệu đồng bằng định dạng `#,##0,,` nhưng ô vẫn lưu đủ VND,
nên cộng trừ không phát sinh sai số làm tròn.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from report_model import (
    NO_PRODUCT,
    PVT_COLUMN,
    PVT_HEADERS,
    REPORTING_STRUCTURE,
    build_model,
)


# ----------------------------------------------------------------------------
# Bảng màu và định dạng dùng chung
# ----------------------------------------------------------------------------

FONT = "Aptos"
TITLE_FONT = "Aptos Display"

NAVY = "FF1F3864"
BLUE = "FF2F5597"
GREY = "FF595959"
GOLD = "FFBF9000"
DEEP_BLUE = "FF1F4E79"
PURPLE = "FF7030A0"
TEAL = "FF007060"
WHITE = "FFFFFFFF"
LIGHT_GREY = "FFD9D9D9"
BAND_BLUE = "FFD9E2F3"
SOFT_GOLD = "FFFFF9E6"
SOFT_BLUE = "FFF2F7FC"

# Tiền: lưu VND, hiển thị triệu đồng. Mỗi dấu phẩy cuối chuỗi chia cho 1.000.
MONEY_FORMAT = '#,##0,,;(#,##0,,);"-"'
PERCENT_FORMAT = '0.0%;(0.0%);"-"'
RAW_MONEY_FORMAT = '#,##0;(#,##0);"-"'

THIN = Side(style="thin", color=LIGHT_GREY)
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Ô chứa "(1)" bị Excel hiểu là số âm; thêm ký tự rỗng để giữ nguyên dạng chữ.
ZWSP = "​"


def money(cell) -> None:
    cell.number_format = MONEY_FORMAT
    cell.alignment = Alignment(horizontal="right", vertical="center")


def percent(cell) -> None:
    cell.number_format = PERCENT_FORMAT
    cell.alignment = Alignment(horizontal="right", vertical="center")


# ----------------------------------------------------------------------------
# PIVOT: 19 cột A..S
# ----------------------------------------------------------------------------

PIVOT_FIRST_DATA_ROW = 6

PIVOT_HEADERS = [
    "Sales Region",
    "Area",
    "Total Target",
    "Vikoda Target",
    "Total MTD",
    "Total % (3) vs (1)",
    "Total Gap",
    "Vikoda MTD",
    "Vikoda % (6) vs (2)",
    "Vikoda Gap",
    "KDT MTD",
    "Total Last Year",
    "Total % (3) vs (10)",
    "Vikoda Last Year",
    "Vikoda % (6) vs (12)",
    "Total Last Month",
    "Total % (3) vs (14)",
    "Vikoda Last Month",
    "Vikoda % (6) vs (16)",
]

# (nhãn nhóm, cột đầu, cột cuối, màu nền)
PIVOT_SECTIONS = [
    ("MARKET STRUCTURE", 1, 2, GREY),
    ("TARGET", 3, 4, GOLD),
    ("MTD / GAP", 5, 11, DEEP_BLUE),
    ("LAST YEAR", 12, 15, PURPLE),
    ("LAST MONTH", 16, 19, TEAL),
]

PIVOT_WIDTHS = [16, 20, 14, 14, 14, 13, 14, 14, 13, 14, 13, 15, 13, 15, 13, 16, 13, 16, 13]

# Cột tiền và cột phần trăm trong PIVOT.
PIVOT_MONEY_COLUMNS = [3, 4, 5, 7, 8, 10, 11, 12, 14, 16, 18]
PIVOT_PERCENT_COLUMNS = [6, 9, 13, 15, 17, 19]
# Cột chênh lệch: âm đỏ, dương xanh.
PIVOT_GAP_COLUMNS = [7, 10]
# Cột tỉ lệ đạt Target: tô theo ngưỡng.
PIVOT_ATTAINMENT_COLUMNS = [6, 9]


def sumifs_by_area_region(pvt_field: str, pvt_last_row: int, row: int) -> str:
    """Tổng một cột PVT_DATA theo cặp Miền/Vùng của dòng PIVOT."""
    column = get_column_letter(PVT_COLUMN[pvt_field])
    return (
        f"=SUMIFS('PVT_DATA'!${column}$2:${column}${pvt_last_row},"
        f"'PVT_DATA'!$A$2:$A${pvt_last_row},$A{row},"
        f"'PVT_DATA'!$B$2:$B${pvt_last_row},$B{row})"
    )


def pivot_metric_formulas(row: int, amounts: dict[str, str]) -> list[str]:
    """17 cột chỉ tiêu, cột phần trăm và chênh lệch luôn là công thức."""
    return [
        amounts["target_total"],                 # C  (1)
        amounts["target_vikoda"],                # D  (2)
        amounts["actual"],                       # E  (3)
        f"=IFERROR(E{row}/C{row},0)",            # F  (4)
        f"=E{row}-C{row}",                       # G  (5)
        amounts["vikoda"],                       # H  (6)
        f"=IFERROR(H{row}/D{row},0)",            # I  (7)
        f"=H{row}-D{row}",                       # J  (8)
        amounts["kdt"],                          # K  (9)
        amounts["last_year"],                    # L  (10)
        f"=IFERROR(E{row}/L{row},0)",            # M  (11)
        amounts["vikoda_last_year"],             # N  (12)
        f"=IFERROR(H{row}/N{row},0)",            # O  (13)
        amounts["prior_month"],                  # P  (14)
        f"=IFERROR(E{row}/P{row},0)",            # Q  (15)
        amounts["vikoda_prior_month"],           # R  (16)
        f"=IFERROR(H{row}/R{row},0)",            # S  (17)
    ]


def build_pivot_sheet(
    ws: Worksheet,
    pvt_last_row: int,
    period_label: str,
    as_of_date: str,
) -> dict[str, Any]:
    ws.sheet_view.showGridLines = False
    last_column = len(PIVOT_HEADERS)
    last_letter = get_column_letter(last_column)

    # Dòng 1-2: tiêu đề và thông tin kỳ.
    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"] = "DAILY REVENUE SELL IN REPORT: VIKODA"
    ws["A1"].font = Font(name=TITLE_FONT, size=15, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws.merge_cells("E2:H2")
    ws["E2"] = "Report for:"
    ws.merge_cells("I2:K2")
    ws["I2"] = as_of_date
    ws.merge_cells(f"L2:{last_letter}2")
    ws["L2"] = f"Period: {period_label} | Unit: VND mn | Sales + Returns"
    for coordinate in ("E2", "I2", "L2"):
        ws[coordinate].font = Font(name=FONT, size=10, italic=True, color=GREY)
        ws[coordinate].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Dòng 3: nhóm cột.
    for label, first, last, color in PIVOT_SECTIONS:
        ws.merge_cells(
            start_row=3, start_column=first, end_row=3, end_column=last
        )
        cell = ws.cell(3, first, label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # Dòng 4: tên cột, tô theo màu nhóm.
    section_color = {}
    for _, first, last, color in PIVOT_SECTIONS:
        for column in range(first, last + 1):
            section_color[column] = color
    for index, title in enumerate(PIVOT_HEADERS, start=1):
        cell = ws.cell(4, index, title)
        cell.fill = PatternFill("solid", fgColor=section_color[index])
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = GRID
    ws.row_dimensions[4].height = 48

    # Dòng 5: số thứ tự cột để tên cột phần trăm tham chiếu được.
    for index in range(3, last_column + 1):
        cell = ws.cell(5, index, f"{ZWSP}({index - 2})")
        cell.font = Font(name=FONT, size=9, bold=True, color=GREY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = GRID
    ws.row_dimensions[5].height = 16

    # Dòng dữ liệu.
    row = PIVOT_FIRST_DATA_ROW
    detail_rows: list[int] = []
    subtotal_rows: list[int] = []
    for area, regions in REPORTING_STRUCTURE:
        first_detail = row
        for region in regions:
            ws.cell(row, 1, area)
            ws.cell(row, 2, region)
            amounts = {
                "target_total": sumifs_by_area_region("TargetTong", pvt_last_row, row),
                "target_vikoda": sumifs_by_area_region("TargetVikoda", pvt_last_row, row),
                "actual": sumifs_by_area_region("Actual", pvt_last_row, row),
                "vikoda": sumifs_by_area_region("Vikoda", pvt_last_row, row),
                "kdt": sumifs_by_area_region("KDT", pvt_last_row, row),
                "last_year": sumifs_by_area_region("CungKyLY", pvt_last_row, row),
                "vikoda_last_year": sumifs_by_area_region("VikodaLY", pvt_last_row, row),
                "prior_month": sumifs_by_area_region("ThangTruoc", pvt_last_row, row),
                "vikoda_prior_month": sumifs_by_area_region(
                    "VikodaThangTruoc", pvt_last_row, row
                ),
            }
            for offset, formula in enumerate(pivot_metric_formulas(row, amounts)):
                ws.cell(row, 3 + offset, formula)
            detail_rows.append(row)
            row += 1

        last_detail = row - 1
        ws.cell(row, 1, f"{area} Total")

        def block_sum(column_index: int) -> str:
            letter = get_column_letter(column_index)
            return f"=SUM({letter}{first_detail}:{letter}{last_detail})"

        amounts = {
            "target_total": block_sum(3),
            "target_vikoda": block_sum(4),
            "actual": block_sum(5),
            "vikoda": block_sum(8),
            "kdt": block_sum(11),
            "last_year": block_sum(12),
            "vikoda_last_year": block_sum(14),
            "prior_month": block_sum(16),
            "vikoda_prior_month": block_sum(18),
        }
        for offset, formula in enumerate(pivot_metric_formulas(row, amounts)):
            ws.cell(row, 3 + offset, formula)
        subtotal_rows.append(row)
        row += 1

    grand_total_row = row
    ws.cell(grand_total_row, 1, "Grand Total")

    def total_of_subtotals(column_index: int) -> str:
        letter = get_column_letter(column_index)
        return "=SUM(" + ",".join(f"{letter}{item}" for item in subtotal_rows) + ")"

    amounts = {
        "target_total": total_of_subtotals(3),
        "target_vikoda": total_of_subtotals(4),
        "actual": total_of_subtotals(5),
        "vikoda": total_of_subtotals(8),
        "kdt": total_of_subtotals(11),
        "last_year": total_of_subtotals(12),
        "vikoda_last_year": total_of_subtotals(14),
        "prior_month": total_of_subtotals(16),
        "vikoda_prior_month": total_of_subtotals(18),
    }
    for offset, formula in enumerate(pivot_metric_formulas(grand_total_row, amounts)):
        ws.cell(grand_total_row, 3 + offset, formula)

    # Định dạng vùng dữ liệu.
    for current in range(PIVOT_FIRST_DATA_ROW, grand_total_row + 1):
        for column in range(1, last_column + 1):
            cell = ws.cell(current, column)
            cell.font = Font(name=FONT, size=10)
            cell.border = GRID
            if column <= 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif column in PIVOT_PERCENT_COLUMNS:
                percent(cell)
            else:
                money(cell)
        for column in (3, 4):
            ws.cell(current, column).fill = PatternFill("solid", fgColor=SOFT_GOLD)
        for column in range(5, 12):
            ws.cell(current, column).fill = PatternFill("solid", fgColor=SOFT_BLUE)
        ws.row_dimensions[current].height = 18

    for current in subtotal_rows:
        for column in range(1, last_column + 1):
            cell = ws.cell(current, column)
            cell.fill = PatternFill("solid", fgColor=BAND_BLUE)
            cell.font = Font(name=FONT, size=10, bold=True)
    for column in range(1, last_column + 1):
        cell = ws.cell(grand_total_row, column)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)

    for index, width in enumerate(PIVOT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "C6"

    # Tô màu theo ngưỡng: dưới 80% đỏ, 80-100% vàng, từ 100% xanh.
    for column in PIVOT_ATTAINMENT_COLUMNS:
        letter = get_column_letter(column)
        area = f"{letter}{PIVOT_FIRST_DATA_ROW}:{letter}{grand_total_row}"
        ws.conditional_formatting.add(area, CellIsRule(
            operator="lessThan", formula=["0.8"],
            fill=PatternFill("solid", fgColor="FFFFC7CE"),
            font=Font(color="FF9C0006"),
        ))
        ws.conditional_formatting.add(area, CellIsRule(
            operator="between", formula=["0.8", "0.999999"],
            fill=PatternFill("solid", fgColor="FFFFEB9C"),
            font=Font(color="FF9C6500"),
        ))
        ws.conditional_formatting.add(area, CellIsRule(
            operator="greaterThanOrEqual", formula=["1"],
            fill=PatternFill("solid", fgColor="FFC6EFCE"),
            font=Font(color="FF006100"),
        ))
    for column in PIVOT_GAP_COLUMNS:
        letter = get_column_letter(column)
        area = f"{letter}{PIVOT_FIRST_DATA_ROW}:{letter}{grand_total_row}"
        ws.conditional_formatting.add(area, CellIsRule(
            operator="lessThan", formula=["0"], font=Font(color="FFC00000"),
        ))
        ws.conditional_formatting.add(area, CellIsRule(
            operator="greaterThan", formula=["0"], font=Font(color="FF00703C"),
        ))

    return {
        "first_data_row": PIVOT_FIRST_DATA_ROW,
        "detail_rows": detail_rows,
        "subtotal_rows": subtotal_rows,
        "grand_total_row": grand_total_row,
        "last_column": last_column,
    }


# ----------------------------------------------------------------------------
# PVT_DATA
# ----------------------------------------------------------------------------

PVT_WIDTHS = [18, 22, 18, 40, 58, 16, 16, 16, 16, 18, 18, 16, 16, 18]


def build_pvt_data_sheet(ws: Worksheet, rows: list[list]) -> int:
    ws.sheet_view.showGridLines = False
    ws.append(PVT_HEADERS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=GREY)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 30

    last_row = len(rows) + 1
    for current in range(2, last_row + 1):
        for column in range(1, 6):
            cell = ws.cell(current, column)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for column in range(6, len(PVT_HEADERS) + 1):
            cell = ws.cell(current, column)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = RAW_MONEY_FORMAT

    for index, width in enumerate(PVT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    return last_row


# ----------------------------------------------------------------------------
# BC_<Miền>: Vùng -> Khách hàng -> Sản phẩm
# ----------------------------------------------------------------------------

BC_HEADERS = [
    "Vùng / Khách hàng / Sản phẩm",
    "Actual",
    "Cùng kỳ LY",
    "% vs LY",
    "Tháng trước",
    "% vs TT",
    "Vikoda",
    "Target",
    "Target Vikoda",
    "% đạt Target",
]

BC_WIDTHS = [58, 16, 16, 11, 16, 11, 16, 16, 16, 12]

BC_TITLE_ROW = 2
BC_SUBTITLE_ROW = 3
BC_KPI_LABEL_ROW = 6
BC_KPI_VALUE_ROW = 7
BC_HEADER_ROW = 10
BC_FIRST_DATA_ROW = 11

# Cột PVT_DATA lấy số cho từng cột BC_.
BC_SUM_FIELDS = {
    2: "Actual",
    3: "CungKyLY",
    5: "ThangTruoc",
    7: "Vikoda",
    8: "TargetTong",
    9: "TargetVikoda",
}
BC_PERCENT_COLUMNS = {4: ("B", "C"), 6: ("B", "E"), 10: ("B", "H")}

BC_KPI_LABELS = [
    (2, "ACTUAL"),
    (4, "CÙNG KỲ LY"),
    (6, "% VS LY"),
    (8, "TARGET"),
    (10, "% ĐẠT TARGET"),
]


def sheet_title_for(area: str) -> str:
    """Tên sheet Excel: tối đa 31 ký tự, không chứa : \\ / ? * [ ]."""
    name = f"BC_{area}"
    for character in ':\\/?*[]':
        name = name.replace(character, " ")
    return name[:31]


def bc_sumifs(field: str, pvt_last_row: int, criteria: list[tuple[str, str]]) -> str:
    column = get_column_letter(PVT_COLUMN[field])
    parts = [f"'PVT_DATA'!${column}$2:${column}${pvt_last_row}"]
    for criteria_column, value in criteria:
        parts.append(
            f"'PVT_DATA'!${criteria_column}$2:${criteria_column}${pvt_last_row},{value}"
        )
    return "=SUMIFS(" + ",".join(parts) + ")"


def quote(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def build_bc_sheet(
    ws: Worksheet,
    area: str,
    tree: dict[str, dict[str, dict[str, list[float]]]],
    pvt_last_row: int,
    period_label: str,
    as_of_date: str,
) -> dict[str, Any]:
    """Một sheet chi tiết cho một miền.

    Dòng vùng và dòng khách hàng dùng SUMIFS về PVT_DATA để số liệu luôn khớp
    PIVOT. Dòng sản phẩm là dữ liệu lá nên ghi thẳng giá trị, tránh hàng chục
    nghìn công thức SUMIFS làm Excel mở chậm.
    """
    ws.sheet_view.showGridLines = False
    last_column = len(BC_HEADERS)
    last_letter = get_column_letter(last_column)

    ws.merge_cells(f"A{BC_TITLE_ROW}:{last_letter}{BC_TITLE_ROW}")
    title_cell = ws.cell(BC_TITLE_ROW, 1, f"BÁO CÁO CHI TIẾT — {area.upper()}")
    title_cell.font = Font(name=TITLE_FONT, size=14, bold=True, color=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[BC_TITLE_ROW].height = 26

    ws.merge_cells(f"A{BC_SUBTITLE_ROW}:{last_letter}{BC_SUBTITLE_ROW}")
    subtitle = (
        f"Kỳ báo cáo: {period_label}  |  Đơn vị: triệu đồng  |  "
        f"Đơn bán + trả hàng  |  Vikoda/KDT theo tên sản phẩm"
    )
    if as_of_date:
        subtitle += f"  |  Cập nhật: {as_of_date}"
    subtitle_cell = ws.cell(BC_SUBTITLE_ROW, 1, subtitle)
    subtitle_cell.font = Font(name=FONT, size=9, italic=True, color=GREY)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Bảng dữ liệu dựng trước để KPI tham chiếu tới dòng Grand Total.
    for index, title in enumerate(BC_HEADERS, start=1):
        cell = ws.cell(BC_HEADER_ROW, index, title)
        cell.fill = PatternFill("solid", fgColor=DEEP_BLUE)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = GRID
    ws.row_dimensions[BC_HEADER_ROW].height = 34

    row = BC_FIRST_DATA_ROW
    region_rows: list[int] = []
    customer_rows: list[int] = []
    product_rows: list[int] = []

    for region, customers in tree.items():
        region_row = row
        ws.cell(region_row, 1, region)
        base = [("A", quote(area)), ("B", quote(region))]
        for column, field in BC_SUM_FIELDS.items():
            ws.cell(region_row, column, bc_sumifs(field, pvt_last_row, base))
        region_rows.append(region_row)
        row += 1

        for (code, customer), products in customers.items():
            customer_row = row
            ws.cell(customer_row, 1, customer)
            criteria = base + [("C", quote(code))]
            for column, field in BC_SUM_FIELDS.items():
                ws.cell(customer_row, column, bc_sumifs(field, pvt_last_row, criteria))
            ws.row_dimensions[customer_row].outlineLevel = 1
            customer_rows.append(customer_row)
            row += 1

            for product, values in products.items():
                label = product if product != NO_PRODUCT else "(chưa gắn sản phẩm)"
                ws.cell(row, 1, f"    {label}")
                for column, field in BC_SUM_FIELDS.items():
                    ws.cell(row, column, values[field])
                ws.row_dimensions[row].outlineLevel = 2
                ws.row_dimensions[row].hidden = True
                product_rows.append(row)
                row += 1

    grand_total_row = row
    ws.cell(grand_total_row, 1, "Grand Total")
    for column in BC_SUM_FIELDS:
        letter = get_column_letter(column)
        formula = "=SUM(" + ",".join(f"{letter}{item}" for item in region_rows) + ")"
        ws.cell(grand_total_row, column, formula if region_rows else 0)

    # Cột phần trăm cho mọi dòng.
    all_rows = sorted(region_rows + customer_rows + product_rows) + [grand_total_row]
    for current in all_rows:
        for column, (numerator, denominator) in BC_PERCENT_COLUMNS.items():
            ws.cell(
                current,
                column,
                f"=IFERROR({numerator}{current}/{denominator}{current},0)",
            )

    # Định dạng.
    region_set, customer_set = set(region_rows), set(customer_rows)
    for current in all_rows:
        for column in range(1, last_column + 1):
            cell = ws.cell(current, column)
            cell.border = GRID
            if column == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif column in BC_PERCENT_COLUMNS:
                percent(cell)
            else:
                money(cell)
        if current == grand_total_row:
            for column in range(1, last_column + 1):
                ws.cell(current, column).fill = PatternFill("solid", fgColor=BLUE)
                ws.cell(current, column).font = Font(
                    name=FONT, size=10, bold=True, color=WHITE
                )
        elif current in region_set:
            for column in range(1, last_column + 1):
                ws.cell(current, column).fill = PatternFill("solid", fgColor=BAND_BLUE)
                ws.cell(current, column).font = Font(name=FONT, size=10, bold=True)
        elif current in customer_set:
            for column in range(1, last_column + 1):
                ws.cell(current, column).font = Font(name=FONT, size=10)
        else:
            for column in range(1, last_column + 1):
                ws.cell(current, column).font = Font(
                    name=FONT, size=9, italic=True, color=GREY
                )

    # KPI đầu trang, lấy thẳng từ dòng Grand Total.
    for column, label in BC_KPI_LABELS:
        label_cell = ws.cell(BC_KPI_LABEL_ROW, column, label)
        label_cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        label_cell.fill = PatternFill("solid", fgColor=GREY)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = GRID
    kpi_source = {2: "B", 4: "C", 6: "D", 8: "H", 10: "J"}
    for column, source_letter in kpi_source.items():
        cell = ws.cell(
            BC_KPI_VALUE_ROW, column, f"={source_letter}{grand_total_row}"
        )
        cell.font = Font(name=FONT, size=12, bold=True, color=NAVY)
        cell.border = GRID
        if column in (6, 10):
            percent(cell)
        else:
            money(cell)
    ws.row_dimensions[BC_KPI_VALUE_ROW].height = 22

    for index, width in enumerate(BC_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    # Dòng tổng nằm TRÊN dòng chi tiết nên summaryBelow phải tắt.
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.applyStyles = False
    ws.freeze_panes = f"A{BC_FIRST_DATA_ROW}"

    for column in (4, 6, 10):
        letter = get_column_letter(column)
        area_ref = f"{letter}{BC_FIRST_DATA_ROW}:{letter}{grand_total_row}"
        ws.conditional_formatting.add(area_ref, CellIsRule(
            operator="lessThan", formula=["0.8"], font=Font(color="FF9C0006"),
        ))
        ws.conditional_formatting.add(area_ref, CellIsRule(
            operator="greaterThanOrEqual", formula=["1"], font=Font(color="FF006100"),
        ))

    return {
        "sheet": ws.title,
        "area": area,
        "region_rows": len(region_rows),
        "customer_rows": len(customer_rows),
        "product_rows": len(product_rows),
        "grand_total_row": grand_total_row,
    }


def build_tree(model_rows: list[list]) -> dict[str, dict]:
    """Gom dòng PVT_DATA thành cây Miền -> Vùng -> Khách hàng -> Sản phẩm.

    Khách hàng gom theo MÃ, không theo tên hiển thị. Cùng một mã có thể xuất
    hiện với hai cách viết tên (một từ Sell In, một từ Target); nếu tách thành
    hai dòng thì dòng khách hàng dùng SUMIFS theo mã sẽ đếm gộp cả hai trong
    khi các dòng sản phẩm con chỉ có một phần.
    """
    tree: dict[str, dict] = {}
    for row in model_rows:
        area, region, code, customer, product = row[0], row[1], row[2], row[3], row[4]
        values = {
            "Actual": row[5],
            "CungKyLY": row[6],
            "ThangTruoc": row[7],
            "Vikoda": row[8],
            "TargetTong": row[9],
            "TargetVikoda": row[10],
        }
        regions = tree.setdefault(area, {})
        customers = regions.setdefault(region, {})
        entry = customers.setdefault(code, {"labels": set(), "products": {}})
        entry["labels"].add(customer)
        current = entry["products"].setdefault(product, dict.fromkeys(values, 0.0))
        for field, value in values.items():
            current[field] += value
    return tree


def customer_display(labels: set[str]) -> str:
    """Tên hiển thị của khách hàng: chọn bản đầy đủ nhất trong các cách viết."""
    known = [label for label in labels if "Không rõ tên" not in label]
    return sorted(known or list(labels), key=lambda text: (-len(text), text))[0]


def order_tree(tree: dict[str, dict]) -> dict[str, dict]:
    """Sắp xếp theo thứ tự miền/vùng của báo cáo, khách hàng theo tên."""
    ordered: dict[str, dict] = {}
    for area, regions in REPORTING_STRUCTURE:
        area_tree = tree.get(area, {})
        ordered_regions: dict[str, dict] = {}
        for region in regions:
            customers = area_tree.get(region, {})
            resolved = {
                (code, customer_display(entry["labels"])): dict(
                    sorted(entry["products"].items())
                )
                for code, entry in customers.items()
            }
            ordered_regions[region] = dict(
                sorted(resolved.items(), key=lambda item: item[0][1])
            )
        ordered[area] = ordered_regions
    return ordered


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--sell-in-data-file", required=True)
    parser.add_argument("--target-data-file", required=True)
    parser.add_argument("--dmkh-data-file", required=True)
    parser.add_argument("--report-file", required=True)
    return parser.parse_args()


def load_json(path: str) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def main() -> None:
    args = parse_args()
    sell_in_payload = load_json(args.sell_in_data_file)
    target_payload = load_json(args.target_data_file)
    dmkh_payload = load_json(args.dmkh_data_file)

    model = build_model(sell_in_payload, target_payload, dmkh_payload)
    model_rows = model.rows()

    workbook_path = Path(args.workbook).resolve()
    workbook = load_workbook(workbook_path)

    for name in list(workbook.sheetnames):
        if name == "PIVOT" or name == "PVT_DATA" or name.startswith("BC_"):
            del workbook[name]

    pivot_ws = workbook.create_sheet("PIVOT")
    pvt_ws = workbook.create_sheet("PVT_DATA")
    pvt_last_row = build_pvt_data_sheet(pvt_ws, model_rows)

    period_label = f"{model.current_month:02d}/{model.current_year}"
    as_of_date = str(sell_in_payload.get("as_of_date") or "")
    pivot_info = build_pivot_sheet(
        pivot_ws, pvt_last_row, period_label, as_of_date
    )

    tree = order_tree(build_tree(model_rows))
    bc_reports = []
    for area, _ in REPORTING_STRUCTURE:
        ws = workbook.create_sheet(sheet_title_for(area))
        bc_reports.append(
            build_bc_sheet(
                ws, area, tree[area], pvt_last_row, period_label, as_of_date
            )
        )

    pvt_ws.sheet_state = "hidden"
    order = ["Target", "Data", "DMKH", "PIVOT", "PVT_DATA"] + [
        report["sheet"] for report in bc_reports
    ]
    workbook._sheets = sorted(
        workbook._sheets,
        key=lambda sheet: order.index(sheet.title) if sheet.title in order else 99,
    )
    workbook.save(workbook_path)

    report = {
        "workbook": str(workbook_path),
        "period": period_label,
        "as_of_date": as_of_date,
        "pvt_rows": len(model_rows),
        "pvt_last_row": pvt_last_row,
        "pivot": pivot_info,
        "bc_sheets": bc_reports,
        "mapping_stats": model.mapping_stats,
    }
    Path(args.report_file).parent.mkdir(parents=True, exist_ok=True)
    Path(args.report_file).write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(str(args.report_file))


if __name__ == "__main__":
    main()
