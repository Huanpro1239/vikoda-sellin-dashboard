from __future__ import annotations

import argparse
import json
import os
import posixpath
import xml.etree.ElementTree as ET
from copy import copy
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.table import Table


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)


def copy_worksheet(source, destination) -> None:
    for row in source.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            destination_cell = destination[source_cell.coordinate]
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.border = copy(source_cell.border)
                destination_cell.alignment = copy(source_cell.alignment)
                destination_cell.number_format = source_cell.number_format
                destination_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                destination_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                destination_cell.comment = copy(source_cell.comment)

    for key, source_dimension in source.column_dimensions.items():
        destination_dimension = destination.column_dimensions[key]
        destination_dimension.width = source_dimension.width
        destination_dimension.hidden = source_dimension.hidden
        destination_dimension.bestFit = source_dimension.bestFit
        destination_dimension.outlineLevel = source_dimension.outlineLevel
        destination_dimension.collapsed = source_dimension.collapsed
        destination_dimension.min = source_dimension.min
        destination_dimension.max = source_dimension.max

    for index, source_dimension in source.row_dimensions.items():
        destination_dimension = destination.row_dimensions[index]
        destination_dimension.height = source_dimension.height
        destination_dimension.hidden = source_dimension.hidden
        destination_dimension.outlineLevel = source_dimension.outlineLevel
        destination_dimension.collapsed = source_dimension.collapsed

    for merged_range in source.merged_cells.ranges:
        destination.merge_cells(str(merged_range))

    for conditional_format in source.conditional_formatting:
        for rule in conditional_format.rules:
            destination.conditional_formatting.add(
                str(conditional_format.sqref),
                copy(rule),
            )

    for table in source.tables.values():
        destination_table = Table(
            displayName=table.displayName,
            ref=table.ref,
        )
        destination_table.tableStyleInfo = copy(table.tableStyleInfo)
        destination.add_table(destination_table)

    destination.freeze_panes = source.freeze_panes
    destination.sheet_view.showGridLines = source.sheet_view.showGridLines
    destination.sheet_format = copy(source.sheet_format)
    destination.sheet_properties = copy(source.sheet_properties)
    destination.page_margins = copy(source.page_margins)
    destination.page_setup = copy(source.page_setup)
    destination.print_options = copy(source.print_options)


def workbook_sheet_path(archive: ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    relationship_id = None
    sheets = workbook_root.find(f"{{{MAIN_NS}}}sheets")
    if sheets is None:
        raise ValueError("Workbook khong co danh sach sheet")
    for sheet in sheets:
        if sheet.attrib.get("name") == sheet_name:
            relationship_id = sheet.attrib.get(f"{{{REL_NS}}}id")
            break
    if relationship_id is None:
        raise ValueError(f"Khong tim thay sheet {sheet_name} trong workbook.xml")

    relationships_root = ET.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    target = None
    for relationship in relationships_root:
        if relationship.attrib.get("Id") == relationship_id:
            target = relationship.attrib.get("Target")
            break
    if not target:
        raise ValueError(f"Khong tim thay relationship cua sheet {sheet_name}")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def numeric_xml_value(value) -> str:
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    raise TypeError(f"Gia tri cache cong thuc khong phai so: {value!r}")


def inject_formula_cache(
    workbook_path: Path,
    sheet_name: str,
    cached_values: dict[str, int | float],
) -> None:
    temporary_file = workbook_path.with_name(
        f"{workbook_path.stem}.cache-building{workbook_path.suffix}"
    )
    if temporary_file.exists():
        temporary_file.unlink()

    with ZipFile(workbook_path, "r") as source_archive:
        sheet_path = workbook_sheet_path(source_archive, sheet_name)
        sheet_root = ET.fromstring(source_archive.read(sheet_path))
        updated = 0
        for cell in sheet_root.iter(f"{{{MAIN_NS}}}c"):
            coordinate = cell.attrib.get("r")
            formula = cell.find(f"{{{MAIN_NS}}}f")
            if formula is None or coordinate not in cached_values:
                continue
            value_element = cell.find(f"{{{MAIN_NS}}}v")
            if value_element is None:
                value_element = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
            value_element.text = numeric_xml_value(cached_values[coordinate])
            cell.attrib.pop("t", None)
            updated += 1

        if updated != len(cached_values):
            raise ValueError(
                f"Chi chen duoc {updated}/{len(cached_values)} "
                f"gia tri cache vao {sheet_name}"
            )

        updated_sheet_xml = ET.tostring(
            sheet_root,
            encoding="utf-8",
            xml_declaration=True,
        )
        with ZipFile(
            temporary_file,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=6,
        ) as destination_archive:
            for item in source_archive.infolist():
                data = (
                    updated_sheet_xml
                    if item.filename == sheet_path
                    else source_archive.read(item.filename)
                )
                destination_archive.writestr(item, data)

    os.replace(temporary_file, workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ghep PIVOT va PVT_DATA vao Bao_Cao_Sell_in.xlsx"
    )
    parser.add_argument("--base-workbook", required=True)
    parser.add_argument("--pivot-workbook", required=True)
    args = parser.parse_args()

    base_path = Path(args.base_workbook).resolve()
    pivot_path = Path(args.pivot_workbook).resolve()
    if not base_path.exists() or not pivot_path.exists():
        raise FileNotFoundError("Thieu workbook nen hoac workbook PIVOT")

    source = load_workbook(pivot_path, data_only=False)
    source_values = load_workbook(pivot_path, data_only=True)
    destination = load_workbook(base_path, data_only=False)
    try:
        cached_values = {
            cell.coordinate: cell.value
            for row in source_values["PIVOT"].iter_rows()
            for cell in row
            if (
                source["PIVOT"][cell.coordinate].data_type == "f"
                and isinstance(cell.value, (int, float))
            )
        }
        expected_formula_count = sum(
            1
            for row in source["PIVOT"].iter_rows()
            for cell in row
            if cell.data_type == "f"
        )
        if len(cached_values) != expected_formula_count:
            raise ValueError(
                f"Workbook PIVOT chi co {len(cached_values)}/"
                f"{expected_formula_count} gia tri cache cong thuc"
            )

        for sheet_name in ("PIVOT", "PVT_DATA"):
            if sheet_name in destination.sheetnames:
                destination.remove(destination[sheet_name])
            target_sheet = destination.create_sheet(sheet_name)
            copy_worksheet(source[sheet_name], target_sheet)

        destination["PIVOT"].sheet_state = "visible"
        destination["PVT_DATA"].sheet_state = "hidden"
        destination["PIVOT"].freeze_panes = "C5"
        destination["PVT_DATA"].freeze_panes = "A2"
        destination.active = destination.sheetnames.index("Target")
        destination.calculation.calcMode = "auto"
        destination.calculation.fullCalcOnLoad = True
        destination.calculation.forceFullCalc = True

        temporary_file = base_path.with_name(
            f"{base_path.stem}.merge-building{base_path.suffix}"
        )
        if temporary_file.exists():
            temporary_file.unlink()
        destination.save(temporary_file)
        os.replace(temporary_file, base_path)
    finally:
        destination.close()
        source_values.close()
        source.close()

    inject_formula_cache(base_path, "PIVOT", cached_values)
    print(
        json.dumps(
            {
                "base_workbook": str(base_path),
                "pivot_workbook": str(pivot_path),
                "cached_formula_values": len(cached_values),
            }
        )
    )


if __name__ == "__main__":
    main()
