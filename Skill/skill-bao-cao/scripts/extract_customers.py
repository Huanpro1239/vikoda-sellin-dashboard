from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_COLUMNS = [
    "MaKhachHangMoi",
    "TenKhachHang",
    "TenKhachHangdaydu",
    "DiaChi",
    "KenhBanHang",
    "Loaikhachhang",
    "Hethong MT",
    "MIEN",
    "VUNG",
    "TINHTHANH",
    "QUANHUYEN",
    "CODE",
    "TENKH",
]

OUTPUT_COLUMNS = SOURCE_COLUMNS + ["ThongTinBoSungNguon"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def select_source_file(source_dir: Path) -> Path:
    preferred = source_dir / "Thong tin khach hang.xlsx"
    if preferred.exists():
        return preferred
    candidates = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if len(candidates) != 1:
        raise ValueError(
            "Can dung mot file .xlsx trong thu muc Danh muc KH; "
            f"tim thay {len(candidates)} file"
        )
    return candidates[0]


def main() -> int:
    parser = argparse.ArgumentParser(description="Chuan hoa danh muc khach hang")
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--staging-dir", required=True)
    args = parser.parse_args()

    source_dir = Path(args.source_dir).resolve()
    staging_dir = Path(args.staging_dir).resolve()
    staging_dir.mkdir(parents=True, exist_ok=True)
    data_file = staging_dir / "dmkh_data.json"
    audit_file = staging_dir / "dmkh_audit.json"

    problems: list[str] = []
    warnings: list[str] = []
    rows: list[list[str]] = []
    source_rows: list[int] = []
    source_file: Path | None = None

    try:
        source_file = select_source_file(source_dir)
    except ValueError as error:
        problems.append(str(error))

    if source_file is not None:
        workbook = load_workbook(source_file, read_only=True, data_only=False)
        try:
            if "DM KH" not in workbook.sheetnames:
                problems.append(f"{source_file.name}: khong co sheet DM KH")
            else:
                worksheet = workbook["DM KH"]
                headers = [
                    worksheet.cell(1, column).value
                    for column in range(1, len(OUTPUT_COLUMNS) + 1)
                ]
                if headers[: len(SOURCE_COLUMNS)] != SOURCE_COLUMNS:
                    problems.append(
                        f"{source_file.name}: 13 cot dau khong dung: "
                        f"{headers[:len(SOURCE_COLUMNS)]}"
                    )
                if headers[13] not in (None, ""):
                    problems.append(
                        f"{source_file.name}: cot N da co tieu de {headers[13]!r}; "
                        "can cap nhat quy tac truoc khi chay"
                    )

                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        min_col=1,
                        max_col=len(OUTPUT_COLUMNS),
                        values_only=True,
                    ),
                    start=2,
                ):
                    if not any(value is not None for value in values):
                        continue
                    if any(
                        isinstance(value, str) and value.startswith("=")
                        for value in values
                    ):
                        problems.append(
                            f"{source_file.name}!R{row_number}: khong duoc co cong thuc"
                        )
                        continue
                    rows.append([clean_text(value) for value in values])
                    source_rows.append(row_number)
        finally:
            workbook.close()

    codes: dict[str, list[int]] = defaultdict(list)
    extra_source_values = 0
    for source_row, row in zip(source_rows, rows):
        code = row[0]
        codes[code].append(source_row)
        if not code:
            warnings.append(
                f"Dong nguon {source_row}: MaKhachHangMoi trong, van giu dong"
            )
        if row[13]:
            extra_source_values += 1

    duplicate_codes = {
        code: row_numbers
        for code, row_numbers in codes.items()
        if code and len(row_numbers) > 1
    }
    for code, row_numbers in duplicate_codes.items():
        warnings.append(
            f"MaKhachHangMoi {code} lap tai dong nguon "
            + ", ".join(str(row_number) for row_number in row_numbers)
            + "; khong tu xoa"
        )
    if extra_source_values:
        warnings.append(
            f"Cot N khong co tieu de co {extra_source_values} gia tri; "
            "luu vao ThongTinBoSungNguon"
        )

    generated_at = datetime.now().astimezone().isoformat()
    payload = {
        "schema_version": 1,
        "generated_at": generated_at,
        "columns": OUTPUT_COLUMNS,
        "rows": rows,
    }
    audit = {
        "generated_at": generated_at,
        "source_dir": str(source_dir),
        "source_file": str(source_file) if source_file else None,
        "source_sha256": sha256_file(source_file) if source_file else None,
        "data_file": str(data_file),
        "rows": len(rows),
        "unique_nonblank_customer_codes": len(
            [code for code in codes if code]
        ),
        "blank_customer_code_rows": len(codes.get("", [])),
        "duplicate_customer_codes": duplicate_codes,
        "extra_source_values": extra_source_values,
        "warnings": warnings,
        "problems": problems,
    }
    data_file.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    audit_file.write_text(
        json.dumps(audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "data_file": str(data_file),
                "audit_file": str(audit_file),
                "records": len(rows),
                "warnings": len(warnings),
                "problems": len(problems),
            },
            ensure_ascii=False,
        )
    )
    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
